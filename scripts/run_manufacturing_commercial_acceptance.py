from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import excel_data_toolbox.server as server_module  # noqa: E402
from excel_data_toolbox.core import export_tables  # noqa: E402
from excel_data_toolbox.enterprise_report import build_enterprise_diagnosis_report  # noqa: E402
from excel_data_toolbox.io_utils import load_tables_from_files  # noqa: E402
from excel_data_toolbox.server import AppSession, ToolboxHandler  # noqa: E402

SOURCE = Path(r"D:\Users\liuyisheng\终极测试_制造企业经营诊断案例.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parents[1] / "outputs" / "enterprise-commercial-refactor-20260826"
OUTPUT = OUTPUT_DIR / "制造企业经营诊断报告_商业验收版.xlsx"
REPORT = OUTPUT_DIR / "商业验收结果.json"


def main() -> None:
    tables = load_tables_from_files(SOURCE)
    prompt = "公司最近经营出现问题，请帮我全面分析企业经营情况，找出主要风险，并给出下一步行动建议。"
    session = AppSession()
    handler = object.__new__(ToolboxHandler)
    try:
        table_ids = [
            session.add_table(name, frame, source=str(SOURCE), original=True) for name, frame in tables.items()
        ]
        with (
            patch.object(server_module, "SESSION", session),
            patch.object(
                server_module,
                "_project_ai_config",
                return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"},
            ),
            patch.object(
                server_module.DeepSeekClient,
                "classify_unified_request",
                side_effect=AssertionError("企业诊断应由本地确定性路由直接识别"),
            ),
        ):
            routed = handler._ai_unified({"prompt": prompt, "table_ids": table_ids})
    finally:
        session.close()
    result = build_enterprise_diagnosis_report(
        list(tables.values()),
        source_names=list(tables),
        user_request=prompt,
    )
    expected = {
        "net_revenue": 2_155_000.0,
        "management_revenue_excluding_refunds": 2_035_000.0,
        "gross_profit": 510_000.0,
        "operating_expense": 740_000.0,
        "estimated_operating_profit": -230_000.0,
        "collection_amount": 950_000.0,
        "collection_rate": 950_000 / 2_155_000,
        "collection_risk_exposure": 670_000.0,
        "inventory_value": 820_000.0,
        "production_cost": 1_420_000.0,
    }
    checks = [
        {
            "check": "natural_language_enterprise_route",
            "passed": routed["plan"]["steps"][0]["operation"] == "enterprise_diagnosis_report",
            "actual": routed["plan"]["steps"][0]["operation"],
            "expected": "enterprise_diagnosis_report",
        }
    ]
    for key, expected_value in expected.items():
        actual = result.report[key]
        passed = actual is not None and math.isclose(float(actual), expected_value, rel_tol=1e-9, abs_tol=0.01)
        checks.append({"check": key, "passed": passed, "actual": actual, "expected": expected_value})
        if not passed:
            raise AssertionError(f"{key}: actual={actual!r}, expected={expected_value!r}")

    customer = result.outputs["客户与回款风险"].set_index("客户")
    overseas = customer.loc["海外客户A"]
    checks.extend(
        [
            {
                "check": "overseas_customer_share",
                "passed": math.isclose(float(overseas["收入占比"]), 1_100_000 / 2_035_000, rel_tol=1e-9),
                "actual": overseas["收入占比"],
                "expected": 1_100_000 / 2_035_000,
            },
            {
                "check": "overseas_customer_risk",
                "passed": overseas["风险等级"] == "高",
                "actual": overseas["风险等级"],
                "expected": "高",
            },
        ]
    )
    checks.extend(
        [
            {
                "check": "refund_customer_reason",
                "passed": customer.loc["华东汽车", "综合风险"] == "中"
                and "退款订单120,000元需复核" in customer.loc["华东汽车", "主要风险"],
                "actual": customer.loc["华东汽车", ["综合风险", "主要风险"]].to_dict(),
                "expected": "综合风险中且明确退款原因",
            },
            {
                "check": "source_risk_not_downgraded",
                "passed": customer.loc["北方机械", "综合风险"] == "中",
                "actual": customer.loc["北方机械", "综合风险"],
                "expected": "中",
            },
            {
                "check": "unmatched_customer_not_low_risk",
                "passed": customer.loc["华北机械", "综合风险"] == "未知/待核验",
                "actual": customer.loc["华北机械", "综合风险"],
                "expected": "未知/待核验",
            },
        ]
    )
    sales = result.outputs["销售团队诊断"]
    checks.append(
        {
            "check": "missing_order_cost_stays_unknown",
            "passed": sales["订单级成本"].isna().all() and sales["毛利口径"].eq("绩效口径").all(),
            "actual": sales[["负责人", "订单级成本", "毛利口径"]].to_dict("records"),
            "expected": "成本为空且参考毛利标记为绩效口径",
        }
    )
    checks.append(
        {
            "check": "sales_quality_action",
            "passed": result.outputs["风险行动计划"]["风险事项"].eq("销售质量与客户体验").any(),
            "actual": result.outputs["风险行动计划"]["风险事项"].tolist(),
            "expected": "销售质量与客户体验",
        }
    )
    audit = result.outputs["数据口径与验收"]
    checks.append(
        {
            "check": "customer_directional_coverage",
            "passed": audit["数据证据/口径"].astype(str).str.contains("88.9%.*83.3%", regex=True).any(),
            "actual": audit["数据证据/口径"].tolist(),
            "expected": "行88.9%/唯一客户83.3%",
        }
    )
    checks.append(
        {
            "check": "audit_open_items_reconcile",
            "passed": result.report["open_definition_count"]
            == int(audit["状态"].astype(str).str.startswith("待").sum())
            == 7,
            "actual": result.report["open_definition_count"],
            "expected": 7,
        }
    )
    checks.append(
        {
            "check": "product_alias_suggestion",
            "passed": audit["审计项"].astype(str).str.contains("工业软件↔工业软件授权", regex=False).any(),
            "actual": audit["审计项"].tolist(),
            "expected": "工业软件↔工业软件授权",
        }
    )
    checks.append(
        {
            "check": "refund_boundary",
            "passed": audit["审计项"].eq("退款处理").any(),
            "actual": audit["审计项"].tolist(),
            "expected": "退款处理",
        }
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    export_tables(result.outputs, OUTPUT, include_log=False, overwrite=True)
    workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    try:
        expected_sheets = list(result.outputs)
        checks.append(
            {
                "check": "sheet_structure",
                "passed": workbook.sheetnames == expected_sheets,
                "actual": workbook.sheetnames,
                "expected": expected_sheets,
            }
        )
        chart_count = len(workbook["经营诊断看板"]._charts)
        checks.append({"check": "native_charts", "passed": chart_count == 4, "actual": chart_count, "expected": 4})
        dashboard = workbook["经营诊断看板"]
        chart_columns = [chart.anchor._from.col for chart in dashboard._charts]
        checks.append(
            {
                "check": "compact_dashboard_layout",
                "passed": dashboard["A4"].value == "销售规模"
                and "A1:N1" in {str(item) for item in dashboard.merged_cells.ranges}
                and max(chart_columns, default=99) <= 7,
                "actual": {
                    "first_card": dashboard["A4"].value,
                    "chart_anchor_columns": chart_columns,
                    "print_area": str(dashboard.print_area),
                },
                "expected": "A:N内的KPI、诊断、风险卡和2×2图表",
            }
        )
        hidden_rows = sum(bool(dimension.hidden) for dimension in workbook["经营诊断看板"].row_dimensions.values())
        checks.append(
            {"check": "chart_helper_rows_hidden", "passed": hidden_rows > 0, "actual": hidden_rows, "expected": ">0"}
        )
    finally:
        workbook.close()

    failed = [check for check in checks if not check["passed"]]
    payload = {
        "status": "passed" if not failed else "failed",
        "source": str(SOURCE),
        "output": str(OUTPUT),
        "report": dict(result.report),
        "checks": checks,
    }
    REPORT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    if failed:
        raise AssertionError(f"商业验收失败：{[item['check'] for item in failed]}")
    print(json.dumps({"status": "passed", "output": str(OUTPUT), "checks": len(checks)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
