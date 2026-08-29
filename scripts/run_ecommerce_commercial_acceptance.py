from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import excel_data_toolbox.server as server_module  # noqa: E402
from excel_data_toolbox.core import export_tables  # noqa: E402
from excel_data_toolbox.ecommerce_report import build_ecommerce_diagnosis_report  # noqa: E402
from excel_data_toolbox.io_utils import load_tables_from_files  # noqa: E402
from excel_data_toolbox.server import AppSession, ToolboxHandler  # noqa: E402

SOURCE = Path(r"D:\Users\liuyisheng\真实项目_多平台电商经营诊断终极测试.xlsx")
OUTPUT_DIR = PROJECT_ROOT / "excel_data_toolbox" / "outputs" / "ecommerce-commercial-refactor-20260828"
OUTPUT = OUTPUT_DIR / "多平台电商经营诊断报告_商业验收版.xlsx"
REPORT = OUTPUT_DIR / "商业验收结果.json"


def main() -> None:
    tables = load_tables_from_files(SOURCE)
    prompt = (
        "公司销售和订单增长，但利润、现金和库存恶化。请完整关联订单、退款、平台结算、广告、采购、"
        "库存和客户，找出风险并生成老板能直接看的经营诊断Excel。"
    )
    session = AppSession()
    handler = object.__new__(ToolboxHandler)
    try:
        table_ids = [session.add_table(name, frame, source=str(SOURCE), original=True) for name, frame in tables.items()]
        with (
            patch.object(server_module, "SESSION", session),
            patch.object(server_module, "_project_ai_config", return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"}),
            patch.object(server_module.DeepSeekClient, "classify_unified_request", side_effect=AssertionError("应使用本地确定性插件")),
        ):
            routed = handler._ai_unified({"prompt": prompt, "table_ids": table_ids})
    finally:
        session.close()

    result = build_ecommerce_diagnosis_report(list(tables.values()), source_names=list(tables), user_request=prompt)
    checks: list[dict[str, object]] = []

    def record(name: str, passed: bool, actual: object, expected: object) -> None:
        checks.append({"check": name, "passed": bool(passed), "actual": actual, "expected": expected})

    record("natural_language_route", routed["plan"]["steps"][0]["operation"] == "enterprise_diagnosis_report", routed["plan"]["steps"][0]["operation"], "enterprise_diagnosis_report")
    record("domain_plan_message", "多平台电商" in routed["plan"]["message"], routed["plan"]["message"], "多平台电商事实域插件")
    expected_metrics = {
        "valid_order_count": 32,
        "valid_line_count": 34,
        "buyer_paid": 26821,
        "refund_amount": 2478,
        "net_management_sales": 24343,
        "standard_cost": 16322,
        "product_gross_profit": 8021,
        "platform_fees": 2507,
        "actual_arrival": 21836,
        "ad_spend": 10440,
        "management_contribution": -4926,
        "latest_inventory_value": 167104,
        "pending_refund_amount": 180,
    }
    for key, expected in expected_metrics.items():
        actual = result.report[key]
        record(key, math.isclose(float(actual), float(expected), abs_tol=0.01), actual, expected)
    record("roas", math.isclose(result.report["roas"], 2.132183908, rel_tol=1e-8), result.report["roas"], 2.132183908)
    record("inventory_growth", math.isclose(result.report["inventory_growth"], 0.605256585, rel_tol=1e-8), result.report["inventory_growth"], 0.605256585)

    channel = result.outputs["渠道与广告诊断"].set_index("渠道")
    record("douyin_contribution", channel.loc["抖音", "管理贡献"] == -6384, channel.loc["抖音", "管理贡献"], -6384)
    record("douyin_roas", math.isclose(channel.loc["抖音", "ROAS"], 1.284210526, rel_tol=1e-8), channel.loc["抖音", "ROAS"], 1.284210526)
    product = result.outputs["商品利润质量"].set_index("SKU")
    record("kit_negative_margin", product.loc["KIT01", "退款后商品毛利"] == -242, product.loc["KIT01", "退款后商品毛利"], -242)
    customer = result.outputs["客户与回款风险"].set_index("客户ID")
    record("c025_high_risk", customer.loc["C025", "风险优先级"] == "P0" and math.isclose(customer.loc["C025", "退款率"], 0.5), customer.loc["C025", ["风险优先级", "退款率"]].to_dict(), "P0 / 50%")
    audit = result.outputs["数据口径与验收"]
    relation_names = audit.loc[audit["审计类型"].eq("关系证据"), "审计项"].astype(str)
    record("business_key_relationship", relation_names.str.contains("原订单号.*订单号", regex=True).any(), relation_names.tolist(), "原订单号→订单号")
    record("no_amount_join", not relation_names.str.contains("退款金额.*退款金额", regex=True).any(), relation_names.tolist(), "禁止退款金额同名关联")
    record("settlement_reconciliation", result.outputs["平台费用与回款"]["勾稽差额"].abs().max() < 0.01, result.outputs["平台费用与回款"]["勾稽差额"].abs().max(), 0)
    record("business_risk_actions", len(result.outputs["风险行动计划"]) >= 5, result.outputs["风险行动计划"]["风险事项"].tolist(), ">=5 evidence-backed actions")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    export_tables(result.outputs, OUTPUT, include_log=False, overwrite=True)
    workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    try:
        record("sheet_contract", workbook.sheetnames == list(result.outputs), workbook.sheetnames, list(result.outputs))
        record("native_charts", len(workbook["经营诊断看板"]._charts) == 4, len(workbook["经营诊断看板"]._charts), 4)
        record("dashboard_title", workbook["经营诊断看板"]["A1"].value == "多平台电商经营驾驶舱", workbook["经营诊断看板"]["A1"].value, "多平台电商经营驾驶舱")
        record("summary_title", workbook["管理层诊断总览"]["A1"].value == "多平台电商经营诊断驾驶舱", workbook["管理层诊断总览"]["A1"].value, "多平台电商经营诊断驾驶舱")
        chart_anchors = [chart.anchor._from.col for chart in workbook["经营诊断看板"]._charts]
        record("compact_dashboard", max(chart_anchors, default=99) <= 7, chart_anchors, "all charts anchored in A:N canvas")
        record("helper_rows_hidden", any(bool(row.hidden) for row in workbook["经营诊断看板"].row_dimensions.values()), "hidden", True)
    finally:
        workbook.close()

    failed = [item for item in checks if not item["passed"]]
    payload = {"status": "passed" if not failed else "failed", "source": str(SOURCE), "output": str(OUTPUT), "report": dict(result.report), "checks": checks}
    REPORT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    if failed:
        raise AssertionError(f"商业验收失败：{[item['check'] for item in failed]}")
    print(json.dumps({"status": "passed", "output": str(OUTPUT), "checks": len(checks)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
