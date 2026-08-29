"""Run production-style V9 acceptance checks and keep auditable artifacts."""

from __future__ import annotations

from datetime import datetime, timedelta
import json
from pathlib import Path
import sqlite3
import sys
import traceback
from typing import Callable

import pandas as pd
from openpyxl import Workbook


PACKAGE = Path(__file__).resolve().parents[1]
ROOT = PACKAGE.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_data_toolbox.advanced_automation import (  # noqa: E402
    build_vba_bundle,
    document_capabilities,
    extract_image_text,
    query_sqlite_read_only,
)
from excel_data_toolbox.core import export_tables  # noqa: E402
from excel_data_toolbox.delivery_qa import verify_delivery, write_acceptance_json  # noqa: E402
from excel_data_toolbox.large_data import duckdb_available, query_files  # noqa: E402
from excel_data_toolbox.nl_agent import build_table_catalog, normalize_plan_envelope, validate_plan  # noqa: E402
from excel_data_toolbox.order_intake import quote_order  # noqa: E402
from excel_data_toolbox.scheduler import LocalScheduler  # noqa: E402
from excel_data_toolbox.secure_secrets import SecureSecretStore  # noqa: E402
from excel_data_toolbox.task_store import TaskRepository  # noqa: E402
from excel_data_toolbox.workbook_fidelity import (  # noqa: E402
    preserve_workbook_export,
    workbook_feature_inventory,
)


def production_sales() -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    regions = ["华东", "华南", "华北", "西南"]
    channels = ["线上商城", "企业直销", "线下门店"]
    products = ["云影显示器", "轻享办公椅", "远山升降桌"]
    for index in range(1, 241):
        month = (index - 1) % 12 + 1
        amount = float(1200 + month * 73 + (index % 17) * 215)
        rows.append(
            {
                "订单编号": f"ACCEPT-{index:05d}",
                "日期": pd.Timestamp(2025, month, (index % 27) + 1),
                "地区": regions[index % len(regions)],
                "渠道": channels[index % len(channels)],
                "产品": products[index % len(products)],
                "销售额": amount,
                "成本": round(amount * (0.55 + (index % 5) * 0.025), 2),
                "数量": index % 6 + 1,
            }
        )
    frame = pd.DataFrame(rows)
    frame["利润"] = frame["销售额"] - frame["成本"]
    return frame


def main() -> int:
    run_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    output = PACKAGE / "outputs" / "v9_acceptance" / run_id
    output.mkdir(parents=True, exist_ok=False)
    checks: list[dict[str, object]] = []

    def check(name: str, function: Callable[[], object]) -> None:
        started = datetime.now()
        try:
            detail = function()
            checks.append(
                {
                    "name": name,
                    "status": "passed",
                    "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
                    "detail": detail,
                }
            )
        except Exception as exc:  # keep running so the report lists every failure
            checks.append(
                {
                    "name": name,
                    "status": "failed",
                    "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
                    "error": f"{type(exc).__name__}: {exc}",
                    "traceback": traceback.format_exc(limit=5),
                }
            )

    sales = production_sales()

    def ai_schema_check() -> dict[str, object]:
        catalog = build_table_catalog({"sales": sales})
        raw = {
            "version": "v1",
            "status": "ready",
            "normalized_request": "按地区汇总销售额并形成经营数据",
            "description": "兼容模型常见字段别名",
            "questions": [],
            "notes": [],
            "assumption_list": [],
            "operations": [{
                "step_id": "step_1",
                "op": "summary",
                "inputs": "sales",
                "output": "月度销售",
                "parameters": {
                    "group_by": "地区",
                    "value_column": "销售额",
                    "aggregation": "sum",
                },
            }],
        }
        plan = validate_plan(normalize_plan_envelope(raw), catalog)
        if not plan.executable or not plan.steps:
            raise AssertionError(plan.message)
        return {"executable": plan.executable, "aggregation": plan.steps[0].params["aggregations"]}

    check("AI JSON 兼容归一化与本地安全校验", ai_schema_check)

    def delivery_check() -> dict[str, object]:
        monthly = (
            sales.assign(月份=sales["日期"].dt.to_period("M").astype(str))
            .groupby("月份", as_index=False)
            .agg(销售额=("销售额", "sum"), 利润=("利润", "sum"))
        )
        target = output / "生产销售交付包.xlsx"
        expected = {"销售明细": sales, "月度经营趋势": monthly}
        export_tables(expected, target, include_log=False)
        report = verify_delivery(target, expected)
        write_acceptance_json(report, output / "生产销售交付包_自动验收.json")
        if report.status != "passed":
            raise AssertionError(report.errors)
        return {
            "artifact": target.name,
            "checks": f"{report.checks_passed}/{report.checks_total}",
            "sha256": report.artifact_sha256,
        }

    check("Excel 导出后重开、指纹与数值合计验收", delivery_check)

    def fidelity_check() -> dict[str, object]:
        source = output / "客户复杂原表.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "经营底稿"
        sheet.append(["指标", "值"])
        sheet.append(["销售额", 100])
        sheet["B3"] = "=B2*0.3"
        sheet.merge_cells("A5:B5")
        hidden = workbook.create_sheet("隐藏参数")
        hidden.sheet_state = "hidden"
        workbook.save(source)
        target = output / "客户复杂原表_保真交付.xlsx"
        preserve_workbook_export(source, target, {"AI分析结果": sales.head(20)})
        inventory = workbook_feature_inventory(target)
        if "隐藏参数" not in inventory["hidden_sheets"] or inventory["merged_ranges"] < 1:
            raise AssertionError(inventory)
        return inventory

    check("原工作簿公式、隐藏表、合并单元格保真导出", fidelity_check)

    def task_check() -> dict[str, object]:
        repository = TaskRepository(output / "task_repository", retention_days=30)
        task_id = "20260823-V9QA"
        repository.create(task_id, "生产验收任务")
        repository.save(
            task_id,
            task_name="生产验收任务",
            tables={"sales": ("销售明细", sales, "自动验收", True)},
            active_table="sales",
            operations=[{"action": "acceptance"}],
            file_names=["生产销售交付包.xlsx"],
            import_warnings=[],
        )
        restored = repository.load(task_id)
        restored_frame = restored["loaded_tables"]["sales"][1]
        pd.testing.assert_frame_equal(restored_frame, sales)
        return {"task_id": task_id, "rows": len(restored_frame), "folders": len(repository.list_tasks())}

    check("一单一目录持久化与恢复", task_check)

    def database_check() -> dict[str, object]:
        database = output / "sales.sqlite"
        connection = sqlite3.connect(database)
        sales[["地区", "销售额", "利润"]].to_sql("sales", connection, index=False)
        connection.close()
        result = query_sqlite_read_only(
            database,
            "SELECT 地区, ROUND(SUM(销售额),2) 销售额 FROM sales GROUP BY 地区 ORDER BY 销售额 DESC",
        )
        result.to_excel(output / "数据库只读查询结果.xlsx", index=False)
        return {"rows": len(result), "leader": result.iloc[0]["地区"]}

    check("SQLite 外部数据库只读查询", database_check)

    def duckdb_check() -> dict[str, object]:
        if not duckdb_available():
            raise RuntimeError("DuckDB unavailable")
        source = output / "large_sales.csv"
        sales.to_csv(source, index=False)
        result = query_files(
            [source],
            "SELECT 地区, SUM(销售额) 销售额 FROM input_1 GROUP BY 地区 ORDER BY 销售额 DESC",
        )
        return {"rows": len(result), "total": round(float(result["销售额"].sum()), 2)}

    check("DuckDB 大 CSV 本地汇总", duckdb_check)

    def vba_check() -> dict[str, object]:
        code = (
            "Option Explicit\nPublic Sub FormatReport()\n"
            "  Worksheets(1).Rows(1).Font.Bold = True\nEnd Sub"
        )
        result = build_vba_bundle(code, output / "VBA安全交付包.zip")
        return {"status": result.status, "artifact": Path(result.artifacts[0]).name}

    check("VBA 静态危险指令扫描与交付包", vba_check)

    def ocr_check() -> dict[str, object]:
        if not document_capabilities().get("image_ocr"):
            raise RuntimeError("OCR engine or chi_sim/eng language packs unavailable")
        image = PACKAGE / "assets" / "xianyu" / "01_主封面_v2_数据分析.png"
        result = extract_image_text(image)
        result.to_excel(output / "OCR识别结果.xlsx", index=False)
        return {"text_blocks": len(result), "mean_confidence": round(float(result["conf"].mean()), 2)}

    check("真实中文图片 OCR", ocr_check)

    def scheduling_check() -> dict[str, object]:
        ran: list[dict[str, object]] = []
        scheduler = LocalScheduler(output / "schedules.sqlite3")
        scheduler.register_job("acceptance", lambda payload: ran.append(dict(payload)))
        schedule = scheduler.add("每周经营报表", "interval_minutes", "1", "acceptance", {"task": "sales"})
        count = scheduler.run_due(datetime.fromisoformat(schedule.next_run) + timedelta(seconds=1))
        if count != 1 or ran != [{"task": "sales"}]:
            raise AssertionError({"count": count, "ran": ran})
        return {"executed": count, "status": scheduler.list()[0].last_status}

    check("登记式本地定时任务", scheduling_check)

    def security_and_quote_check() -> dict[str, object]:
        store = SecureSecretStore(output / "test_secrets.dpapi")
        store.set("acceptance", "not-a-real-secret")
        if store.get("acceptance") != "not-a-real-secret":
            raise AssertionError("DPAPI roundtrip failed")
        store.delete("acceptance")
        quote = quote_order(
            "合并多渠道订单，完成复杂对账、经营看板、数据库读取和每周自动运行",
            table_count=5,
            total_rows=800_000,
            has_sample=True,
        )
        return {
            "dpapi": "passed",
            "capability": quote.capability,
            "complexity": quote.complexity_score,
            "suggested_price": list(quote.suggested_price),
        }

    check("Windows DPAPI 密钥保险箱与智能接单报价", security_and_quote_check)

    failed = [item for item in checks if item["status"] != "passed"]
    report = {
        "version": "9.0.0",
        "run_id": run_id,
        "status": "passed" if not failed else "failed",
        "checks_passed": len(checks) - len(failed),
        "checks_total": len(checks),
        "output_directory": str(output),
        "checks": checks,
    }
    report_path = output / "V9生产验收总报告.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps({key: report[key] for key in ("status", "checks_passed", "checks_total", "output_directory")}, ensure_ascii=False))
    return 0 if not failed else 1


if __name__ == "__main__":
    raise SystemExit(main())
