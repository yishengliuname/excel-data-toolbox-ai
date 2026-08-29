"""Run the real customer workbook through the five-sheet sales report engine."""

from __future__ import annotations

import json
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from excel_data_toolbox.core import export_tables
from excel_data_toolbox.nl_agent import build_table_catalog, execute_plan, validate_plan
from excel_data_toolbox.sales_report import infer_sales_report_columns


ROOT = Path(__file__).resolve().parents[2]
INPUT = Path(r"D:\Users\liuyisheng\Excel工具客户测试数据.xlsx")
OUTPUT_DIR = ROOT / "excel_data_toolbox" / "outputs" / "customer_sales_20260824"
OUTPUT_XLSX = OUTPUT_DIR / "销售经营分析报告_自动验收.xlsx"
OUTPUT_JSON = OUTPUT_DIR / "销售经营分析报告_自动验收.json"


def run() -> dict[str, object]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    frame = pd.read_excel(INPUT, sheet_name="销售数据原始表", dtype=object)
    params: dict[str, object] = {
        **infer_sales_report_columns(frame),
        "satisfaction_threshold": 4,
    }
    catalog = build_table_catalog({"customer_sales": frame}, display_names={"customer_sales": "销售数据原始表"})
    plan = validate_plan(
        {
            "schema_version": 1,
            "status": "ready",
            "summary": "生成五张工作表的销售经营管理报告",
            "message": "可执行",
            "clarification_questions": [],
            "assumptions": ["平均利润率采用总利润除以总销售额。"],
            "warnings": ["异常提醒需结合业务凭证复核。"],
            "steps": [
                {
                    "id": "sales_report_1",
                    "operation": "sales_management_report",
                    "input_ids": ["customer_sales"],
                    "output_name": "销售经营分析报告",
                    "params": params,
                }
            ],
        },
        catalog,
    )
    executed = execute_plan(plan, {"customer_sales": frame}, dry_run=False)
    export_tables(executed.tables, OUTPUT_XLSX, include_log=False, overwrite=True)

    report = dict(executed.reports["sales_report_1"])
    expected_names = ["管理层数据总览", "产品分析", "销售人员分析", "异常数据提醒", "图表展示"]
    assert list(executed.tables) == expected_names
    assert report["total_sales"] == 442500.0
    assert report["total_cost"] == 202000.0
    assert report["total_profit"] == 240500.0
    assert report["top_product_by_sales"] == "数据服务"
    assert report["top_product_by_profit"] == "数据服务"
    assert report["top_salesperson"] == "赵敏"
    assert report["attention_rows"] == 2

    workbook = load_workbook(OUTPUT_XLSX, data_only=False)
    try:
        assert workbook.sheetnames == expected_names
        assert len(workbook["图表展示"]._charts) == 3
        assert workbook["管理层数据总览"]["B5"].value == 442500.0
        assert workbook["管理层数据总览"]["B8"].number_format == "0.00%"
        assert workbook["图表展示"]["A40"].value == "月份"
        assert workbook["图表展示"].sheet_view.showGridLines is False
        assert workbook["产品分析"].column_dimensions["B"].width >= 17
        chart_titles = []
        for chart in workbook["图表展示"]._charts:
            title = ""
            try:
                title = chart.title.tx.rich.p[0].r[0].t
            except (AttributeError, IndexError, TypeError):
                pass
            chart_titles.append(title)
    finally:
        workbook.close()

    result = {
        "status": "passed",
        "input": str(INPUT),
        "output": str(OUTPUT_XLSX),
        "sheets": expected_names,
        "chart_count": 3,
        "chart_titles": chart_titles,
        "metrics": report,
    }
    OUTPUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


if __name__ == "__main__":
    print(json.dumps(run(), ensure_ascii=False, indent=2))
