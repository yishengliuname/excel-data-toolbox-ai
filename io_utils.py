"""File loading and exporting helpers for the Excel data toolbox."""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
import json
import math
import os
from pathlib import Path
import re
import tempfile
from typing import Any
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

from .models import ExportResult, OperationLog, OperationRecord


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
SUPPORTED_TEXT_SUFFIXES = {".csv", ".tsv"}
INVALID_EXCEL_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

_EMBEDDED_SALES_HEADER_GROUPS: tuple[tuple[str, ...], ...] = (
    ("订单编号", "流水号", "单号", "订单号"),
    ("下单日期", "日期", "业务日期", "交易日期"),
    ("产品类别", "产品", "品类"),
    ("销售区域", "地区", "区域"),
    ("业务员", "销售人员", "负责人", "销售员"),
    ("销售金额", "成交额", "含税销售额", "销售额"),
    ("成本", "采购成本", "采购/服务成本"),
    ("客户满意度", "满意度评分", "评分"),
    ("订单状态", "状态", "是否有效"),
)


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    try:
        if bool(pd.isna(value)):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(value, str) and not value.strip()


def _sales_header_score(values: Iterable[Any]) -> int:
    normalised = {re.sub(r"\s+", "", str(value)).casefold() for value in values if not _is_blank_cell(value)}
    return sum(
        any(re.sub(r"\s+", "", alias).casefold() in normalised for alias in group)
        for group in _EMBEDDED_SALES_HEADER_GROUPS
    )


def _promote_embedded_sales_header(frame: pd.DataFrame) -> pd.DataFrame:
    """Promote a later sales header row while leaving ordinary sheets intact."""

    if frame.empty or _sales_header_score(frame.columns) >= 4:
        return frame
    best_index: Any | None = None
    best_score = 0
    for index, row in frame.head(12).iterrows():
        score = _sales_header_score(row.tolist())
        if score > best_score:
            best_index, best_score = index, score
    if best_index is None or best_score < 4:
        return frame

    position = frame.index.get_loc(best_index)
    raw_headers = frame.iloc[position].tolist()
    used: set[str] = set()
    headers: list[str] = []
    for column_index, value in enumerate(raw_headers, start=1):
        candidate = "" if _is_blank_cell(value) else str(value).strip()
        headers.append(_unique_name(candidate or f"未命名列{column_index}", used))
    promoted = frame.iloc[position + 1 :].copy(deep=True)
    promoted.columns = headers
    promoted = promoted.loc[~promoted.apply(lambda row: all(_is_blank_cell(value) for value in row), axis=1)]
    keep_columns = [column for column in promoted.columns if not promoted[column].map(_is_blank_cell).all()]
    return promoted.loc[:, keep_columns].reset_index(drop=True)


def _normalise_paths(paths: str | Path | Iterable[str | Path]) -> list[Path]:
    if isinstance(paths, (str, Path)):
        result = [Path(paths)]
    else:
        result = [Path(path) for path in paths]
    if not result:
        raise ValueError("至少需要提供一个 Excel 或 CSV 文件")
    return result


def _unique_name(candidate: str, used: set[str], *, separator: str = "_") -> str:
    """Return a case-insensitively unique name."""

    name = candidate
    counter = 2
    while name.casefold() in used:
        name = f"{candidate}{separator}{counter}"
        counter += 1
    used.add(name.casefold())
    return name


def load_tables_from_files(
    paths: str | Path | Iterable[str | Path],
    *,
    csv_encoding: str | None = None,
    csv_options: Mapping[str, Any] | None = None,
    excel_options: Mapping[str, Any] | None = None,
) -> dict[str, pd.DataFrame]:
    """Load Excel/CSV files into independent DataFrames.

    Excel files load every worksheet. Keys have the stable form
    ``"文件名.xlsx::工作表"``; CSV/TSV keys use ``"文件名.csv::CSV"``.
    Duplicate keys receive a ``#2`` suffix. Returned DataFrames do not share
    mutable state with pandas' reader objects.
    """

    csv_kwargs = dict(csv_options or {})
    excel_kwargs = dict(excel_options or {})
    if "sheet_name" in excel_kwargs:
        raise ValueError("excel_options 不应包含 sheet_name；本工具始终读取全部工作表")
    if csv_encoding is not None:
        csv_kwargs["encoding"] = csv_encoding
    # Preserve business identifiers and literal codes such as "NA". Explicit
    # caller options still override these conservative import defaults.
    csv_kwargs.setdefault("dtype", str)
    csv_kwargs.setdefault("keep_default_na", False)
    excel_kwargs.setdefault("dtype", object)
    excel_kwargs.setdefault("keep_default_na", False)

    tables: dict[str, pd.DataFrame] = {}
    used: set[str] = set()
    for path in _normalise_paths(paths):
        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{path}")
        if not path.is_file():
            raise ValueError(f"不是文件：{path}")

        suffix = path.suffix.lower()
        if suffix in SUPPORTED_EXCEL_SUFFIXES:
            try:
                sheets = pd.read_excel(path, sheet_name=None, **excel_kwargs)
            except ImportError as exc:
                if suffix == ".xls":
                    raise RuntimeError("读取旧版 .xls 需要可选依赖 xlrd；建议先另存为 .xlsx") from exc
                raise
            for sheet_name, frame in sheets.items():
                key = _unique_name(f"{path.name}::{sheet_name}", used, separator="#")
                tables[key] = _promote_embedded_sales_header(frame.copy(deep=True))
        elif suffix in SUPPORTED_TEXT_SUFFIXES:
            options = dict(csv_kwargs)
            if suffix == ".tsv":
                options.setdefault("sep", "\t")
            if "encoding" in options:
                frame = pd.read_csv(path, **options)
            else:
                # utf-8-sig also accepts plain UTF-8. GB18030 covers common
                # Chinese Windows exports (GBK/GB2312) without an extra package.
                decode_error: UnicodeDecodeError | None = None
                for encoding in ("utf-8-sig", "gb18030"):
                    try:
                        frame = pd.read_csv(path, encoding=encoding, **options)
                        break
                    except UnicodeDecodeError as exc:
                        decode_error = exc
                else:  # pragma: no cover - defensive; both codecs rarely fail
                    assert decode_error is not None
                    raise decode_error
            key = _unique_name(f"{path.name}::CSV", used, separator="#")
            tables[key] = frame.copy(deep=True)
        else:
            raise ValueError(f"不支持的文件类型 {suffix or '<无扩展名>'}：{path.name}；支持 .xlsx/.xlsm/.xls/.csv/.tsv")
    return tables


def sanitise_sheet_name(name: str, *, fallback: str = "Sheet") -> str:
    cleaned = INVALID_EXCEL_SHEET_CHARS.sub("_", str(name)).strip().strip("'")
    return (cleaned or fallback)[:31]


def sanitise_filename(name: str, *, fallback: str = "table") -> str:
    cleaned = INVALID_FILENAME_CHARS.sub("_", str(name)).strip().rstrip(".")
    return cleaned or fallback


def _escape_formula_cells(frame: pd.DataFrame) -> pd.DataFrame:
    """Neutralise text that spreadsheet programs may execute as a formula."""

    result = frame.copy(deep=True)

    def escape(value: Any) -> Any:
        if not isinstance(value, str) or value.startswith("'"):
            return value
        stripped = value.lstrip(" \t\r\n")
        if stripped.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    for position in range(result.shape[1]):
        series = result.iloc[:, position]
        is_text = pd.api.types.is_object_dtype(series.dtype) or pd.api.types.is_string_dtype(series.dtype)
        if not is_text:
            continue
        result.isetitem(position, series.map(escape, na_action="ignore"))
    if isinstance(result.columns, pd.MultiIndex):
        result.columns = pd.MultiIndex.from_tuples(
            [tuple(escape(part) for part in key) for key in result.columns],
            names=[escape(name) for name in result.columns.names],
        )
    else:
        result.columns = pd.Index([escape(column) for column in result.columns], name=escape(result.columns.name))
    return result


def _normalise_tables(tables: Mapping[str, pd.DataFrame], *, escape_formulas: bool) -> dict[str, pd.DataFrame]:
    if not tables:
        raise ValueError("没有可导出的数据表")
    result: dict[str, pd.DataFrame] = {}
    normalised_names: set[str] = set()
    for raw_name, frame in tables.items():
        name = str(raw_name).strip()
        if not name:
            raise ValueError("数据表名称不能为空")
        if name in normalised_names:
            raise ValueError(f"规范化后出现重复数据表名称：{name!r}")
        if not isinstance(frame, pd.DataFrame):
            raise TypeError(f"数据表 {name!r} 不是 pandas DataFrame")
        copied = frame.copy(deep=True)
        result[name] = _escape_formula_cells(copied) if escape_formulas else copied
        normalised_names.add(name)
    return result


def _unique_sheet_name(candidate: str, used: set[str]) -> str:
    """Return an Excel-valid, case-insensitively unique sheet name."""

    base = sanitise_sheet_name(candidate)
    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        name = f"{base[: 31 - len(suffix)]}{suffix}"
        if name.casefold() not in used:
            used.add(name.casefold())
            return name
        counter += 1


def _operation_records(
    operation_log: OperationLog | Sequence[OperationRecord] | None,
) -> tuple[OperationRecord, ...]:
    if operation_log is None:
        return ()
    if isinstance(operation_log, OperationLog):
        return operation_log.entries
    records = tuple(operation_log)
    if not all(isinstance(record, OperationRecord) for record in records):
        raise TypeError("operation_log 必须是 OperationLog 或 OperationRecord 序列")
    return records


def operation_log_frame(records: Sequence[OperationRecord]) -> pd.DataFrame:
    columns = ["时间(UTC)", "操作", "输入表", "输出表", "详情"]
    rows = []
    for record in records:
        rows.append(
            {
                "时间(UTC)": record.timestamp,
                "操作": record.action,
                "输入表": "；".join(record.input_tables),
                "输出表": "；".join(record.output_tables),
                "详情": json.dumps(dict(record.details), ensure_ascii=False, default=str, sort_keys=True),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _prepare_destination(output_path: str | Path, *, overwrite: bool) -> Path:
    destination = Path(output_path)
    if destination.exists() and not overwrite:
        raise FileExistsError(f"输出文件已存在：{destination}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def _atomic_temp_path(destination: Path) -> Path:
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{destination.stem}_", suffix=destination.suffix, dir=destination.parent
    )
    os.close(descriptor)
    return Path(temp_name)


_HEADER_FILL = PatternFill("solid", fgColor="0B6B46")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_BORDER = Border(bottom=Side(style="medium", color="084C35"))
_SUMMARY_KEY_FILL = PatternFill("solid", fgColor="EAF7F0")
_LONG_TEXT_HEADER = re.compile(
    r"问题|说明|描述|备注|意见|评价|评语|摘要|原因|内容|详情|风险|提示|建议|依据|口径|规则|结论|动作|措施|"
    r"comment|description|remark|note|summary|reason|risk|advice|evidence",
    re.IGNORECASE,
)
_COMPACT_TEXT_HEADER = re.compile(
    r"^(?:风险等级|风险级别|风险扣分|风险评分|风险分|风险数量|风险记录数|入选状态|状态|类型|类别|级别)$",
    re.IGNORECASE,
)
_PERCENT_HEADER = re.compile(
    r"差异率|增长率|利润率|毛利率|净利率|资产负债率|收益率|ROA|ROE|占比|百分比|margin|percentage",
    re.IGNORECASE,
)
_INTEGER_HEADER = re.compile(r"排名|订单数|记录数|数量|件数|count|rank", re.IGNORECASE)
_MULTIPLE_HEADER = re.compile(
    r"流动比率|速动比率|现金比率|周转率|倍数|multiple|turnover",
    re.IGNORECASE,
)
_FINANCIAL_AMOUNT_HEADER = re.compile(
    r"金额|收入|支出|利润|预算|实际金额|成本|现金流|余额|应收|应付|借方|贷方|差额|未结|合计",
    re.IGNORECASE,
)


def _apply_semantic_number_formats(worksheet: Any, *, header_row: int = 1) -> None:
    """Give common finance outputs audit-friendly Excel number formats."""

    for column_index in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(row=header_row, column=column_index).value or "").strip()
        if _INTEGER_HEADER.search(header):
            number_format = "#,##0"
        elif _PERCENT_HEADER.search(header):
            number_format = "0.00%;[Red](0.00%);-"
        elif _MULTIPLE_HEADER.search(header):
            number_format = "0.00x;[Red](0.00x);-"
        elif _FINANCIAL_AMOUNT_HEADER.search(header):
            number_format = "#,##0.00;[Red](#,##0.00);-"
        else:
            continue
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _display_width(value: Any) -> int:
    """Estimate a readable Excel column width for Chinese and Latin text."""

    text = "" if value is None else str(value)
    return sum(2 if "\u2e80" <= character <= "\uffff" else 1 for character in text)


def _style_worksheet(
    worksheet: Any,
    *,
    summary: bool = False,
    long_text_detail: bool = False,
    header_row: int = 1,
) -> None:
    """Apply a compact, professional style without touching cell values.

    Width sampling is deliberately bounded so styling remains cheap for large
    exports.  The function only formats the header and (for summary sheets) the
    first column; it does not materialise hundreds of thousands of styled cells.
    """

    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = "159A62" if summary else "6BCB9B"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.orientation = "landscape"
    worksheet.print_title_rows = f"{header_row}:{header_row}"
    worksheet.sheet_view.zoomScale = 90 if summary or long_text_detail else 80

    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    last_column_letter = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{worksheet.max_row}"
    worksheet.row_dimensions[header_row].height = 30
    for cell in worksheet[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = Font(name="微软雅黑", size=10, color="FFFFFF", bold=True)
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_end = min(worksheet.max_row, 250)
    long_text_columns: dict[int, float] = {}
    for column_index in range(1, worksheet.max_column + 1):
        sampled_values = [
            worksheet.cell(row=row_index, column=column_index).value for row_index in range(header_row, sample_end + 1)
        ]
        observed_width = max((_display_width(value) for value in sampled_values), default=10)
        header = str(sampled_values[0] or "") if sampled_values else ""
        is_long_text = bool(_LONG_TEXT_HEADER.search(header)) or observed_width > 64
        if summary and column_index == 2:
            is_long_text = True
        if is_long_text:
            # Source tables are an at-a-glance overview.  Keep narrative
            # columns compact there and expose every full item in the
            # companion ``长文本明细`` sheet.
            width = 56 if summary or long_text_detail else 34
            long_text_columns[column_index] = width
        else:
            width = min(28, max(10, observed_width + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    if long_text_columns:
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            estimated_lines = 1
            for column_index, width in long_text_columns.items():
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
                display_width = _display_width(cell.value)
                explicit_lines = str(cell.value or "").count("\n") + 1
                estimated_lines = max(
                    estimated_lines,
                    explicit_lines,
                    math.ceil(display_width / max(width, 1)),
                )
            row_height_cap = 135 if summary or long_text_detail else 105
            worksheet.row_dimensions[row_index].height = min(
                row_height_cap,
                max(20, 15 * estimated_lines),
            )

    if summary and worksheet.max_column >= 1:
        for row_index in range(header_row + 1, min(worksheet.max_row, 5000) + 1):
            cell = worksheet.cell(row=row_index, column=1)
            cell.fill = _SUMMARY_KEY_FILL
            cell.font = Font(bold=True, color="0B5D3B")

    _apply_semantic_number_formats(worksheet, header_row=header_row)


def _estimated_wrapped_lines(value: Any, width: float) -> int:
    """Estimate Excel display lines for mixed Chinese/Latin narrative text."""

    text = "" if value is None else str(value)
    if not text:
        return 1
    usable_width = max(float(width) - 2.0, 8.0)
    lines = 0
    for paragraph in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        lines += max(1, math.ceil(_display_width(paragraph) / usable_width))
    return max(lines, 1)


def _finalize_adaptive_layout(worksheet: Any, *, header_row: int) -> None:
    """Apply the final readability pass shared by every generated worksheet.

    Specialist report styles establish colours and hierarchy first.  This pass
    then normalises usable widths, wraps every narrative field, and expands each
    populated row according to its longest visible cell.  It deliberately caps
    columns and rows so one verbose value cannot make the workbook unusably wide
    or tall.
    """

    if worksheet.max_row < header_row or worksheet.max_column < 1:
        return
    worksheet.sheet_format.defaultRowHeight = 21
    worksheet.row_dimensions[header_row].height = max(
        worksheet.row_dimensions[header_row].height or 0,
        32,
    )
    for cell in worksheet[header_row]:
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    if worksheet.title == "经营诊断看板":
        return

    sample_end = min(worksheet.max_row, header_row + 500)
    narrative_columns: dict[int, float] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(header_row, column_index).value or "").strip()
        sampled_values = [
            worksheet.cell(row, column_index).value
            for row in range(header_row + 1, sample_end + 1)
            if worksheet.cell(row, column_index).value not in (None, "")
        ]
        widths = [_display_width(header), *(_display_width(value) for value in sampled_values)]
        observed = max(widths, default=10)
        existing = float(worksheet.column_dimensions[get_column_letter(column_index)].width or 10)
        narrative = (
            bool(_LONG_TEXT_HEADER.search(header)) and not _COMPACT_TEXT_HEADER.fullmatch(header)
        ) or observed >= 72
        if narrative:
            # 42 works for ordinary notes; denser review text receives more
            # horizontal room but never grows into an unscannable wall.
            target = 56.0 if observed >= 180 else 46.0
            width = min(58.0, max(36.0, min(existing, target) if existing >= 36 else target))
            narrative_columns[column_index] = width
        else:
            target = min(30.0, max(10.0, observed + 2.0))
            width = min(32.0, max(target, min(existing, 32.0)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    if not narrative_columns:
        return

    # Excel supports at most 409.5 points.  Keeping a little headroom avoids
    # viewer-specific clipping while still showing roughly 25 lines of text.
    maximum_height = 390.0
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        longest = 1
        has_narrative = False
        for column_index, width in narrative_columns.items():
            cell = worksheet.cell(row_index, column_index)
            if cell.value in (None, ""):
                continue
            has_narrative = True
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
            longest = max(longest, _estimated_wrapped_lines(cell.value, width))
        if has_narrative:
            calculated = min(maximum_height, max(24.0, 15.5 * longest + 5.0))
            worksheet.row_dimensions[row_index].height = max(
                worksheet.row_dimensions[row_index].height or 0,
                calculated,
            )


_SALES_REPORT_SHEETS = (
    "管理层数据总览",
    "季度合并数据",
    "产品分析",
    "地区分析",
    "销售人员分析",
    "异常数据提醒",
    "清洗审计",
    "图表展示",
)

_INVENTORY_REPORT_SHEETS = (
    "管理层库存总览",
    "商品库存分析",
    "补货建议",
    "积压清单",
    "采购分析",
    "销售分析",
    "人工核验",
    "数据审计",
    "库存图表看板",
)

_INVENTORY_REPORT_TITLES = {
    "管理层库存总览": ("库存经营管理驾驶舱", "采购、销售、库存与风险预警｜口径和阈值已在数据审计中披露"),
    "商品库存分析": ("商品库存全景分析", "期初、入库、出库、调整、可售库存、库存天数和管理建议"),
    "补货建议": ("采购补货建议", "结合安全库存、采购提前期、近30天销量、目标库存与在途库存"),
    "积压清单": ("库存积压与清仓清单", "超过目标库存天数阈值、无近期动销或停售仍有库存的商品"),
    "采购分析": ("采购入库分析", "仅统计已入库采购｜待入库单独列示为在途"),
    "销售分析": ("销售出库分析", "仅统计已完成销售｜毛利按商品资料采购单价估算"),
    "人工核验": ("需要人工确认的业务事项", "退货是否重新入库、缺失数量、待确认调整、未知商品或负库存"),
    "数据审计": ("库存核算与数据清洗审计", "重复、取消、退货、待确认和缺失字段均保留处理原因"),
    "库存图表看板": ("采购销售库存可视化看板", "月度进销、品类库存金额和库存状态结构｜图表可在 Excel / WPS 中继续编辑"),
}

_INVENTORY_TAB_COLOURS = {
    "管理层库存总览": "17324D",
    "商品库存分析": "2F75B5",
    "补货建议": "00A389",
    "积压清单": "D99614",
    "采购分析": "6B5FD2",
    "销售分析": "2F75B5",
    "人工核验": "E26A45",
    "数据审计": "B42318",
    "库存图表看板": "0B6B46",
}

_HR_REPORT_SHEETS = (
    "管理层人效总览",
    "员工综合分析",
    "表现优秀员工",
    "重点关注员工",
    "考勤分析",
    "绩效分析",
    "薪资分析",
    "人工核验",
    "数据审计",
    "人力图表看板",
)

_HR_REPORT_TITLES = {
    "管理层人效总览": ("员工经营管理驾驶舱", "考勤、绩效、薪资与管理预警｜评分口径和人工边界已在数据审计中披露"),
    "员工综合分析": ("员工综合表现全景", "统一查看考勤、绩效、薪资、综合排名与可核验事实依据"),
    "表现优秀员工": ("表现优秀员工候选", "达到综合阈值、迟到不超过1次且无缺勤｜仍需直属主管复核岗位成果"),
    "重点关注员工": ("重点关注与改进清单", "仅基于考勤和绩效事实触发预警，不替代正式人事结论"),
    "考勤分析": ("月度考勤分析", "出勤、迟到、早退、请假、缺勤和加班情况"),
    "绩效分析": ("绩效与目标完成分析", "目标完成率、客户评分、绩效得分和综合排名"),
    "薪资分析": ("薪资与调整分析", "基本工资、奖金/提成/补贴/扣款与预计薪资"),
    "人工核验": ("需要人工确认的人员事项", "工号、状态、考勤、绩效和薪资调整冲突不自动推断"),
    "数据审计": ("评分口径与数据审计", "所有阈值、公式、代理指标含义和数据质量问题均可追溯"),
    "人力图表看板": ("员工经营可视化看板", "部门人力成本、员工综合表现、风险结构和考勤异常｜图表可继续编辑"),
}

_HR_TAB_COLOURS = {
    "管理层人效总览": "17324D",
    "员工综合分析": "2F75B5",
    "表现优秀员工": "00A389",
    "重点关注员工": "E26A45",
    "考勤分析": "6B5FD2",
    "绩效分析": "2F75B5",
    "薪资分析": "D99614",
    "人工核验": "B42318",
    "数据审计": "6B7C8F",
    "人力图表看板": "0B6B46",
}

_ADAPTIVE_REPORT_SHEETS = (
    "管理层通用总览",
    "主数据分析",
    "数据字典",
    "数据质量",
    "表关系建议",
    "分类排名",
    "时间趋势",
    "异常数据",
    "自适应图表看板",
)

_ADAPTIVE_REPORT_TITLES = {
    "管理层通用总览": (
        "AI 通用经营分析驾驶舱",
        "需求意图、领域语义、指标口径、证据缺口和图表均由当前数据动态编译",
    ),
    "主数据分析": ("自适应主数据分析", "同构分表自动合并并删除完全重复记录｜保留来源数据表便于追溯"),
    "数据字典": ("字段角色与数据字典", "逐字段展示类型、非空、唯一性、示例值与推断角色"),
    "数据质量": ("数据质量概览", "缺失、重复、行列规模和质量状态按表展示"),
    "表关系建议": ("候选表关系建议", "依据同名字段、值覆盖率和键唯一性推断｜正式连接前必须核验业务含义"),
    "分类排名": ("动态分类表现", "仅在分类维度与可聚合指标同时存在时生成排名和结构分析"),
    "时间趋势": ("动态时间趋势", "仅在时间字段与可聚合指标同时存在时生成，比例指标按业务口径重算"),
    "异常数据": ("风险、异常与证据缺口", "区分数据异常和分析证据不足｜不自动删除、不强行下结论"),
    "自适应图表看板": ("AI 动态经营可视化", "只展示当前证据支持的排名、结构、趋势和风险图｜原生 Excel 图表可继续编辑"),
}

_ADAPTIVE_TAB_COLOURS = {
    "管理层通用总览": "17324D",
    "主数据分析": "2F75B5",
    "数据字典": "6B5FD2",
    "数据质量": "D99614",
    "表关系建议": "00A389",
    "分类排名": "2F75B5",
    "时间趋势": "6B5FD2",
    "异常数据": "E26A45",
    "自适应图表看板": "0B6B46",
}

_SELECTION_REPORT_SHEETS = (
    "评选管理总览",
    "建议入选名单",
    "全部候选排序",
    "风险复核清单",
    "评选规则与字段",
    "评选图表看板",
)

_SELECTION_REPORT_TITLES = {
    "评选管理总览": ("候选对象结构化评选驾驶舱", "自动识别标识、历次得分和评语｜评分口径、扣分与人工边界完整披露"),
    "建议入选名单": ("建议入选名单", "按综合推荐分排序｜每位候选均保留得分、风险、理由和源行号"),
    "全部候选排序": ("全部候选综合排序", "有效均分、最新表现、完整率、文本风险和正向依据共同形成可解释排序"),
    "风险复核清单": ("入选前风险复核", "中高风险只作为核验提示，不由程序自动取消参赛资格"),
    "评选规则与字段": ("评选口径与字段识别", "识别字段、计算公式、缺失处理、风险规则和最终人工边界"),
    "评选图表看板": ("候选表现可视化看板", "综合推荐分、有效均分、最新得分与风险扣分｜原生Excel图表可继续编辑"),
}

_SELECTION_TAB_COLOURS = {
    "评选管理总览": "17324D",
    "建议入选名单": "00A389",
    "全部候选排序": "2F75B5",
    "风险复核清单": "E26A45",
    "评选规则与字段": "6B7C8F",
    "评选图表看板": "0B6B46",
}

_ENTERPRISE_REPORT_SHEETS = (
    "管理层诊断总览",
    "数据源与事实域",
    "门店经营诊断",
    "渠道与外卖分析",
    "菜品盈利分析",
    "原料采购与损耗",
    "人工效率分析",
    "客户评价与退款",
    "利润驱动分析",
    "渠道与广告诊断",
    "商品利润质量",
    "退款售后风险",
    "平台费用与回款",
    "广告效率分析",
    "采购成本分析",
    "客户与回款风险",
    "销售团队诊断",
    "成本费用分析",
    "库存风险分析",
    "风险行动计划",
    "诊断底稿",
    "数据口径与验收",
    "经营诊断看板",
)

_ENTERPRISE_REPORT_TITLES = {
    "管理层诊断总览": (
        "企业集团经营诊断驾驶舱",
        "增长、利润、现金、客户、销售、成本与库存的跨表经营诊断｜所有口径可追溯",
    ),
    "数据源与事实域": ("数据源与事实域", "本次任务明确输入的工作表与事实粒度｜用于审计输入范围和防止历史输出回流"),
    "门店经营诊断": ("门店经营诊断", "营业实付、退款、标准食材成本、平台成本、人工与固定费用的可比口径"),
    "渠道与外卖分析": ("渠道与外卖分析", "平台结算、费用、到账和可比经营贡献｜到账不等同销售收入"),
    "菜品盈利分析": ("菜品盈利分析", "销量、实付、标准食材成本与贡献率｜标准成本缺失需人工核验"),
    "原料采购与损耗": ("原料采购与损耗", "采购入库与盘点差异分开披露｜损耗金额是管理代理指标"),
    "人工效率分析": ("人工效率分析", "工时、加班、人工成本和门店投入产出线索"),
    "客户评价与退款": ("客户评价与退款", "低评分、投诉标签、退款状态和原因可追溯关联"),
    "利润驱动分析": ("利润驱动与月度趋势", "收入、业务或生产成本、期间费用和趋势性经营贡献的月度桥接｜不替代法定利润"),
    "渠道与广告诊断": ("渠道真实盈利与广告效率", "成交、退款、平台费用、到账、标准成本和广告联合诊断｜识别买来的增长"),
    "商品利润质量": ("商品退款后利润质量", "优惠、退款和标准成本共同形成商品管理毛利｜不强行分摊无法可靠归属的广告费用"),
    "退款售后风险": ("退款售后经营风险", "已退款与处理中严格分开｜按原订单号和SKU追溯渠道、客户及原因"),
    "平台费用与回款": ("平台费用与现金转化", "结算基数、平台费用、退款冲减和实际到账逐月逐渠道勾稽"),
    "广告效率分析": ("广告投入与归因效率", "广告花费、归因成交、ROAS与点击效率｜平台归因不与订单收入相加"),
    "采购成本分析": ("采购成本变化线索", "实际采购价与标准成本、首次采购价对比｜待质检与重复入库单独审计"),
    "客户与回款风险": ("客户价值与回款风险", "客户收入贡献、毛利质量、订单风险敞口、信用与满意度联合诊断"),
    "销售团队诊断": (
        "销售规模与质量诊断",
        "收入、绩效口径参考毛利、回款、客户评分和投诉的综合评价｜缺失订单成本不显示为零",
    ),
    "成本费用分析": ("部门毛利与期间费用结构", "业务成本与期间费用分开披露，避免重复计算"),
    "库存风险分析": (
        "库存周转与资金占用线索",
        "库存月数、金额集中和偏高/缺货线索｜未提供安全库存与季节性时不作绝对判定",
    ),
    "风险行动计划": ("90天经营改进行动表", "每项风险均给出证据、动作、责任人、时限、审批点和验收指标"),
    "诊断底稿": ("经营诊断可追溯底稿", "逐笔流水关联客户风险并保留所有计算字段和触发标记"),
    "数据口径与验收": ("数据口径、勾稽与人工边界", "通过项、待核验项和计算限制集中披露，便于客户验收"),
    "经营诊断看板": ("企业经营管理驾驶舱", "核心指标、首要诊断、优先风险和经营趋势集中呈现｜原生 Excel 图表可编辑"),
}

_ENTERPRISE_TAB_COLOURS = {
    "管理层诊断总览": "17324D",
    "数据源与事实域": "0B6B46",
    "门店经营诊断": "2F75B5",
    "渠道与外卖分析": "008C72",
    "菜品盈利分析": "6B5FD2",
    "原料采购与损耗": "D99614",
    "人工效率分析": "00A389",
    "客户评价与退款": "E26A45",
    "利润驱动分析": "2F75B5",
    "渠道与广告诊断": "B42318",
    "商品利润质量": "6B5FD2",
    "退款售后风险": "E26A45",
    "平台费用与回款": "008C72",
    "广告效率分析": "D99614",
    "采购成本分析": "6B7C8F",
    "客户与回款风险": "E26A45",
    "销售团队诊断": "00A389",
    "成本费用分析": "6B5FD2",
    "库存风险分析": "D99614",
    "风险行动计划": "B42318",
    "诊断底稿": "6B7C8F",
    "数据口径与验收": "0B6B46",
    "经营诊断看板": "17324D",
}


_REPORT_TITLES = {
    "管理层数据总览": ("销售经营管理驾驶舱", "关键经营指标、排名与异常监控｜数据来源：当前上传销售明细"),
    "季度合并数据": ("季度有效订单明细", "三月数据已统一字段、日期、金额和文本｜仅包含去重后有效订单"),
    "产品分析": ("产品经营分析", "销售额、成本、利润、利润率与贡献排名"),
    "地区分析": ("地区经营分析", "区域销售规模、利润贡献与排名"),
    "销售人员分析": ("销售团队业绩分析", "个人业绩、利润贡献与排名"),
    "异常数据提醒": ("重点关注与异常数据", "满意度、利润与金额完整性自动检查｜请结合业务凭证复核"),
    "清洗审计": ("数据清洗与排除审计", "原始记录、重复剔除、无效订单排除及完整原因"),
    "图表展示": ("销售经营可视化看板", "趋势、结构与区域表现｜源数据表位于本页下方，图表可在 Excel / WPS 中继续编辑"),
}

_REPORT_TAB_COLOURS = {
    "管理层数据总览": "17324D",
    "季度合并数据": "0B6B46",
    "产品分析": "2F75B5",
    "地区分析": "D99614",
    "销售人员分析": "00A389",
    "异常数据提醒": "E26A45",
    "清洗审计": "B42318",
    "图表展示": "6B5FD2",
}


def _report_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor="17324D")
    title_cell.font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    subtitle_cell = worksheet.cell(row=2, column=1)
    subtitle_cell.value = subtitle
    subtitle_cell.fill = PatternFill("solid", fgColor="EAF1F7")
    subtitle_cell.font = Font(name="微软雅黑", size=10, color="4B6275")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    for column_index in range(2, end_column + 1):
        worksheet.cell(row=1, column=column_index).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(row=2, column=column_index).fill = PatternFill("solid", fgColor="EAF1F7")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _set_report_widths(worksheet: Any, widths: Mapping[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _style_report_data_grid(worksheet: Any, *, header_row: int) -> None:
    light_border = Border(bottom=Side(style="hair", color="D9E2EA"))
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = max(worksheet.row_dimensions[row_index].height or 0, 23)
        for cell in worksheet[row_index]:
            cell.font = Font(name="微软雅黑", size=10, color="263746")
            cell.border = light_border
            cell.alignment = Alignment(
                horizontal="right"
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
                else "left",
                vertical="center",
                wrap_text=False,
            )
            if row_index % 2 == 0 and cell.fill.fill_type is None:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")


def _add_management_cards(worksheet: Any, *, header_row: int) -> None:
    metrics = {
        str(worksheet.cell(row=row_index, column=1).value or ""): worksheet.cell(row=row_index, column=2).value
        for row_index in range(header_row + 1, worksheet.max_row + 1)
    }
    cards = (
        ("F4:H4", "F5:H7", "总销售额", metrics.get("总销售额", metrics.get("季度总销售额")), '#,##0 "元"', "2F75B5"),
        ("J4:L4", "J5:L7", "总成本", metrics.get("总成本", metrics.get("季度总成本")), '#,##0 "元"', "6B7C8F"),
        ("N4:P4", "N5:P7", "总利润", metrics.get("总利润", metrics.get("季度总利润")), '#,##0 "元"', "00A389"),
        ("F9:H9", "F10:H12", "平均利润率", metrics.get("平均利润率", metrics.get("整体利润率")), "0.0%", "6B5FD2"),
        ("J9:L9", "J10:L12", "最佳产品", metrics.get("销售额最高产品"), "@", "D99614"),
        ("N9:P9", "N10:P12", "最佳销售", metrics.get("业绩最佳销售人员"), "@", "E26A45"),
    )
    for label_range, value_range, label, value, number_format, colour in cards:
        worksheet.merge_cells(label_range)
        worksheet.merge_cells(value_range)
        label_cell = worksheet[label_range.split(":")[0]]
        value_cell = worksheet[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=colour)
        label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.value = value if value not in (None, "") else "—"
        value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
        value_cell.font = Font(name="微软雅黑", size=18, bold=True, color="17324D")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
    for column in ("F", "G", "H", "J", "K", "L", "N", "O", "P"):
        worksheet.column_dimensions[column].width = 9.5
    worksheet.column_dimensions["I"].width = 2.5
    worksheet.column_dimensions["M"].width = 2.5


def _style_sales_management_sheet(worksheet: Any, *, header_row: int) -> None:
    """Build an executive-ready, WPS-safe sales management report layout."""

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _REPORT_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)

    if worksheet.title == "管理层数据总览":
        _report_title_band(worksheet, end_column=16)
        _set_report_widths(worksheet, {"A": 24, "B": 20, "C": 10, "D": 46})
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row=row_index, column=1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
        for row_index in range(header_row + 1, min(header_row + 3, worksheet.max_row) + 1):
            worksheet.cell(row=row_index, column=2).number_format = "#,##0.00;[Red](#,##0.00);-"
        if worksheet.max_row >= header_row + 4:
            worksheet.cell(row=header_row + 4, column=2).number_format = "0.00%"
        _add_management_cards(worksheet, header_row=header_row)
    elif worksheet.title == "季度合并数据":
        _report_title_band(worksheet, end_column=max(16, worksheet.max_column))
        _set_report_widths(
            worksheet,
            {
                "A": 15,
                "B": 14,
                "C": 11,
                "D": 16,
                "E": 11,
                "F": 13,
                "G": 11,
                "H": 16,
                "I": 16,
                "J": 16,
                "K": 14,
                "L": 14,
                "M": 14,
                "N": 22,
                "O": 14,
                "P": 24,
            },
        )
        header_map = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            if header_map.get("日期"):
                worksheet.cell(row=row_index, column=header_map["日期"]).number_format = "yyyy-mm-dd"
            for header in ("销售额", "成本", "利润"):
                if header_map.get(header):
                    worksheet.cell(
                        row=row_index, column=header_map[header]
                    ).number_format = "#,##0.00;[Red](#,##0.00);-"
            if header_map.get("利润率"):
                worksheet.cell(row=row_index, column=header_map["利润率"]).number_format = "0.00%"
    elif worksheet.title in {"产品分析", "地区分析", "销售人员分析"}:
        _report_title_band(worksheet, end_column=9)
        _set_report_widths(
            worksheet,
            {"A": 18, "B": 17, "C": 17, "D": 17, "E": 11, "F": 14, "G": 14, "H": 13, "I": 13},
        )
        header_map = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for header in ("销售额", "成本", "利润"):
            column = header_map.get(header)
            if column:
                for row_index in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column).number_format = "#,##0.00;[Red](#,##0.00);-"
        for header in ("利润率", "销售占比"):
            column = header_map.get(header)
            if column:
                for row_index in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column).number_format = "0.00%"
        sales_column = header_map.get("销售额")
        if sales_column and worksheet.max_row > header_row:
            data_range = f"{get_column_letter(sales_column)}{header_row + 1}:{get_column_letter(sales_column)}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                data_range,
                DataBarRule(start_type="min", end_type="max", color="8EC5F5", showValue=True),
            )
        if worksheet.max_row > header_row:
            for cell in worksheet[header_row + 1]:
                cell.fill = PatternFill("solid", fgColor="FFF4D6")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="6B4F00")
    elif worksheet.title == "异常数据提醒":
        _report_title_band(worksheet, end_column=max(12, worksheet.max_column))
        _set_report_widths(
            worksheet,
            {
                "A": 14,
                "B": 16,
                "C": 11,
                "D": 13,
                "E": 12,
                "F": 17,
                "G": 17,
                "H": 14,
                "I": 17,
                "J": 13,
                "K": 12,
                "L": 38,
            },
        )
        header_map = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for header in ("销售金额", "销售额", "成本", "利润"):
            column = header_map.get(header)
            if column:
                for row_index in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column).number_format = "#,##0.00;[Red](#,##0.00);-"
        reason_column = header_map.get("关注原因")
        level_column = header_map.get("关注级别")
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = 30
            if reason_column:
                worksheet.cell(row=row_index, column=reason_column).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if level_column:
                cell = worksheet.cell(row=row_index, column=level_column)
                if cell.value == "高":
                    cell.fill = PatternFill("solid", fgColor="FDE8E7")
                    cell.font = Font(name="微软雅黑", color="B42318", bold=True)
                elif cell.value == "中":
                    cell.fill = PatternFill("solid", fgColor="FFF4D6")
                    cell.font = Font(name="微软雅黑", color="9A6700", bold=True)
    elif worksheet.title == "清洗审计":
        _report_title_band(worksheet, end_column=max(9, worksheet.max_column))
        _set_report_widths(
            worksheet,
            {"A": 12, "B": 16, "C": 26, "D": 14, "E": 48, "F": 16, "G": 16, "H": 16, "I": 28},
        )
        header_map = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = 30
            reason_column = header_map.get("排除原因")
            if reason_column:
                worksheet.cell(row=row_index, column=reason_column).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if worksheet.cell(row=row_index, column=1).value == "明细":
                for cell in worksheet[row_index]:
                    cell.fill = PatternFill("solid", fgColor="FFF1F0")
    elif worksheet.title == "图表展示":
        _report_title_band(worksheet, end_column=19)
        worksheet.freeze_panes = None
        # Clearing ``freeze_panes`` alone leaves openpyxl's old
        # ``pane=\"bottomLeft\"`` selection behind. Desktop Excel repairs that
        # inconsistent view record when opening the workbook, so replace the
        # selection with a normal A1 selection as well.
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.zoomScale = 80
        _set_report_widths(
            worksheet,
            {
                "A": 14,
                "B": 17,
                "C": 18,
                "D": 17,
                "E": 14,
                "F": 17,
                "G": 3,
                "H": 10,
                "I": 10,
                "J": 3,
                "K": 11,
                "L": 11,
                "M": 11,
                "N": 11,
                "O": 11,
                "P": 11,
                "Q": 11,
                "R": 11,
                "S": 11,
            },
        )


def _last_data_row(worksheet: Any, column_index: int, *, header_row: int = 1) -> int:
    for row_index in range(worksheet.max_row, header_row, -1):
        if worksheet.cell(row=row_index, column=column_index).value not in (None, ""):
            return row_index
    return header_row


def _style_line_series(series: Any, colour: str) -> None:
    series.graphicalProperties.line.solidFill = colour
    series.graphicalProperties.line.width = 28575
    series.smooth = True
    series.marker.symbol = "circle"
    series.marker.size = 7
    series.marker.graphicalProperties.solidFill = colour
    series.marker.graphicalProperties.line.solidFill = "FFFFFF"


def _add_chart_insight_panel(
    worksheet: Any,
    *,
    headers: Mapping[str, int],
    header_row: int,
    month_end: int,
    product_end: int,
    region_end: int,
) -> None:
    insights: list[tuple[str, str]] = []
    if month_end > header_row:
        rows = [
            (
                str(worksheet.cell(row=row, column=headers["月份"]).value or ""),
                worksheet.cell(row=row, column=headers["月度销售额"]).value,
            )
            for row in range(header_row + 1, month_end + 1)
        ]
        rows = [(label, value) for label, value in rows if isinstance(value, (int, float))]
        if rows:
            label, value = max(rows, key=lambda item: float(item[1]))
            insights.append(("销售峰值月份", f"{label}｜{float(value):,.0f} 元"))
    if product_end > header_row:
        label = worksheet.cell(row=header_row + 1, column=headers["产品类别"]).value
        value = worksheet.cell(row=header_row + 1, column=headers["产品销售额"]).value
        if label not in (None, "") and isinstance(value, (int, float)):
            insights.append(("销售贡献最高产品", f"{label}｜{float(value):,.0f} 元"))
    if region_end > header_row:
        rows = [
            (
                worksheet.cell(row=row, column=headers["地区"]).value,
                worksheet.cell(row=row, column=headers["地区销售额"]).value,
            )
            for row in range(header_row + 1, region_end + 1)
        ]
        rows = [(label, value) for label, value in rows if label not in (None, "") and isinstance(value, (int, float))]
        if rows:
            label, value = max(rows, key=lambda item: float(item[1]))
            insights.append(("销售贡献最高地区", f"{label}｜{float(value):,.0f} 元"))

    worksheet.merge_cells("K21:S22")
    worksheet["K21"] = "经营要点"
    worksheet["K21"].fill = PatternFill("solid", fgColor="17324D")
    worksheet["K21"].font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
    worksheet["K21"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(12, 20):
        worksheet.cell(row=21, column=column).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(row=22, column=column).fill = PatternFill("solid", fgColor="17324D")
    start_row = 24
    for offset, (label, value) in enumerate(insights[:3]):
        row = start_row + offset * 3
        worksheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=19)
        worksheet.merge_cells(start_row=row + 1, start_column=11, end_row=row + 1, end_column=19)
        worksheet.cell(row=row, column=11).value = label
        worksheet.cell(row=row, column=11).font = Font(name="微软雅黑", size=9, color="6B7C8F")
        worksheet.cell(row=row + 1, column=11).value = value
        worksheet.cell(row=row + 1, column=11).font = Font(name="微软雅黑", size=13, bold=True, color="17324D")
        worksheet.cell(row=row + 1, column=11).fill = PatternFill("solid", fgColor="F2F6FA")
        for column in range(12, 20):
            worksheet.cell(row=row + 1, column=column).fill = PatternFill("solid", fgColor="F2F6FA")


def _add_sales_management_charts(worksheet: Any, *, header_row: int) -> None:
    """Create restrained, editable charts that render cleanly in Excel and WPS."""

    if worksheet.title != "图表展示":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    required = {"月份", "月度销售额", "产品类别", "产品销售额", "地区", "地区销售额"}
    if not required.issubset(headers):
        return
    for header in ("月度销售额", "产品销售额", "地区销售额"):
        column = headers[header]
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column).number_format = "#,##0.00;[Red](#,##0.00);-"

    month_end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
    if month_end > header_row:
        chart = LineChart()
        chart.title = "月度销售趋势（元）"
        chart.y_axis.title = "销售额（元）"
        chart.x_axis.title = "月份"
        chart.style = 2
        chart.height = 7.2
        chart.width = 14.2
        chart.legend = None
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        chart.y_axis.scaling.min = 0
        chart.y_axis.numFmt = "#,##0"
        chart.add_data(
            Reference(worksheet, min_col=headers["月度销售额"], min_row=header_row, max_row=month_end),
            titles_from_data=True,
        )
        chart.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=month_end))
        if chart.series:
            _style_line_series(chart.series[0], "2F75B5")
        worksheet.add_chart(chart, "A4")

    product_end = _last_data_row(worksheet, headers["产品类别"], header_row=header_row)
    if product_end > header_row:
        chart = DoughnutChart()
        chart.title = "产品销售占比（%）"
        chart.style = 2
        chart.height = 7.2
        chart.width = 12.5
        chart.holeSize = 58
        chart.firstSliceAng = 270
        chart.varyColors = True
        chart.add_data(
            Reference(worksheet, min_col=headers["产品销售额"], min_row=header_row, max_row=product_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["产品类别"], min_row=header_row + 1, max_row=product_end)
        )
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = False
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showVal = False
        # Do not emit dLblPos="bestFit" for a doughnut chart. The value is
        # accepted by openpyxl and some renderers, but Desktop Excel rejects
        # the resulting chart part and then removes the entire drawing layer.
        # Omitting the position lets Excel choose its compatible default while
        # preserving editable percentage labels.
        chart.legend.position = "r"
        if chart.series:
            colours = ("2F75B5", "00A389", "D99614", "6B5FD2", "E26A45", "6B7C8F")
            chart.series[0].dPt = [
                DataPoint(idx=index, spPr=GraphicalProperties(solidFill=colour))
                for index, colour in enumerate(colours[: product_end - header_row])
            ]
        worksheet.add_chart(chart, "K4")

    region_end = _last_data_row(worksheet, headers["地区"], header_row=header_row)
    if region_end > header_row:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "地区销售分布（元）"
        chart.x_axis.title = "销售额（元）"
        chart.y_axis.title = None
        chart.style = 2
        chart.height = 7.2
        chart.width = 14.2
        chart.legend = None
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
        chart.x_axis.numFmt = "#,##0"
        chart.gapWidth = 55
        chart.add_data(
            Reference(worksheet, min_col=headers["地区销售额"], min_row=header_row, max_row=region_end),
            titles_from_data=True,
        )
        chart.set_categories(Reference(worksheet, min_col=headers["地区"], min_row=header_row + 1, max_row=region_end))
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = "00A389"
            chart.series[0].graphicalProperties.line.solidFill = "00A389"
        worksheet.add_chart(chart, "A21")

    _add_chart_insight_panel(
        worksheet,
        headers=headers,
        header_row=header_row,
        month_end=month_end,
        product_end=product_end,
        region_end=region_end,
    )


def _inventory_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _INVENTORY_REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet.cell(row=1, column=1).value = title
    worksheet.cell(row=2, column=1).value = subtitle
    for column_index in range(1, end_column + 1):
        title_cell = worksheet.cell(row=1, column=column_index)
        title_cell.fill = PatternFill("solid", fgColor="17324D")
        title_cell.font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        subtitle_cell = worksheet.cell(row=2, column=column_index)
        subtitle_cell.fill = PatternFill("solid", fgColor="EAF1F7")
        subtitle_cell.font = Font(name="微软雅黑", size=10, color="4B6275")
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _add_inventory_cards(worksheet: Any, *, header_row: int) -> None:
    metrics = {
        str(worksheet.cell(row=row_index, column=1).value or ""): worksheet.cell(row=row_index, column=2).value
        for row_index in range(header_row + 1, worksheet.max_row + 1)
    }
    cards = (
        ("F4:H4", "F5:H7", "可售库存金额", metrics.get("可销售库存金额"), '#,##0 "元"', "2F75B5"),
        ("J4:L4", "J5:L7", "采购入库金额", metrics.get("采购入库金额"), '#,##0 "元"', "6B5FD2"),
        ("N4:P4", "N5:P7", "销售出库金额", metrics.get("销售出库金额"), '#,##0 "元"', "00A389"),
        ("F9:H9", "F10:H12", "需补货 SKU", metrics.get("需要补货SKU数"), '0 "个"', "E26A45"),
        ("J9:L9", "J10:L12", "积压 SKU", metrics.get("积压SKU数"), '0 "个"', "D99614"),
        ("N9:P9", "N10:P12", "积压金额", metrics.get("积压库存金额"), '#,##0 "元"', "B42318"),
    )
    for label_range, value_range, label, value, number_format, colour in cards:
        worksheet.merge_cells(label_range)
        worksheet.merge_cells(value_range)
        label_cell = worksheet[label_range.split(":")[0]]
        value_cell = worksheet[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=colour)
        label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.value = value if value not in (None, "") else "—"
        value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
        value_cell.font = Font(name="微软雅黑", size=18, bold=True, color="17324D")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
    for column in ("F", "G", "H", "J", "K", "L", "N", "O", "P"):
        worksheet.column_dimensions[column].width = 9.5
    worksheet.column_dimensions["I"].width = 2.5
    worksheet.column_dimensions["M"].width = 2.5


def _inventory_number_formats(worksheet: Any, *, header_row: int) -> None:
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    currency_headers = {
        "采购单价",
        "零售价",
        "可售库存金额",
        "积压金额",
        "预计采购金额",
        "已入库金额",
        "在途金额",
        "平均采购单价",
        "销售额",
        "估算销售成本",
        "估算毛利",
        "金额",
    }
    percentage_headers = {"估算毛利率"}
    decimal_headers = {"近30天日均销量", "可售库存天数"}
    date_headers = {"日期", "分析截止日期"}
    for header, column in headers.items():
        if header in currency_headers:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "#,##0.00;[Red](#,##0.00);-"
        elif header in percentage_headers:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "0.00%"
        elif header in decimal_headers:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "#,##0.0"
        elif header in date_headers:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "yyyy-mm-dd"


def _style_inventory_management_sheet(worksheet: Any, *, header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _INVENTORY_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 88
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)
    _inventory_number_formats(worksheet, header_row=header_row)

    if worksheet.title == "管理层库存总览":
        _inventory_title_band(worksheet, end_column=16)
        _set_report_widths(worksheet, {"A": 24, "B": 21, "C": 10, "D": 54})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row, column=1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row=row, column=1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
            if worksheet.cell(row=row, column=1).value == "分析截止日期":
                worksheet.cell(row=row, column=2).number_format = "yyyy-mm-dd"
            metric = str(worksheet.cell(row=row, column=1).value or "")
            if "金额" in metric or "毛利" in metric:
                worksheet.cell(row=row, column=2).number_format = "#,##0.00;[Red](#,##0.00);-"
            elif any(token in metric for token in ("库存", "数量", "SKU", "事项")) and metric != "分析截止日期":
                worksheet.cell(row=row, column=2).number_format = "#,##0"
        _add_inventory_cards(worksheet, header_row=header_row)
    elif worksheet.title == "商品库存分析":
        _inventory_title_band(worksheet, end_column=max(29, worksheet.max_column))
        widths = {
            "A": 13,
            "B": 22,
            "C": 14,
            "D": 16,
            "E": 12,
            "F": 11,
            "G": 10,
            "H": 10,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 14,
            "M": 14,
            "N": 11,
            "O": 11,
            "P": 14,
            "Q": 14,
            "R": 12,
            "S": 15,
            "T": 14,
            "U": 14,
            "V": 14,
            "W": 13,
            "X": 12,
            "Y": 13,
            "Z": 16,
            "AA": 16,
            "AB": 13,
            "AC": 42,
        }
        _set_report_widths(worksheet, widths)
        headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        status_col = headers.get("库存状态")
        suggestion_col = headers.get("管理建议")
        for row in range(header_row + 1, worksheet.max_row + 1):
            if suggestion_col:
                worksheet.cell(row=row, column=suggestion_col).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if status_col:
                cell = worksheet.cell(row=row, column=status_col)
                colours = {
                    "需要补货": ("FDE8E7", "B42318"),
                    "库存积压": ("FFF4D6", "9A6700"),
                    "停售积压": ("FCE8E6", "A61B1B"),
                    "库存异常": ("F9D7D5", "8A1010"),
                    "正常": ("E7F6EE", "0B6B46"),
                    "停售清零": ("EEF2F6", "4B6275"),
                }
                if cell.value in colours:
                    fill, font = colours[cell.value]
                    cell.fill = PatternFill("solid", fgColor=fill)
                    cell.font = Font(name="微软雅黑", size=10, bold=True, color=font)
    elif worksheet.title in {"补货建议", "积压清单", "采购分析", "销售分析"}:
        _inventory_title_band(worksheet, end_column=max(14, worksheet.max_column))
        _set_report_widths(
            worksheet,
            {
                "A": 13,
                "B": 22,
                "C": 14,
                "D": 16,
                "E": 14,
                "F": 14,
                "G": 14,
                "H": 15,
                "I": 15,
                "J": 15,
                "K": 14,
                "L": 14,
                "M": 17,
                "N": 40,
            },
        )
        if worksheet.max_row > header_row:
            for cell in worksheet[header_row + 1]:
                cell.fill = PatternFill("solid", fgColor="FFF4D6" if worksheet.title == "积压清单" else "E7F6EE")
                cell.font = Font(
                    name="微软雅黑", size=10, bold=True, color="6B4F00" if worksheet.title == "积压清单" else "0B5D3B"
                )
        headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        advice_column = headers.get("管理建议")
        if advice_column:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 48
                worksheet.cell(row=row, column=advice_column).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
    elif worksheet.title in {"人工核验", "数据审计"}:
        _inventory_title_band(worksheet, end_column=max(12, worksheet.max_column))
        _set_report_widths(
            worksheet,
            {
                "A": 12,
                "B": 13,
                "C": 15,
                "D": 15,
                "E": 14,
                "F": 28,
                "G": 14,
                "H": 12,
                "I": 16,
                "J": 18,
                "K": 48,
                "L": 38,
            },
        )
        headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 32
            for header in ("原因", "说明"):
                column = headers.get(header)
                if column:
                    worksheet.cell(row=row, column=column).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
            if worksheet.title == "人工核验" or worksheet.cell(row=row, column=1).value == "明细":
                for cell in worksheet[row]:
                    cell.fill = PatternFill("solid", fgColor="FFF1F0")
    elif worksheet.title == "库存图表看板":
        _inventory_title_band(worksheet, end_column=19)
        worksheet.freeze_panes = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.zoomScale = 80
        _set_report_widths(
            worksheet,
            {
                "A": 14,
                "B": 16,
                "C": 16,
                "D": 18,
                "E": 18,
                "F": 16,
                "G": 13,
                "H": 3,
                "I": 10,
                "J": 3,
                "K": 11,
                "L": 11,
                "M": 11,
                "N": 11,
                "O": 11,
                "P": 11,
                "Q": 11,
                "R": 11,
                "S": 11,
            },
        )


def _add_inventory_management_charts(worksheet: Any, *, header_row: int) -> None:
    if worksheet.title != "库存图表看板":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    required = {"月份", "采购入库数量", "销售出库数量", "品类", "可售库存金额", "库存状态", "SKU数量"}
    if not required.issubset(headers):
        return

    month_end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
    if month_end > header_row:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "月度采购入库与销售出库（件）"
        chart.y_axis.title = "数量（件）"
        chart.x_axis.title = "月份"
        chart.style = 2
        chart.height = 7.2
        chart.width = 14.2
        chart.y_axis.scaling.min = 0
        chart.y_axis.numFmt = "#,##0"
        chart.add_data(
            Reference(
                worksheet,
                min_col=headers["采购入库数量"],
                max_col=headers["销售出库数量"],
                min_row=header_row,
                max_row=month_end,
            ),
            titles_from_data=True,
        )
        chart.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=month_end))
        colours = ("2F75B5", "00A389")
        series_titles = ("采购入库数量", "销售出库数量")
        for index, series in enumerate(chart.series):
            series.graphicalProperties.solidFill = colours[index % len(colours)]
            series.graphicalProperties.line.solidFill = colours[index % len(colours)]
            series.tx = SeriesLabel(v=series_titles[index % len(series_titles)])
        chart.legend.position = "b"
        worksheet.add_chart(chart, "A4")

    category_end = _last_data_row(worksheet, headers["品类"], header_row=header_row)
    if category_end > header_row:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "品类可销售库存金额（元）"
        chart.x_axis.title = "库存金额（元）"
        chart.style = 2
        chart.height = 7.2
        chart.width = 12.5
        chart.legend = None
        chart.x_axis.scaling.min = 0
        chart.x_axis.numFmt = "#,##0"
        chart.add_data(
            Reference(worksheet, min_col=headers["可售库存金额"], min_row=header_row, max_row=category_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["品类"], min_row=header_row + 1, max_row=category_end)
        )
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = "6B5FD2"
            chart.series[0].graphicalProperties.line.solidFill = "6B5FD2"
        worksheet.add_chart(chart, "K4")

    status_end = _last_data_row(worksheet, headers["库存状态"], header_row=header_row)
    if status_end > header_row:
        chart = DoughnutChart()
        chart.title = "库存状态 SKU 结构（%）"
        chart.style = 2
        chart.height = 7.2
        chart.width = 14.2
        chart.holeSize = 58
        chart.firstSliceAng = 270
        chart.varyColors = True
        chart.add_data(
            Reference(worksheet, min_col=headers["SKU数量"], min_row=header_row, max_row=status_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["库存状态"], min_row=header_row + 1, max_row=status_end)
        )
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = False
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showVal = False
        chart.legend.position = "r"
        if chart.series:
            colours = ("E26A45", "D99614", "00A389", "6B5FD2", "6B7C8F", "B42318")
            chart.series[0].dPt = [
                DataPoint(idx=index, spPr=GraphicalProperties(solidFill=colour))
                for index, colour in enumerate(colours[: status_end - header_row])
            ]
        worksheet.add_chart(chart, "A21")

    worksheet.merge_cells("K21:S22")
    worksheet["K21"] = "管理提示"
    worksheet["K21"].fill = PatternFill("solid", fgColor="17324D")
    worksheet["K21"].font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
    worksheet["K21"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(12, 20):
        worksheet.cell(row=21, column=column).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(row=22, column=column).fill = PatternFill("solid", fgColor="17324D")
    messages = (
        "红色补货项优先结合在途和交期下单",
        "积压项优先暂停采购、促销或跨仓调拨",
        "退货与待确认调整必须凭仓库单据人工确认",
    )
    for offset, message in enumerate(messages):
        row = 24 + offset * 3
        worksheet.merge_cells(start_row=row, start_column=11, end_row=row + 1, end_column=19)
        cell = worksheet.cell(row=row, column=11)
        cell.value = message
        cell.fill = PatternFill("solid", fgColor="F2F6FA")
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="17324D")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for column in range(12, 20):
            worksheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor="F2F6FA")
            worksheet.cell(row=row + 1, column=column).fill = PatternFill("solid", fgColor="F2F6FA")


def _hr_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _HR_REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet.cell(row=1, column=1).value = title
    worksheet.cell(row=2, column=1).value = subtitle
    for column_index in range(1, end_column + 1):
        title_cell = worksheet.cell(row=1, column=column_index)
        title_cell.fill = PatternFill("solid", fgColor="17324D")
        title_cell.font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        subtitle_cell = worksheet.cell(row=2, column=column_index)
        subtitle_cell.fill = PatternFill("solid", fgColor="EAF1F7")
        subtitle_cell.font = Font(name="微软雅黑", size=10, color="4B6275")
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _hr_number_formats(worksheet: Any, *, header_row: int) -> None:
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    currency = {"基本工资", "薪资调整金额", "预计薪资", "销售额", "预计在职薪资"}
    percent = {"出勤率", "目标完成率", "销售额/预计薪资", "平均出勤率", "平均目标完成率"}
    score = {"考勤得分", "绩效得分", "综合得分", "平均综合得分", "客户评分"}
    date_headers = {"入职日期"}
    for header, column in headers.items():
        if header in currency:
            fmt = "#,##0.00;[Red](#,##0.00);-"
        elif header in percent:
            fmt = "0.0%"
        elif header in score:
            fmt = "0.0"
        elif header in date_headers:
            fmt = "yyyy-mm-dd"
        else:
            continue
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = fmt


def _add_hr_cards(worksheet: Any, *, header_row: int) -> None:
    metrics = {
        str(worksheet.cell(row=row, column=1).value or ""): worksheet.cell(row=row, column=2).value
        for row in range(header_row + 1, worksheet.max_row + 1)
    }
    cards = (
        ("F4:H4", "F5:H7", "在职员工", metrics.get("在职员工数"), '0 "人"', "2F75B5"),
        ("J4:L4", "J5:L7", "预计在职薪资", metrics.get("预计在职薪资"), '#,##0 "元"', "6B5FD2"),
        ("N4:P4", "N5:P7", "平均综合得分", metrics.get("平均综合得分"), '0.0 "分"', "00A389"),
        ("F9:H9", "F10:H12", "综合表现第一", metrics.get("综合表现第一"), "@", "D99614"),
        ("J9:L9", "J10:L12", "优秀员工", metrics.get("表现优秀员工"), '0 "人"', "00A389"),
        ("N9:P9", "N10:P12", "重点关注", metrics.get("重点关注员工"), '0 "人"', "E26A45"),
    )
    for label_range, value_range, label, value, number_format, colour in cards:
        worksheet.merge_cells(label_range)
        worksheet.merge_cells(value_range)
        label_cell = worksheet[label_range.split(":")[0]]
        value_cell = worksheet[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=colour)
        label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.value = value if value not in (None, "") else "—"
        value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
        value_cell.font = Font(name="微软雅黑", size=18, bold=True, color="17324D")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
    for column in ("F", "G", "H", "J", "K", "L", "N", "O", "P"):
        worksheet.column_dimensions[column].width = 9.5
    worksheet.column_dimensions["I"].width = 2.5
    worksheet.column_dimensions["M"].width = 2.5


def _style_hr_management_sheet(worksheet: Any, *, header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _HR_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 88
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)
    _hr_number_formats(worksheet, header_row=header_row)
    _hr_title_band(worksheet, end_column=max(16, worksheet.max_column))

    if worksheet.title == "管理层人效总览":
        _set_report_widths(worksheet, {"A": 24, "B": 22, "C": 10, "D": 58})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row, column=1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row=row, column=1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
            metric = str(worksheet.cell(row=row, column=1).value or "")
            if "薪资" in metric:
                worksheet.cell(row=row, column=2).number_format = "#,##0.00;[Red](#,##0.00);-"
            elif "率" in metric:
                worksheet.cell(row=row, column=2).number_format = "0.0%"
            elif "得分" in metric:
                worksheet.cell(row=row, column=2).number_format = "0.0"
        _add_hr_cards(worksheet, header_row=header_row)
    elif worksheet.title == "员工综合分析":
        widths = {
            "A": 12,
            "B": 11,
            "C": 13,
            "D": 15,
            "E": 11,
            "F": 11,
            "G": 11,
            "H": 11,
            "I": 11,
            "J": 11,
            "K": 11,
            "L": 14,
            "M": 14,
            "N": 12,
            "O": 12,
            "P": 14,
            "Q": 15,
            "R": 15,
            "S": 12,
            "T": 13,
            "U": 14,
            "V": 17,
            "W": 34,
            "X": 46,
        }
        _set_report_widths(worksheet, widths)
        headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for row in range(header_row + 1, worksheet.max_row + 1):
            for header in ("事实依据", "建议动作"):
                column = headers.get(header)
                if column:
                    worksheet.cell(row=row, column=column).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
            worksheet.row_dimensions[row].height = 42
            status = worksheet.cell(row=row, column=headers.get("管理分类", 1))
            if status.value == "表现优秀":
                status.fill = PatternFill("solid", fgColor="E7F6EE")
                status.font = Font(name="微软雅黑", size=10, bold=True, color="0B6B46")
            elif status.value == "重点关注":
                status.fill = PatternFill("solid", fgColor="FFF1F0")
                status.font = Font(name="微软雅黑", size=10, bold=True, color="B42318")
    elif worksheet.title in {"表现优秀员工", "重点关注员工", "人工核验"}:
        _set_report_widths(
            worksheet,
            {
                "A": 12,
                "B": 11,
                "C": 13,
                "D": 15,
                "E": 13,
                "F": 13,
                "G": 12,
                "H": 12,
                "I": 12,
                "J": 13,
                "K": 18,
                "L": 18,
                "M": 36,
                "N": 48,
            },
        )
        headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 44
            for header in ("事实依据", "建议动作"):
                column = headers.get(header)
                if column:
                    worksheet.cell(row=row, column=column).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
            fill = "E7F6EE" if worksheet.title == "表现优秀员工" else "FFF1F0"
            for cell in worksheet[row]:
                if cell.fill.fill_type is None:
                    cell.fill = PatternFill("solid", fgColor=fill)
    elif worksheet.title in {"考勤分析", "绩效分析", "薪资分析"}:
        for column_index in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 14
        worksheet.column_dimensions["B"].width = 11
        worksheet.column_dimensions["C"].width = 13
        worksheet.column_dimensions["D"].width = 15
        if worksheet.title == "薪资分析" and worksheet.max_column >= 10:
            worksheet.column_dimensions["J"].width = 26
    elif worksheet.title == "数据审计":
        _set_report_widths(worksheet, {"A": 24, "B": 46, "C": 14, "D": 72})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 44
            worksheet.cell(row=row, column=2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            worksheet.cell(row=row, column=4).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
    elif worksheet.title == "人力图表看板":
        worksheet.freeze_panes = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.zoomScale = 80
        for column in range(1, 20):
            worksheet.column_dimensions[get_column_letter(column)].width = 11
        worksheet.column_dimensions["G"].width = 17
        worksheet.column_dimensions["I"].width = 15
        worksheet.column_dimensions["J"].width = 15
        worksheet.column_dimensions["H"].width = 3
        worksheet.column_dimensions["J"].width = 3


def _add_hr_management_charts(worksheet: Any, *, header_row: int) -> None:
    if worksheet.title != "人力图表看板":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    required = {"部门", "预计薪资", "姓名", "综合得分", "离职风险代理等级", "风险人数", "考勤员工", "考勤异常次数"}
    if not required.issubset(headers):
        return

    department_end = _last_data_row(worksheet, headers["部门"], header_row=header_row)
    if department_end > header_row:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "部门预计薪资（元）"
        chart.x_axis.title = "预计薪资（元)"
        chart.height = 7.2
        chart.width = 13.8
        chart.style = 2
        chart.legend = None
        chart.x_axis.scaling.min = 0
        chart.x_axis.numFmt = "#,##0"
        chart.add_data(
            Reference(worksheet, min_col=headers["预计薪资"], min_row=header_row, max_row=department_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["部门"], min_row=header_row + 1, max_row=department_end)
        )
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = "2F75B5"
            chart.series[0].graphicalProperties.line.solidFill = "2F75B5"
        worksheet.add_chart(chart, "A4")

    employee_end = _last_data_row(worksheet, headers["姓名"], header_row=header_row)
    if employee_end > header_row:
        chart = BarChart()
        chart.type = "col"
        chart.title = "在职员工综合得分"
        chart.y_axis.title = "综合得分（分）"
        chart.x_axis.title = "员工"
        chart.height = 7.2
        chart.width = 13.8
        chart.style = 2
        chart.legend = None
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.y_axis.numFmt = "0"
        chart.add_data(
            Reference(worksheet, min_col=headers["综合得分"], min_row=header_row, max_row=employee_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["姓名"], min_row=header_row + 1, max_row=employee_end)
        )
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = "00A389"
            chart.series[0].graphicalProperties.line.solidFill = "00A389"
        worksheet.add_chart(chart, "K4")

    risk_end = _last_data_row(worksheet, headers["离职风险代理等级"], header_row=header_row)
    if risk_end > header_row:
        chart = DoughnutChart()
        chart.title = "风险代理等级结构（%）"
        chart.height = 7.2
        chart.width = 13.8
        chart.style = 2
        chart.holeSize = 58
        chart.firstSliceAng = 270
        chart.varyColors = True
        chart.add_data(
            Reference(worksheet, min_col=headers["风险人数"], min_row=header_row, max_row=risk_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["离职风险代理等级"], min_row=header_row + 1, max_row=risk_end)
        )
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = False
        chart.dataLabels.showLegendKey = False
        chart.legend.position = "r"
        if chart.series:
            colours = ("E26A45", "D99614", "00A389", "6B7C8F")
            chart.series[0].dPt = [
                DataPoint(idx=index, spPr=GraphicalProperties(solidFill=colour))
                for index, colour in enumerate(colours[: risk_end - header_row])
            ]
        worksheet.add_chart(chart, "A21")

    attendance_end = _last_data_row(worksheet, headers["考勤员工"], header_row=header_row)
    if attendance_end > header_row:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "迟到与早退次数（次）"
        chart.x_axis.title = "异常次数"
        chart.height = 7.2
        chart.width = 13.8
        chart.style = 2
        chart.legend = None
        chart.x_axis.scaling.min = 0
        chart.x_axis.numFmt = "0"
        chart.add_data(
            Reference(worksheet, min_col=headers["考勤异常次数"], min_row=header_row, max_row=attendance_end),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=headers["考勤员工"], min_row=header_row + 1, max_row=attendance_end)
        )
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = "E26A45"
            chart.series[0].graphicalProperties.line.solidFill = "E26A45"
        worksheet.add_chart(chart, "K21")


def _adaptive_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _ADAPTIVE_REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet.cell(1, 1).value = title
    worksheet.cell(1, 1).font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
    worksheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    worksheet.cell(2, 1).value = subtitle
    worksheet.cell(2, 1).font = Font(name="微软雅黑", size=10, color="4B6275")
    worksheet.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, end_column + 1):
        worksheet.cell(1, column).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(2, column).fill = PatternFill("solid", fgColor="EAF1F7")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _style_adaptive_management_sheet(worksheet: Any, *, header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _ADAPTIVE_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 88
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)
    _adaptive_title_band(worksheet, end_column=max(16, worksheet.max_column))
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}

    if worksheet.title == "管理层通用总览":
        _set_report_widths(worksheet, {"A": 25, "B": 25, "C": 11, "D": 68})
        metrics = {}
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row, 1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row, 1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
            metrics[str(worksheet.cell(row, 1).value or "")] = worksheet.cell(row, 2).value
            unit = str(worksheet.cell(row, 3).value or "")
            if unit == "%":
                worksheet.cell(row, 2).number_format = "0.0%"
            elif unit == "倍":
                worksheet.cell(row, 2).number_format = '0.00"x"'
            elif unit == "元":
                worksheet.cell(row, 2).number_format = '#,##0 "元";[Red](#,##0) "元";-'
            elif unit in {"单", "行", "个", "条"}:
                worksheet.cell(row, 2).number_format = "#,##0"
        cards = (
            ("F4:H4", "F5:H7", "主数据记录", metrics.get("主数据记录数"), "#,##0", "2F75B5"),
            ("J4:L4", "J5:L7", "识别指标", metrics.get("识别数值指标"), "#,##0", "00A389"),
            ("N4:P4", "N5:P7", "识别维度", metrics.get("识别分类维度"), "#,##0", "6B5FD2"),
            ("F9:H9", "F10:H12", "候选关系", metrics.get("建议表关系"), "#,##0", "D99614"),
            ("J9:L9", "J10:L12", "异常线索", metrics.get("检测异常记录"), "#,##0", "E26A45"),
            ("N9:P9", "N10:P12", "主分析表", metrics.get("主分析表"), "@", "17324D"),
        )
        for label_range, value_range, label, value, number_format, colour in cards:
            worksheet.merge_cells(label_range)
            worksheet.merge_cells(value_range)
            label_cell = worksheet[label_range.split(":")[0]]
            value_cell = worksheet[value_range.split(":")[0]]
            label_cell.value = label
            label_cell.fill = PatternFill("solid", fgColor=colour)
            label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.value = value if value not in (None, "") else "—"
            value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
            value_cell.font = Font(name="微软雅黑", size=16, bold=True, color="17324D")
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.number_format = number_format
        for column in ("F", "G", "H", "J", "K", "L", "N", "O", "P"):
            worksheet.column_dimensions[column].width = 10
    elif worksheet.title == "自适应图表看板":
        worksheet.freeze_panes = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.zoomScale = 78
        for column in range(1, max(20, worksheet.max_column + 1)):
            worksheet.column_dimensions[get_column_letter(column)].width = 11
    else:
        widths = {
            "A": 18,
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 16,
            "F": 16,
            "G": 17,
            "H": 16,
            "I": 18,
            "J": 28,
            "K": 34,
            "L": 42,
        }
        _set_report_widths(worksheet, widths)
        for header, column in headers.items():
            if any(token in header for token in ("率", "占比", "唯一性", "覆盖率")):
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "0.0%"
            elif any(token in header for token in ("日期", "时间")):
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "yyyy-mm-dd"
            elif any(
                token in header
                for token in ("金额", "销售", "收入", "成本", "利润", "工资", "薪资", "指标值", "异常值")
            ):
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "#,##0.00;[Red](#,##0.00);-"
        if worksheet.title in {"数据字典", "表关系建议", "异常数据"}:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 34
                for cell in worksheet[row]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if worksheet.title == "异常数据":
            for row in range(header_row + 1, worksheet.max_row + 1):
                for cell in worksheet[row]:
                    if cell.fill.fill_type is None:
                        cell.fill = PatternFill("solid", fgColor="FFF1F0")


def _add_adaptive_management_charts(worksheet: Any, *, header_row: int) -> None:
    if worksheet.title != "自适应图表看板":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}

    if {"排名分类", "排名指标值"}.issubset(headers):
        end = _last_data_row(worksheet, headers["排名分类"], header_row=header_row)
        if end > header_row:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "首要分类指标排名"
            chart.x_axis.title = "指标值"
            chart.y_axis.title = "分类"
            chart.height = 7.2
            chart.width = 13.8
            chart.style = 2
            chart.legend = None
            chart.add_data(
                Reference(worksheet, min_col=headers["排名指标值"], min_row=header_row, max_row=end),
                titles_from_data=True,
            )
            chart.set_categories(Reference(worksheet, min_col=headers["排名分类"], min_row=header_row + 1, max_row=end))
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = "2F75B5"
                chart.series[0].graphicalProperties.line.solidFill = "2F75B5"
            worksheet.add_chart(chart, "A4")

    trend_columns = [name for name in headers if name.startswith("趋势_")]
    if "月份" in headers and trend_columns:
        end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
        if end > header_row:
            chart = LineChart()
            chart.title = "核心指标时间趋势"
            chart.y_axis.title = "指标值"
            chart.x_axis.title = "月份"
            chart.height = 7.2
            chart.width = 13.8
            chart.style = 13
            for index, name in enumerate(trend_columns[:3]):
                chart.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=end), titles_from_data=True
                )
                series = chart.series[-1]
                colour = ("00A389", "6B5FD2", "D99614")[index]
                series.graphicalProperties.line.solidFill = colour
                series.graphicalProperties.line.width = 26000
            chart.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=end))
            chart.legend.position = "b"
            worksheet.add_chart(chart, "K4")

    if {"结构分类", "结构指标值"}.issubset(headers):
        end = _last_data_row(worksheet, headers["结构分类"], header_row=header_row)
        if end > header_row:
            chart = DoughnutChart()
            chart.title = "首要指标结构占比"
            chart.height = 7.2
            chart.width = 13.8
            chart.style = 10
            chart.holeSize = 58
            chart.add_data(
                Reference(worksheet, min_col=headers["结构指标值"], min_row=header_row, max_row=end),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(worksheet, min_col=headers["结构分类"], min_row=header_row + 1, max_row=end)
            )
            chart.legend.position = "r"
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showLeaderLines = True
            worksheet.add_chart(chart, "A21")

    if {"异常类型", "风险数量"}.issubset(headers):
        end = _last_data_row(worksheet, headers["异常类型"], header_row=header_row)
        if end > header_row:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "风险与异常线索分布"
            chart.x_axis.title = "线索数量"
            chart.y_axis.title = "异常类型"
            chart.height = 7.2
            chart.width = 13.8
            chart.style = 2
            chart.legend = None
            chart.add_data(
                Reference(worksheet, min_col=headers["风险数量"], min_row=header_row, max_row=end),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(worksheet, min_col=headers["异常类型"], min_row=header_row + 1, max_row=end)
            )
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = "E26A45"
                chart.series[0].graphicalProperties.line.solidFill = "E26A45"
            worksheet.add_chart(chart, "K21")


def _selection_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _SELECTION_REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet.cell(1, 1).value = title
    worksheet.cell(2, 1).value = subtitle
    for column in range(1, end_column + 1):
        worksheet.cell(1, column).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(1, column).font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
        worksheet.cell(1, column).alignment = Alignment(horizontal="left", vertical="center")
        worksheet.cell(2, column).fill = PatternFill("solid", fgColor="EAF1F7")
        worksheet.cell(2, column).font = Font(name="微软雅黑", size=10, color="4B6275")
        worksheet.cell(2, column).alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _style_selection_management_sheet(worksheet: Any, *, header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _SELECTION_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 88
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)
    _selection_title_band(worksheet, end_column=max(16, worksheet.max_column))
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}

    for header in ("综合推荐分", "基础表现分", "有效平均分", "最新得分", "得分趋势", "风险扣分", "正向加分"):
        column = headers.get(header)
        if column:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row, column).number_format = "0.00"
    for header in ("得分完整率",):
        column = headers.get(header)
        if column:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row, column).number_format = "0.0%"

    if worksheet.title == "评选管理总览":
        _set_report_widths(worksheet, {"A": 24, "B": 42, "C": 10, "D": 70})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row, 1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row, 1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
    elif worksheet.title in {"建议入选名单", "全部候选排序"}:
        for column in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = 14
        for header in ("风险提示", "推荐理由", "评语摘要", "正向依据"):
            column = headers.get(header)
            if column:
                worksheet.column_dimensions[get_column_letter(column)].width = 42 if header != "评语摘要" else 68
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    worksheet.row_dimensions[row].height = max(worksheet.row_dimensions[row].height or 0, 42)
        status_column = headers.get("入选状态")
        if status_column:
            for row in range(header_row + 1, worksheet.max_row + 1):
                cell = worksheet.cell(row, status_column)
                if cell.value == "建议入选":
                    cell.fill = PatternFill("solid", fgColor="E7F6EE")
                    cell.font = Font(name="微软雅黑", size=10, bold=True, color="0B6B46")
        score_column = headers.get("综合推荐分")
        if score_column and worksheet.max_row > header_row:
            address = f"{get_column_letter(score_column)}{header_row + 1}:{get_column_letter(score_column)}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                address, DataBarRule(start_type="min", end_type="max", color="2F75B5", showValue=True)
            )
    elif worksheet.title == "风险复核清单":
        _set_report_widths(worksheet, {"A": 12, "B": 16, "C": 14, "D": 12, "E": 42, "F": 72, "G": 10})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 48
            for header in ("风险提示", "评语摘要"):
                column = headers.get(header)
                if column:
                    worksheet.cell(row, column).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
            level = worksheet.cell(row, headers.get("风险等级", 1))
            if level.value == "高":
                level.fill = PatternFill("solid", fgColor="FFF1F0")
                level.font = Font(name="微软雅黑", bold=True, color="B42318")
            elif level.value == "中":
                level.fill = PatternFill("solid", fgColor="FFF4D6")
                level.font = Font(name="微软雅黑", bold=True, color="9A6700")
    elif worksheet.title == "评选规则与字段":
        _set_report_widths(worksheet, {"A": 16, "B": 24, "C": 96})
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 34
            worksheet.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    elif worksheet.title == "评选图表看板":
        worksheet.freeze_panes = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.zoomScale = 80
        for column in range(1, max(worksheet.max_column, 19) + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = 12


def _add_selection_management_charts(worksheet: Any, *, header_row: int) -> None:
    if worksheet.title != "评选图表看板":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    metric_headers = {"综合推荐分", "有效平均分", "最新得分", "风险扣分", "入选状态"}
    candidate_headers = [header for header in headers if header and header not in metric_headers]
    if not candidate_headers or "综合推荐分" not in headers:
        return
    candidate = candidate_headers[0]
    end = _last_data_row(worksheet, headers[candidate], header_row=header_row)
    if end <= header_row:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "候选综合推荐分"
    chart.y_axis.title = "综合推荐分"
    chart.x_axis.title = "候选"
    chart.height = 8
    chart.width = 14
    chart.style = 2
    chart.legend = None
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.y_axis.numFmt = "0"
    chart.add_data(
        Reference(worksheet, min_col=headers["综合推荐分"], min_row=header_row, max_row=end), titles_from_data=True
    )
    chart.set_categories(Reference(worksheet, min_col=headers[candidate], min_row=header_row + 1, max_row=end))
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "00A389"
        chart.series[0].graphicalProperties.line.solidFill = "00A389"
    worksheet.add_chart(chart, "A4")

    if {"有效平均分", "最新得分"}.issubset(headers):
        compare = BarChart()
        compare.type = "col"
        compare.title = "平均表现与最新表现"
        compare.y_axis.title = "得分"
        compare.x_axis.title = "候选"
        compare.height = 8
        compare.width = 14
        compare.style = 2
        compare.y_axis.scaling.min = 0
        compare.y_axis.scaling.max = 100
        compare.y_axis.numFmt = "0"
        compare.add_data(
            Reference(
                worksheet, min_col=headers["有效平均分"], max_col=headers["最新得分"], min_row=header_row, max_row=end
            ),
            titles_from_data=True,
        )
        compare.set_categories(Reference(worksheet, min_col=headers[candidate], min_row=header_row + 1, max_row=end))
        if len(compare.series) >= 2:
            compare.series[0].graphicalProperties.solidFill = "2F75B5"
            compare.series[1].graphicalProperties.solidFill = "D99614"
        worksheet.add_chart(compare, "K4")


def _enterprise_title_band(worksheet: Any, *, end_column: int) -> None:
    title, subtitle = _ENTERPRISE_REPORT_TITLES[worksheet.title]
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet.cell(1, 1).value = title
    worksheet.cell(2, 1).value = subtitle
    for column in range(1, end_column + 1):
        top = worksheet.cell(1, column)
        sub = worksheet.cell(2, column)
        top.fill = PatternFill("solid", fgColor="17324D")
        top.font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
        top.alignment = Alignment(horizontal="left", vertical="center")
        sub.fill = PatternFill("solid", fgColor="EAF1F7")
        sub.font = Font(name="微软雅黑", size=10, color="4B6275")
        sub.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 9


def _style_enterprise_management_sheet(worksheet: Any, *, header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _ENTERPRISE_TAB_COLOURS[worksheet.title]
    worksheet.sheet_view.zoomScale = 88
    worksheet.print_title_rows = f"1:{header_row}"
    _style_report_data_grid(worksheet, header_row=header_row)
    title_end_column = 14 if worksheet.title == "经营诊断看板" else max(18, worksheet.max_column)
    _enterprise_title_band(worksheet, end_column=title_end_column)
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}

    if worksheet.title == "管理层诊断总览":
        _set_report_widths(worksheet, {"A": 28, "B": 24, "C": 10, "D": 58})
        metrics = {}
        for row in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row, 1).fill = _SUMMARY_KEY_FILL
            worksheet.cell(row, 1).font = Font(name="微软雅黑", size=10, bold=True, color="0B5D3B")
            metrics[str(worksheet.cell(row, 1).value or "")] = worksheet.cell(row, 2).value
            unit = str(worksheet.cell(row, 3).value or "")
            if unit == "%":
                worksheet.cell(row, 2).number_format = "0.0%"
                worksheet.cell(row, 3).value = None
            elif unit == "倍":
                worksheet.cell(row, 2).number_format = '0.00"x"'
                worksheet.cell(row, 3).value = None
            elif unit == "元":
                worksheet.cell(row, 2).number_format = '#,##0 "元";[Red](#,##0) "元";-'
                worksheet.cell(row, 3).value = None
            elif unit in {"单", "行", "个", "条"}:
                worksheet.cell(row, 2).number_format = "#,##0"
        ecommerce = "买家实付" in metrics and "广告花费" in metrics
        compact_restaurant = "季度净营业收入" in metrics and "管理利润" in metrics
        if ecommerce:
            worksheet["A1"] = "多平台电商经营诊断驾驶舱"
            worksheet["A2"] = "订单、退款、平台结算、标准成本、广告、采购、库存和客户的证据化经营诊断"
        elif compact_restaurant:
            worksheet["A1"] = "连锁餐饮经营诊断驾驶舱"
            worksheet["A2"] = "门店与月份加权经营口径｜规模、利润、成本结构、风险和行动均可追溯"
        gross_label = next(
            (label for label in ("绩效口径毛利", "流水口径毛利", "毛利") if label in metrics),
            "参考毛利",
        )
        margin_label = next(
            (label for label in ("绩效口径毛利率", "流水口径毛利率", "整体毛利率") if label in metrics),
            "参考毛利率",
        )
        cards = (
            (
                ("F4:H4", "F5:H7", "买家实付", metrics.get("买家实付"), '#,##0 "元"', "2F75B5"),
                ("J4:L4", "J5:L7", "实际到账", metrics.get("实际到账"), '#,##0 "元"', "008C72"),
                ("N4:P4", "N5:P7", "趋势性管理贡献", metrics.get("趋势性管理贡献"), '#,##0 "元"', "E26A45"),
                ("F9:H9", "F10:H12", "退款后商品毛利率", metrics.get("管理口径商品毛利率"), "0.0%", "6B5FD2"),
                ("J9:L9", "J10:L12", "整体ROAS", metrics.get("整体ROAS"), '0.00"x"', "D99614"),
                ("N9:P9", "N10:P12", "期末库存金额", metrics.get("期末库存金额"), '#,##0 "元"', "B42318"),
            )
            if ecommerce
            else (
                ("F4:H4", "F5:H7", "净营业收入", metrics.get("季度净营业收入"), '#,##0 "元"', "2F75B5"),
                ("J4:L4", "J5:L7", "管理利润", metrics.get("管理利润"), '#,##0 "元"', "00A389"),
                ("N4:P4", "N5:P7", "管理利润率", metrics.get("管理利润率"), "0.0%", "6B5FD2"),
                ("F9:H9", "F10:H12", "平台费", metrics.get("平台费"), '#,##0 "元"', "D99614"),
                ("J9:L9", "J10:L12", "最佳经营门店", metrics.get("最佳经营门店"), "@", "008C72"),
                ("N9:P9", "N10:P12", "重点关注门店", metrics.get("重点关注门店"), "@", "B42318"),
            )
            if compact_restaurant
            else (
                ("F4:H4", "F5:H7", "原始订单金额", metrics.get("原始订单金额"), '#,##0 "元"', "2F75B5"),
                ("J4:L4", "J5:L7", gross_label, metrics.get(gross_label), '#,##0 "元"', "00A389"),
                ("N4:P4", "N5:P7", "估算经营利润", metrics.get("估算经营利润"), '#,##0 "元"', "E26A45"),
                ("F9:H9", "F10:H12", margin_label, metrics.get(margin_label), "0.0%", "6B5FD2"),
                ("J9:L9", "J10:L12", "回款率", metrics.get("回款率"), "0.0%", "D99614"),
                ("N9:P9", "N10:P12", "前三客户集中度", metrics.get("前三客户收入集中度"), "0.0%", "B42318"),
            )
        )
        for label_range, value_range, label, value, fmt, colour in cards:
            worksheet.merge_cells(label_range)
            worksheet.merge_cells(value_range)
            label_cell = worksheet[label_range.split(":")[0]]
            value_cell = worksheet[value_range.split(":")[0]]
            label_cell.value = label
            label_cell.fill = PatternFill("solid", fgColor=colour)
            label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.value = value if value not in (None, "") else "—"
            value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
            value_cell.font = Font(name="微软雅黑", size=17, bold=True, color="17324D")
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.number_format = fmt
        for column in ("F", "G", "H", "J", "K", "L", "N", "O", "P"):
            worksheet.column_dimensions[column].width = 10
    elif worksheet.title == "经营诊断看板":
        worksheet.freeze_panes = None
        worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        worksheet.sheet_view.topLeftCell = "A1"
        worksheet.sheet_view.zoomScale = 90
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_area = "A1:N58"
        for column in range(1, 15):
            worksheet.column_dimensions[get_column_letter(column)].width = 10.5
        worksheet.column_dimensions["E"].width = 2.5
        worksheet.column_dimensions["J"].width = 2.5
        for column in range(15, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column)].hidden = True
        for row in range(header_row, worksheet.max_row + 1):
            worksheet.row_dimensions[row].hidden = True
            worksheet.row_dimensions[row].outlineLevel = 1

        def dashboard_value(*names: str) -> Any:
            for name in names:
                column = headers.get(name)
                if column:
                    return worksheet.cell(header_row + 1, column).value
            return None

        restaurant = "门店_门店" in headers
        compact_restaurant = {"月度_月份", "月度_营业额", "月度_管理利润", "月度_管理利润率", "门店_营业额", "门店_管理利润", "门店_管理利润率"}.issubset(headers)
        ecommerce = "渠道_渠道" in headers and not restaurant
        if restaurant:
            worksheet["A1"] = "餐饮门店经营驾驶舱"
            worksheet["A2"] = (
                "门店与月度经营规模、管理利润、成本结构和行动优先级｜比例指标全部使用加权口径"
                if compact_restaurant
                else "销售、退款、外卖到账、菜品成本、人工、固定费用与损耗线索｜尺度不匹配时只提示人工核验"
            )
        if ecommerce:
            worksheet["A1"] = "多平台电商经营驾驶舱"
            worksheet["A2"] = "增长质量、现金转化、渠道真实盈利、广告效率和库存占用｜原生 Excel 图表可编辑"
        kpi_cards = (
            (1, 4, "成交实付" if ecommerce else ("净营业收入" if compact_restaurant else ("营业实付" if restaurant else "销售规模")), dashboard_value("KPI_销售规模"), '#,##0 "元"', "2F75B5"),
            (6, 9, "现金转化率" if ecommerce else ("管理利润率" if compact_restaurant else ("平台到账率" if restaurant else "回款率")), dashboard_value("KPI_管理利润率", "KPI_平台到账率", "KPI_回款率"), "0.0%", "008C72"),
            (
                11,
                14,
                str(dashboard_value("KPI_毛利标题") or "参考毛利率"),
                dashboard_value("KPI_毛利率"),
                "0.0%",
                "6B5FD2",
            ),
            (1, 4, "趋势性管理贡献" if ecommerce else ("管理利润" if compact_restaurant else ("情景经营结果（待核验）" if restaurant else "估算经营结果")), dashboard_value("KPI_估算经营结果"), '#,##0 "元"', "E26A45"),
            (6, 9, "已发生退款" if ecommerce else ("已发生退款" if restaurant else "风险订单金额"), dashboard_value("KPI_已发生退款", "KPI_风险订单"), '#,##0 "元"', "B42318"),
            (11, 14, "期末库存金额" if ecommerce else ("平台费" if compact_restaurant else ("报损金额线索" if restaurant else "库存金额")), dashboard_value("KPI_平台费", "KPI_报损金额", "KPI_库存金额"), '#,##0 "元"', "D99614"),
        )
        for index, (start_col, end_col, label, value, fmt, colour) in enumerate(kpi_cards):
            label_row, value_start, value_end = (4, 5, 7) if index < 3 else (9, 10, 12)
            worksheet.merge_cells(
                start_row=label_row,
                start_column=start_col,
                end_row=label_row,
                end_column=end_col,
            )
            worksheet.merge_cells(
                start_row=value_start,
                start_column=start_col,
                end_row=value_end,
                end_column=end_col,
            )
            label_cell = worksheet.cell(label_row, start_col)
            value_cell = worksheet.cell(value_start, start_col)
            label_cell.value = label
            label_cell.fill = PatternFill("solid", fgColor=colour)
            label_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.value = value if value not in (None, "") else "—"
            value_cell.number_format = fmt
            value_cell.fill = PatternFill("solid", fgColor="F5F8FB")
            value_cell.font = Font(name="微软雅黑", size=18, bold=True, color="17324D")
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(value_start, value_end + 1):
                for column in range(start_col, end_col + 1):
                    worksheet.cell(row, column).fill = PatternFill("solid", fgColor="F5F8FB")

        worksheet.merge_cells("A14:N14")
        worksheet.merge_cells("A15:N17")
        worksheet["A14"] = "核心经营诊断"
        worksheet["A14"].fill = PatternFill("solid", fgColor="17324D")
        worksheet["A14"].font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        worksheet["A14"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet["A15"] = dashboard_value("核心诊断") or "当前数据尚不足以形成确定性经营诊断。"
        worksheet["A15"].fill = PatternFill("solid", fgColor="EEF4F8")
        worksheet["A15"].font = Font(name="微软雅黑", size=12, bold=True, color="17324D")
        worksheet["A15"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for index, (start_col, end_col) in enumerate(((1, 4), (6, 9), (11, 14)), start=1):
            worksheet.merge_cells(start_row=19, start_column=start_col, end_row=19, end_column=end_col)
            worksheet.merge_cells(start_row=20, start_column=start_col, end_row=24, end_column=end_col)
            title = dashboard_value(f"风险卡{index}_标题") or "暂无更多高优先级风险"
            evidence = dashboard_value(f"风险卡{index}_证据") or "未触发"
            action = dashboard_value(f"风险卡{index}_行动") or "持续监控"
            title_cell = worksheet.cell(19, start_col)
            body_cell = worksheet.cell(20, start_col)
            title_cell.value = title
            title_cell.fill = PatternFill("solid", fgColor="B42318" if str(title).startswith("P0") else "D99614")
            title_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            body_cell.value = f"证据：{evidence}\n行动：{action}"
            body_cell.fill = PatternFill("solid", fgColor="FFF8F0")
            body_cell.font = Font(name="微软雅黑", size=9, color="3E4C59")
            body_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.row_dimensions[14].height = 24
        worksheet.row_dimensions[15].height = 30
        worksheet.row_dimensions[16].height = 24
        worksheet.row_dimensions[17].height = 24
        worksheet.row_dimensions[19].height = 24
        for row in range(20, 25):
            worksheet.row_dimensions[row].height = 24
    else:
        if worksheet.title == "门店经营诊断" and {"净营业收入", "管理利润", "管理利润率"}.issubset(headers):
            worksheet["A2"] = "门店规模、加权管理利润率、成本率和利润转化诊断｜收入最高不等于经营最好"
        elif worksheet.title == "利润驱动分析" and {"净营业收入", "管理利润", "管理利润率"}.issubset(headers):
            worksheet["A2"] = "营业额、退款、净营业收入、成本费用和加权利润率的月度桥接｜识别增收不增利"
        elif worksheet.title == "成本费用分析" and {"成本费用项目", "占净营业收入"}.issubset(headers):
            worksheet["A1"] = "餐饮成本费用结构"
            worksheet["A2"] = "食材、人工、平台、租金、水电与营销占净营业收入比重｜金额统一求和"
        for column in range(1, worksheet.max_column + 1):
            header = str(worksheet.cell(header_row, column).value or "")
            letter = get_column_letter(column)
            worksheet.column_dimensions[letter].width = (
                42
                if any(token in header for token in ("证据", "动作", "建议", "核验", "风险说明", "管理诊断", "边界"))
                else 18
            )
            if header == "ROAS":
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = '0.00"x"'
            elif "库存覆盖月数" in header:
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = '0.0"月"'
            elif any(token in header for token in ("率", "占比", "完成率", "偏差", "变化")):
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "0.0%"
            elif any(token in header for token in ("收入", "成本", "毛利", "利润", "金额", "费用", "风险敞口")):
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "#,##0.00;[Red](#,##0.00);-"
            elif "日期" in header:
                for row in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row, column).number_format = "yyyy-mm-dd"
        if worksheet.title in {"风险行动计划", "数据口径与验收"}:
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 48
                for cell in worksheet[row]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if worksheet.title in {"客户与回款风险", "库存风险分析", "风险行动计划"}:
            for row in range(header_row + 1, worksheet.max_row + 1):
                values = [str(worksheet.cell(row, col).value or "") for col in range(1, worksheet.max_column + 1)]
                if any(value in {"高", "P0"} or "偏高线索" in value or "缺货线索" in value for value in values):
                    for cell in worksheet[row]:
                        if cell.fill.fill_type is None:
                            cell.fill = PatternFill("solid", fgColor="FFF1F0")
                elif any(value in {"中", "P1", "未知/待核验"} for value in values):
                    for cell in worksheet[row]:
                        if cell.fill.fill_type is None:
                            cell.fill = PatternFill("solid", fgColor="FFF8E6")


def _style_source_confirmation_sheet(worksheet: Any, *, header_row: int) -> None:
    """Make the source audit readable before any business report styling."""
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = "0B6B46"
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(11, worksheet.max_column))
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(11, worksheet.max_column))
    worksheet.cell(1, 1).value = "数据源确认"
    worksheet.cell(2, 1).value = "本页只记录当前任务明确上传并纳入分析的原始工作表；历史输出文件会在上传阶段阻断。"
    for col in range(1, max(11, worksheet.max_column) + 1):
        worksheet.cell(1, col).fill = PatternFill("solid", fgColor="17324D")
        worksheet.cell(1, col).font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        worksheet.cell(2, col).fill = PatternFill("solid", fgColor="EAF1F7")
        worksheet.cell(2, col).font = Font(name="微软雅黑", size=10, color="4B6275")
    widths = [8, 18, 38, 28, 12, 10, 14, 14, 14, 18, 62]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(header_row, worksheet.max_row + 1):
        for cell in worksheet[row]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row == header_row:
            for cell in worksheet[row]:
                cell.fill = PatternFill("solid", fgColor="0B6B46")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        elif str(worksheet.cell(row, 9).value or "") == "通过":
            worksheet.cell(row, 9).fill = PatternFill("solid", fgColor="E6F4EA")
        elif str(worksheet.cell(row, 9).value or "") == "阻断":
            worksheet.cell(row, 9).fill = PatternFill("solid", fgColor="FDE8E7")
    worksheet.row_dimensions[1].height = 32
    worksheet.row_dimensions[2].height = 24
    for row in range(header_row + 1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 42


def _add_enterprise_management_charts(worksheet: Any, *, header_row: int) -> None:
    if worksheet.title != "经营诊断看板":
        return
    headers = {str(cell.value or ""): cell.column for cell in worksheet[header_row]}
    compact_restaurant_fields = {
        "月度_月份",
        "月度_营业额",
        "月度_管理利润",
        "月度_管理利润率",
        "门店_门店",
        "门店_营业额",
        "门店_管理利润",
        "门店_管理利润率",
    }
    if compact_restaurant_fields.issubset(headers):
        store_end = _last_data_row(worksheet, headers["门店_门店"], header_row=header_row)
        if store_end > header_row:
            store_amount = BarChart()
            store_amount.type = "col"
            store_amount.title = "门店营业额与管理利润（元）"
            store_amount.y_axis.title = "金额（元）"
            store_amount.x_axis.title = "门店"
            store_amount.height = 6.2
            store_amount.width = 11.2
            store_amount.style = 10
            store_amount.y_axis.numFmt = "#,##0"
            store_amount.y_axis.majorGridlines = None
            for name, label, colour in (
                ("门店_营业额", "营业额", "2F75B5"),
                ("门店_管理利润", "管理利润", "00A389"),
            ):
                store_amount.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=store_end),
                    titles_from_data=True,
                )
                store_amount.series[-1].tx = SeriesLabel(v=label)
                store_amount.series[-1].graphicalProperties.solidFill = colour
                store_amount.series[-1].graphicalProperties.line.solidFill = colour
            store_amount.set_categories(
                Reference(worksheet, min_col=headers["门店_门店"], min_row=header_row + 1, max_row=store_end)
            )
            store_amount.legend.position = "b"
            worksheet.add_chart(store_amount, "A27")

            store_margin = BarChart()
            store_margin.type = "col"
            store_margin.title = "门店管理利润率（加权口径）"
            store_margin.y_axis.title = "利润率"
            store_margin.x_axis.title = "门店"
            store_margin.height = 6.2
            store_margin.width = 11.2
            store_margin.style = 10
            store_margin.legend = None
            store_margin.y_axis.numFmt = "0.0%"
            store_margin.y_axis.majorGridlines = None
            store_margin.add_data(
                Reference(worksheet, min_col=headers["门店_管理利润率"], min_row=header_row, max_row=store_end),
                titles_from_data=True,
            )
            store_margin.set_categories(
                Reference(worksheet, min_col=headers["门店_门店"], min_row=header_row + 1, max_row=store_end)
            )
            if store_margin.series:
                store_margin.series[0].tx = SeriesLabel(v="管理利润率")
                store_margin.series[0].graphicalProperties.solidFill = "6B5FD2"
                store_margin.series[0].graphicalProperties.line.solidFill = "6B5FD2"
            worksheet.add_chart(store_margin, "H27")

        month_end = _last_data_row(worksheet, headers["月度_月份"], header_row=header_row)
        if month_end > header_row:
            monthly_amount = LineChart()
            monthly_amount.title = "月度营业额与管理利润趋势（元）"
            monthly_amount.y_axis.title = "金额（元）"
            monthly_amount.x_axis.title = "月份"
            monthly_amount.height = 6.2
            monthly_amount.width = 11.2
            monthly_amount.style = 13
            monthly_amount.y_axis.numFmt = "#,##0"
            monthly_amount.y_axis.majorGridlines = None
            for name, label, colour in (
                ("月度_营业额", "营业额", "2F75B5"),
                ("月度_管理利润", "管理利润", "00A389"),
            ):
                monthly_amount.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=month_end),
                    titles_from_data=True,
                )
                monthly_amount.series[-1].tx = SeriesLabel(v=label)
                _style_line_series(monthly_amount.series[-1], colour)
            monthly_amount.set_categories(
                Reference(worksheet, min_col=headers["月度_月份"], min_row=header_row + 1, max_row=month_end)
            )
            monthly_amount.legend.position = "b"
            worksheet.add_chart(monthly_amount, "A43")

            monthly_margin = LineChart()
            monthly_margin.title = "月度管理利润率趋势（加权口径）"
            monthly_margin.y_axis.title = "利润率"
            monthly_margin.x_axis.title = "月份"
            monthly_margin.height = 6.2
            monthly_margin.width = 11.2
            monthly_margin.style = 13
            monthly_margin.legend = None
            monthly_margin.y_axis.numFmt = "0.0%"
            monthly_margin.y_axis.majorGridlines = None
            monthly_margin.add_data(
                Reference(worksheet, min_col=headers["月度_管理利润率"], min_row=header_row, max_row=month_end),
                titles_from_data=True,
            )
            monthly_margin.set_categories(
                Reference(worksheet, min_col=headers["月度_月份"], min_row=header_row + 1, max_row=month_end)
            )
            if monthly_margin.series:
                monthly_margin.series[0].tx = SeriesLabel(v="管理利润率")
                _style_line_series(monthly_margin.series[0], "E26A45")
            worksheet.add_chart(monthly_margin, "H43")
        return
    restaurant_fields = {"月份", "营业实付", "标准食材成本", "人工成本", "固定费用", "门店_门店", "门店_可比经营贡献", "渠道_渠道", "渠道_可比经营贡献", "损耗_原料", "损耗_报损金额"}
    if restaurant_fields.issubset(headers):
        end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
        if end > header_row:
            monthly = LineChart()
            monthly.title = "月度营业实付与成本投入（元）"
            monthly.y_axis.title = "金额（元）"
            monthly.x_axis.title = "月份"
            monthly.height = 6.2; monthly.width = 11.2; monthly.style = 13
            for name, colour in (("营业实付", "2F75B5"), ("标准食材成本", "D99614"), ("人工成本", "E26A45"), ("固定费用", "6B7C8F")):
                monthly.add_data(Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=end), titles_from_data=True)
                _style_line_series(monthly.series[-1], colour)
            monthly.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=end))
            monthly.legend.position = "b"; worksheet.add_chart(monthly, "A27")
        for anchor, cat, value, title, colour in (("H27", "门店_门店", "门店_可比经营贡献", "门店可比经营贡献（元）", "2F75B5"), ("A43", "渠道_渠道", "渠道_可比经营贡献", "渠道可比经营贡献（元）", "008C72"), ("H43", "损耗_原料", "损耗_报损金额", "原料报损金额线索（元）", "B42318")):
            last = _last_data_row(worksheet, headers[cat], header_row=header_row)
            if last <= header_row: continue
            chart = BarChart(); chart.type = "bar"; chart.title = title; chart.x_axis.title = "金额（元）"; chart.height = 6.2; chart.width = 11.2; chart.style = 10; chart.legend = None
            chart.add_data(Reference(worksheet, min_col=headers[value], min_row=header_row, max_row=last), titles_from_data=True)
            chart.set_categories(Reference(worksheet, min_col=headers[cat], min_row=header_row + 1, max_row=last))
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = colour
                chart.series[0].graphicalProperties.line.solidFill = colour
            worksheet.add_chart(chart, anchor)
        return
    ecommerce_fields = {
        "月份",
        "成交实付",
        "实际到账",
        "标准成本",
        "广告费",
        "渠道_渠道",
        "渠道_管理贡献",
        "渠道_ROAS",
        "库存_产品",
        "库存_库存月数",
    }
    if ecommerce_fields.issubset(headers):
        month_end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
        if month_end > header_row:
            chart = LineChart()
            chart.title = "销售增长未转化为经营贡献（元）"
            chart.y_axis.title = "金额（元）"
            chart.x_axis.title = "月份"
            chart.height = 6.2
            chart.width = 11.2
            chart.style = 13
            chart.y_axis.numFmt = "#,##0"
            chart.y_axis.majorGridlines = None
            colours = {"成交实付": "2F75B5", "实际到账": "00A389", "标准成本": "6B7C8F", "广告费": "D99614"}
            for name in ("成交实付", "实际到账", "标准成本", "广告费"):
                chart.add_data(Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=month_end), titles_from_data=True)
                chart.series[-1].tx = SeriesLabel(v=name)
                _style_line_series(chart.series[-1], colours[name])
            chart.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=month_end))
            chart.legend.position = "b"
            worksheet.add_chart(chart, "A27")

        channel_end = _last_data_row(worksheet, headers["渠道_渠道"], header_row=header_row)
        if channel_end > header_row:
            contribution = BarChart()
            contribution.type = "bar"
            contribution.title = "渠道趋势性管理贡献（元）"
            contribution.x_axis.title = "金额（元）"
            contribution.height = 6.2
            contribution.width = 11.2
            contribution.style = 10
            contribution.legend = None
            contribution.x_axis.numFmt = "#,##0"
            contribution.x_axis.majorGridlines = None
            contribution.add_data(Reference(worksheet, min_col=headers["渠道_管理贡献"], min_row=header_row, max_row=channel_end), titles_from_data=True)
            contribution.set_categories(Reference(worksheet, min_col=headers["渠道_渠道"], min_row=header_row + 1, max_row=channel_end))
            if contribution.series:
                contribution.series[0].graphicalProperties.solidFill = "E26A45"
                contribution.series[0].graphicalProperties.line.solidFill = "E26A45"
            worksheet.add_chart(contribution, "H27")

            roas = BarChart()
            roas.type = "col"
            roas.title = "渠道广告效率 ROAS（倍）"
            roas.y_axis.title = "倍"
            roas.height = 6.2
            roas.width = 11.2
            roas.style = 10
            roas.legend = None
            roas.y_axis.numFmt = "0.00"
            roas.y_axis.majorGridlines = None
            roas.add_data(Reference(worksheet, min_col=headers["渠道_ROAS"], min_row=header_row, max_row=channel_end), titles_from_data=True)
            roas.set_categories(Reference(worksheet, min_col=headers["渠道_渠道"], min_row=header_row + 1, max_row=channel_end))
            if roas.series:
                roas.series[0].graphicalProperties.solidFill = "D99614"
                roas.series[0].graphicalProperties.line.solidFill = "D99614"
            worksheet.add_chart(roas, "A43")

        inventory_end = _last_data_row(worksheet, headers["库存_产品"], header_row=header_row)
        if inventory_end > header_row:
            inventory_chart = BarChart()
            inventory_chart.type = "bar"
            inventory_chart.title = "期末库存覆盖月数（管理线索）"
            inventory_chart.x_axis.title = "月"
            inventory_chart.height = 6.2
            inventory_chart.width = 11.2
            inventory_chart.style = 10
            inventory_chart.legend = None
            inventory_chart.x_axis.numFmt = "0.0"
            inventory_chart.x_axis.majorGridlines = None
            inventory_chart.add_data(Reference(worksheet, min_col=headers["库存_库存月数"], min_row=header_row, max_row=inventory_end), titles_from_data=True)
            inventory_chart.set_categories(Reference(worksheet, min_col=headers["库存_产品"], min_row=header_row + 1, max_row=inventory_end))
            if inventory_chart.series:
                inventory_chart.series[0].graphicalProperties.solidFill = "B42318"
                inventory_chart.series[0].graphicalProperties.line.solidFill = "B42318"
            worksheet.add_chart(inventory_chart, "H43")
        return
    if {"月份", "管理口径收入", "业务或生产成本", "费用金额", "估算经营贡献"}.issubset(headers):
        end = _last_data_row(worksheet, headers["月份"], header_row=header_row)
        if end > header_row:
            chart = LineChart()
            chart.title = "收入增长是否转化为经营贡献（元）"
            chart.y_axis.title = "金额（元）"
            chart.x_axis.title = "月份"
            chart.height = 6.2
            chart.width = 11.2
            chart.style = 13
            chart.visible_cells_only = False
            chart.y_axis.numFmt = "#,##0"
            chart.y_axis.majorGridlines = None
            chart.x_axis.majorGridlines = None
            colours = ("2F75B5", "6B5FD2", "D99614", "E26A45")
            for index, name in enumerate(("管理口径收入", "业务或生产成本", "费用金额", "估算经营贡献")):
                chart.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=end), titles_from_data=True
                )
                chart.series[-1].tx = SeriesLabel(v=name)
                _style_line_series(chart.series[-1], colours[index])
            chart.set_categories(Reference(worksheet, min_col=headers["月份"], min_row=header_row + 1, max_row=end))
            chart.legend.position = "b"
            worksheet.add_chart(chart, "A27")
    if {"客户_客户", "客户_管理口径收入", "客户_风险订单金额"}.issubset(headers):
        end = _last_data_row(worksheet, headers["客户_客户"], header_row=header_row)
        if end > header_row:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "高贡献客户及风险订单（元）"
            chart.x_axis.title = "金额（元）"
            chart.height = 6.2
            chart.width = 11.2
            chart.style = 10
            chart.visible_cells_only = False
            chart.x_axis.numFmt = "#,##0"
            chart.x_axis.majorGridlines = None
            for name, colour in (("客户_管理口径收入", "2F75B5"), ("客户_风险订单金额", "E26A45")):
                chart.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=end), titles_from_data=True
                )
                chart.series[-1].tx = SeriesLabel(v=name.replace("客户_", ""))
                chart.series[-1].graphicalProperties.solidFill = colour
                chart.series[-1].graphicalProperties.line.solidFill = colour
            chart.set_categories(
                Reference(worksheet, min_col=headers["客户_客户"], min_row=header_row + 1, max_row=end)
            )
            chart.legend.position = "b"
            worksheet.add_chart(chart, "H27")
    if {"销售_负责人", "销售_绩效表销售额", "销售_回款金额"}.issubset(headers):
        end = _last_data_row(worksheet, headers["销售_负责人"], header_row=header_row)
        if end > header_row:
            chart = BarChart()
            chart.type = "col"
            chart.title = "销售规模与现金转化（元）"
            chart.y_axis.title = "金额（元）"
            chart.x_axis.title = "负责人"
            chart.height = 6.2
            chart.width = 11.2
            chart.style = 10
            chart.visible_cells_only = False
            chart.y_axis.numFmt = "#,##0"
            chart.y_axis.majorGridlines = None
            for name in ("销售_绩效表销售额", "销售_回款金额"):
                chart.add_data(
                    Reference(worksheet, min_col=headers[name], min_row=header_row, max_row=end), titles_from_data=True
                )
                chart.series[-1].tx = SeriesLabel(v=name.replace("销售_", ""))
            chart.set_categories(
                Reference(worksheet, min_col=headers["销售_负责人"], min_row=header_row + 1, max_row=end)
            )
            chart.legend.position = "b"
            if len(chart.series) >= 2:
                chart.series[0].graphicalProperties.solidFill = "2F75B5"
                chart.series[0].graphicalProperties.line.solidFill = "2F75B5"
                chart.series[1].graphicalProperties.solidFill = "00A389"
                chart.series[1].graphicalProperties.line.solidFill = "00A389"
            worksheet.add_chart(chart, "A43")
    if {"库存_产品", "库存_库存月数"}.issubset(headers):
        end = _last_data_row(worksheet, headers["库存_产品"], header_row=header_row)
        if end > header_row:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "库存覆盖月数（偏高仅为线索）"
            chart.x_axis.title = "月"
            chart.height = 6.2
            chart.width = 11.2
            chart.style = 10
            chart.legend = None
            chart.visible_cells_only = False
            chart.x_axis.numFmt = "0.0"
            chart.x_axis.majorGridlines = None
            chart.add_data(
                Reference(worksheet, min_col=headers["库存_库存月数"], min_row=header_row, max_row=end),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(worksheet, min_col=headers["库存_产品"], min_row=header_row + 1, max_row=end)
            )
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = "D99614"
                chart.series[0].graphicalProperties.line.solidFill = "D99614"
            worksheet.add_chart(chart, "H43")


def _write_xlsx(
    tables: Mapping[str, pd.DataFrame],
    destination: Path,
    *,
    records: Sequence[OperationRecord],
    include_log: bool,
    index: bool,
) -> None:
    used: set[str] = set()
    expectations: list[tuple[str, int, int]] = []
    hr_report = "管理层人效总览" in tables
    adaptive_report = "管理层通用总览" in tables
    selection_report = "评选管理总览" in tables
    enterprise_report = "管理层诊断总览" in tables
    temp_path = _atomic_temp_path(destination)
    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            for table_name, frame in tables.items():
                sheet_name = _unique_sheet_name(table_name, used)
                if sheet_name == "经营诊断看板":
                    # Keep helper data below the visible 1:58 dashboard canvas
                    # so chart anchors remain stable in Excel and WPS.
                    start_row = 64
                elif sheet_name == "数据源确认":
                    start_row = 3
                elif sheet_name in {
                    "图表展示",
                    "库存图表看板",
                    "人力图表看板",
                    "自适应图表看板",
                    "评选图表看板",
                }:
                    start_row = 39
                elif (
                    sheet_name in _SALES_REPORT_SHEETS
                    or sheet_name in _INVENTORY_REPORT_SHEETS
                    or sheet_name in _HR_REPORT_SHEETS
                    or sheet_name in _ADAPTIVE_REPORT_SHEETS
                    or sheet_name in _SELECTION_REPORT_SHEETS
                    or sheet_name in _ENTERPRISE_REPORT_SHEETS
                ):
                    start_row = 3
                else:
                    start_row = 0
                header_row = start_row + 1
                frame.to_excel(writer, sheet_name=sheet_name, index=index, startrow=start_row)
                worksheet = writer.book[sheet_name]
                _style_worksheet(
                    worksheet,
                    summary=sheet_name
                    in {
                        "处理摘要",
                        "验收清单",
                        "质量概览",
                        "管理层数据总览",
                        "管理层库存总览",
                        "管理层人效总览",
                        "管理层通用总览",
                        "评选管理总览",
                        "管理层诊断总览",
                    },
                    long_text_detail=sheet_name in {"长文本明细", "空值清单"},
                    header_row=header_row,
                )
                if sheet_name == "数据源确认":
                    _style_source_confirmation_sheet(worksheet, header_row=header_row)
                elif enterprise_report and sheet_name in _ENTERPRISE_REPORT_SHEETS:
                    _style_enterprise_management_sheet(worksheet, header_row=header_row)
                    _add_enterprise_management_charts(worksheet, header_row=header_row)
                elif adaptive_report and sheet_name in _ADAPTIVE_REPORT_SHEETS:
                    _style_adaptive_management_sheet(worksheet, header_row=header_row)
                    _add_adaptive_management_charts(worksheet, header_row=header_row)
                elif selection_report and sheet_name in _SELECTION_REPORT_SHEETS:
                    _style_selection_management_sheet(worksheet, header_row=header_row)
                    _add_selection_management_charts(worksheet, header_row=header_row)
                elif hr_report and sheet_name in _HR_REPORT_SHEETS:
                    _style_hr_management_sheet(worksheet, header_row=header_row)
                    _add_hr_management_charts(worksheet, header_row=header_row)
                elif sheet_name in _SALES_REPORT_SHEETS:
                    _style_sales_management_sheet(worksheet, header_row=header_row)
                    _add_sales_management_charts(worksheet, header_row=header_row)
                elif sheet_name in _INVENTORY_REPORT_SHEETS:
                    _style_inventory_management_sheet(worksheet, header_row=header_row)
                    _add_inventory_management_charts(worksheet, header_row=header_row)
                _finalize_adaptive_layout(worksheet, header_row=header_row)
                expectations.append((sheet_name, worksheet.max_row, worksheet.max_column))

            if include_log:
                log_sheet = _unique_sheet_name("操作日志", used)
                log_frame = _escape_formula_cells(operation_log_frame(records))
                log_frame.to_excel(writer, sheet_name=log_sheet, index=False)
                _style_worksheet(writer.book[log_sheet], summary=True)
                _finalize_adaptive_layout(writer.book[log_sheet], header_row=1)
                expectations.append((log_sheet, len(log_frame) + 1, len(log_frame.columns)))
        _verify_xlsx_dimensions(temp_path, expectations)
        os.replace(temp_path, destination)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _verify_xlsx_dimensions(
    path: Path,
    expectations: Sequence[tuple[str, int, int]],
) -> None:
    """Fail closed if an XLSX write silently drops rows, columns or sheets."""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        actual_names = workbook.sheetnames
        expected_names = [name for name, _, _ in expectations]
        if actual_names != expected_names:
            raise ValueError("Excel 导出完整性校验失败：工作表数量或顺序不一致")
        for sheet_name, expected_rows, expected_columns in expectations:
            worksheet = workbook[sheet_name]
            if expected_columns and worksheet.max_column != expected_columns:
                raise ValueError(
                    f"Excel 导出完整性校验失败：{sheet_name} 列数应为 {expected_columns}，实际为 {worksheet.max_column}"
                )
            if worksheet.max_row != max(1, expected_rows):
                raise ValueError(
                    f"Excel 导出完整性校验失败：{sheet_name} 行数应为 {expected_rows}，实际为 {worksheet.max_row}"
                )
    finally:
        workbook.close()


def _write_single_csv(table: pd.DataFrame, destination: Path, *, index: bool, encoding: str) -> None:
    temp_path = _atomic_temp_path(destination)
    try:
        table.to_csv(temp_path, index=index, encoding=encoding)
        os.replace(temp_path, destination)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_csv_with_log(
    table: pd.DataFrame,
    destination: Path,
    log_table: pd.DataFrame,
    log_path: Path,
    *,
    index: bool,
    encoding: str,
) -> None:
    """Prepare both CSV files before replacing either destination."""

    table_temp = _atomic_temp_path(destination)
    log_temp = _atomic_temp_path(log_path)
    backups: dict[Path, Path] = {}
    committed: list[Path] = []
    try:
        table.to_csv(table_temp, index=index, encoding=encoding)
        log_table.to_csv(log_temp, index=False, encoding=encoding)
        for target in (destination, log_path):
            if target.exists():
                backup = _atomic_temp_path(target)
                backup.unlink(missing_ok=True)
                os.replace(target, backup)
                backups[target] = backup
        os.replace(table_temp, destination)
        committed.append(destination)
        os.replace(log_temp, log_path)
        committed.append(log_path)
    except Exception:
        for target in committed:
            target.unlink(missing_ok=True)
        for target, backup in backups.items():
            if backup.exists():
                os.replace(backup, target)
        raise
    finally:
        table_temp.unlink(missing_ok=True)
        log_temp.unlink(missing_ok=True)
        for backup in backups.values():
            backup.unlink(missing_ok=True)


def _write_csv_zip(
    tables: Mapping[str, pd.DataFrame],
    destination: Path,
    *,
    records: Sequence[OperationRecord],
    include_log: bool,
    index: bool,
    encoding: str,
) -> None:
    used: set[str] = set()
    temp_path = _atomic_temp_path(destination)
    try:
        with zipfile.ZipFile(temp_path, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
            for table_name, frame in tables.items():
                filename = sanitise_filename(table_name)
                filename = _unique_name(filename, used)
                archive.writestr(f"{filename}.csv", frame.to_csv(index=index).encode(encoding))
            if include_log:
                log_name = _unique_name("操作日志", used)
                archive.writestr(
                    f"{log_name}.csv",
                    _escape_formula_cells(operation_log_frame(records)).to_csv(index=False).encode(encoding),
                )
        os.replace(temp_path, destination)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def export_tables_to_path(
    tables: Mapping[str, pd.DataFrame],
    output_path: str | Path,
    *,
    operation_log: OperationLog | Sequence[OperationRecord] | None = None,
    include_log: bool = True,
    index: bool = False,
    csv_encoding: str = "utf-8-sig",
    escape_formulas: bool = True,
    overwrite: bool = False,
) -> ExportResult:
    """Export tables to XLSX, a single CSV, or a ZIP of CSV files.

    Text beginning with ``=``, ``+``, ``-`` or ``@`` is escaped by default to
    protect clients from CSV/Excel formula injection. Numeric negative values
    are not changed.
    """

    copied = _normalise_tables(tables, escape_formulas=escape_formulas)
    destination = _prepare_destination(output_path, overwrite=overwrite)
    records = _operation_records(operation_log)
    suffix = destination.suffix.lower()
    extra_files: list[Path] = []

    if suffix == ".xlsx":
        _write_xlsx(
            copied,
            destination,
            records=records,
            include_log=include_log,
            index=index,
        )
        export_format = "xlsx"
    elif suffix == ".zip":
        _write_csv_zip(
            copied,
            destination,
            records=records,
            include_log=include_log,
            index=index,
            encoding=csv_encoding,
        )
        export_format = "zip"
    elif suffix == ".csv":
        if len(copied) != 1:
            raise ValueError("导出 .csv 时只能有一个数据表；多表请导出为 .xlsx 或 .zip")
        frame = next(iter(copied.values()))
        if include_log:
            log_path = destination.with_name(f"{destination.stem}_操作日志.csv")
            if log_path.exists() and not overwrite:
                raise FileExistsError(f"操作日志文件已存在：{log_path}")
            _write_csv_with_log(
                frame,
                destination,
                _escape_formula_cells(operation_log_frame(records)),
                log_path,
                index=index,
                encoding=csv_encoding,
            )
            extra_files.append(log_path)
        else:
            _write_single_csv(frame, destination, index=index, encoding=csv_encoding)
        export_format = "csv"
    else:
        raise ValueError("output_path 扩展名必须是 .xlsx、.csv 或 .zip")

    return ExportResult(
        output_path=destination,
        format=export_format,
        files=(destination, *extra_files),
        rows_by_table={name: len(frame) for name, frame in copied.items()},
        operation_count=len(records),
    )
