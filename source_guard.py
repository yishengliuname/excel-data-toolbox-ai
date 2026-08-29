"""Source identity, generated-output blocking, and prompt/data alignment.

This module is deliberately independent from report builders.  Every upload
and every analytical route can therefore apply the same guardrails instead of
re-implementing one-off checks for each industry.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


_NON_WORD = re.compile(r"[\s_\-（）()【】\[\]：:/.]+")


def _normalise(value: Any) -> str:
    return _NON_WORD.sub("", str(value or "")).casefold()


_GENERATED_SIGNATURES: Mapping[str, frozenset[str]] = {
    "通用自适应经营分析报告": frozenset(
        {
            "管理层通用总览",
            "主数据分析",
            "数据字典",
            "数据质量",
            "表关系建议",
            "分类排名",
            "时间趋势",
            "异常数据",
            "自适应图表看板",
        }
    ),
    "企业经营诊断报告": frozenset(
        {
            "管理层诊断总览",
            "利润驱动分析",
            "风险行动计划",
            "诊断底稿",
            "数据口径与验收",
            "经营诊断看板",
        }
    ),
    "销售经营分析报告": frozenset(
        {"管理层数据总览", "产品分析", "销售人员分析", "异常数据提醒", "图表展示"}
    ),
    "库存经营分析报告": frozenset(
        {"管理层库存总览", "商品库存分析", "补货建议", "积压清单", "数据审计", "库存图表看板"}
    ),
    "人力经营分析报告": frozenset(
        {"管理层人效总览", "员工综合分析", "重点关注员工", "人工核验", "数据审计", "人力图表看板"}
    ),
    "候选对象评选报告": frozenset(
        {"评选管理总览", "建议入选名单", "全部候选排序", "风险复核清单", "评选规则与字段"}
    ),
}


@dataclass(frozen=True)
class GeneratedWorkbookAssessment:
    generated: bool
    report_kind: str
    matched_sheets: tuple[str, ...]
    sheet_names: tuple[str, ...]


def detect_generated_workbook(path: str | Path) -> GeneratedWorkbookAssessment:
    """Recognise a toolbox-produced management workbook from its sheet contract."""

    source = Path(path)
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        return GeneratedWorkbookAssessment(False, "", (), ())
    workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
    try:
        sheet_names = tuple(str(name) for name in workbook.sheetnames)
    finally:
        workbook.close()
    sheet_set = set(sheet_names)
    best_kind = ""
    best_matches: set[str] = set()
    best_ratio = 0.0
    for kind, signature in _GENERATED_SIGNATURES.items():
        matches = sheet_set & set(signature)
        ratio = len(matches) / len(signature)
        if len(matches) >= 3 and ratio >= 0.5 and (len(matches) > len(best_matches) or ratio > best_ratio):
            best_kind = kind
            best_matches = matches
            best_ratio = ratio
    return GeneratedWorkbookAssessment(
        bool(best_kind),
        best_kind,
        tuple(sorted(best_matches)),
        sheet_names,
    )


_DOMAIN_TOKENS: Mapping[str, tuple[str, ...]] = {
    "餐饮门店": (
        "门店",
        "pos",
        "菜品",
        "餐品",
        "食材",
        "原料",
        "bom",
        "堂食",
        "桌号",
        "外卖",
        "报损",
        "盘点损耗",
        "出餐",
    ),
    "多平台电商": (
        "天猫",
        "京东",
        "买家",
        "会员",
        "广告投放",
        "广告花费",
        "roas",
        "sku库存",
        "月末库存",
        "商品利润",
        "渠道与广告",
    ),
    "员工人效": ("员工", "考勤", "薪资", "工资", "绩效", "加班", "缺勤", "工时", "人效"),
    "采购库存": (
        "期初库存",
        "销售出库",
        "库存调整",
        "安全库存",
        "目标库存",
        "补货",
        "积压",
        "仓库",
    ),
    "财务分析": ("会计科目", "凭证", "借方", "贷方", "预算", "应收账款", "现金流", "资产负债"),
    "候选评选": ("候选", "参赛", "比赛", "评选", "入选", "评语", "轮得分", "综合得分"),
}


def _domain_scores(text: str) -> dict[str, int]:
    folded = _normalise(text)
    return {
        domain: sum(1 for token in tokens if _normalise(token) in folded)
        for domain, tokens in _DOMAIN_TOKENS.items()
    }


def _top_domain(scores: Mapping[str, int], *, minimum: int) -> tuple[str, int, int]:
    ordered = sorted(scores.items(), key=lambda item: (item[1], item[0]), reverse=True)
    if not ordered or ordered[0][1] < minimum:
        return "通用/未确定", 0, 0
    second = ordered[1][1] if len(ordered) > 1 else 0
    return ordered[0][0], ordered[0][1], second


@dataclass(frozen=True)
class SourceSemanticAssessment:
    prompt_domain: str
    data_domain: str
    prompt_score: int
    data_score: int
    aligned: bool
    reason: str
    prompt_scores: Mapping[str, int]
    data_scores: Mapping[str, int]


def assess_prompt_data_alignment(
    prompt: str,
    frames: Sequence[pd.DataFrame],
    source_names: Sequence[str],
) -> SourceSemanticAssessment:
    """Compare explicit business language with observed table/field semantics."""

    schema_parts = list(source_names)
    for frame in frames:
        if isinstance(frame, pd.DataFrame):
            schema_parts.extend(str(column) for column in frame.columns)
    prompt_scores = _domain_scores(prompt)
    data_scores = _domain_scores(" ".join(schema_parts))
    prompt_domain, prompt_score, prompt_second = _top_domain(prompt_scores, minimum=2)
    data_domain, data_score, data_second = _top_domain(data_scores, minimum=4)
    decisive_prompt = prompt_domain != "通用/未确定" and prompt_score - prompt_second >= 1
    decisive_data = data_domain != "通用/未确定" and data_score - data_second >= 2
    aligned = not (decisive_prompt and decisive_data and prompt_domain != data_domain)
    if aligned:
        reason = f"需求语义={prompt_domain}；数据语义={data_domain}；未发现明确冲突"
    else:
        reason = (
            f"需求语义识别为“{prompt_domain}”，但当前表名和字段更符合“{data_domain}”；"
            "疑似上传了历史报告或错误项目文件"
        )
    return SourceSemanticAssessment(
        prompt_domain,
        data_domain,
        prompt_score,
        data_score,
        aligned,
        reason,
        prompt_scores,
        data_scores,
    )


def source_confirmation_frame(
    frames: Sequence[pd.DataFrame],
    source_names: Sequence[str],
    *,
    file_names: Sequence[str],
    task_id: str,
    user_request: str,
) -> pd.DataFrame:
    assessment = assess_prompt_data_alignment(user_request, frames, source_names)
    rows: list[dict[str, Any]] = []
    known_files = list(dict.fromkeys(str(name) for name in file_names if str(name).strip()))
    for index, (frame, source_name) in enumerate(zip(frames, source_names, strict=True), start=1):
        source_text = str(source_name)
        file_name = next(
            (
                name
                for name in known_files
                if source_text.startswith(Path(name).name + "::")
                or source_text.startswith(Path(name).name + "__")
            ),
            "",
        )
        if "::" in source_text:
            sheet_name = source_text.split("::", 1)[1]
        else:
            sheet_name = source_text.split("__", 1)[1] if "__" in source_text else source_text
        rows.append(
            {
                "序号": index,
                "任务编号": task_id,
                "本次明确上传文件": file_name or "由当前任务明确选择",
                "纳入分析工作表": sheet_name,
                "数据行数": len(frame),
                "字段数": frame.shape[1],
                "需求语义": assessment.prompt_domain,
                "数据语义": assessment.data_domain,
                "一致性结论": "通过" if assessment.aligned else "阻断",
                "来源状态": "原始输入（允许分析）",
                "校验说明": assessment.reason,
            }
        )
    return pd.DataFrame(rows)


__all__ = [
    "GeneratedWorkbookAssessment",
    "SourceSemanticAssessment",
    "assess_prompt_data_alignment",
    "detect_generated_workbook",
    "source_confirmation_frame",
]
