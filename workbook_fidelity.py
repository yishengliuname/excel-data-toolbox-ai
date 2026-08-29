"""Workbook-preserving exports for orders where formatting is part of the job.

Pandas is the correct engine for table transformations, but rebuilding a
workbook from data frames necessarily discards workbook-level objects.  This
module instead clones the customer's source workbook and updates or appends only
the requested rectangular data areas.  Existing charts, conditional formatting,
named ranges, hidden sheets and VBA streams remain in the cloned workbook.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
import os
import re
import tempfile
from typing import Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("_", str(value)).strip()[:31] or "处理结果"
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f"_{index}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def _write_frame(worksheet, frame: pd.DataFrame, *, replace_existing: bool) -> None:
    old_rows = worksheet.max_row
    old_columns = worksheet.max_column
    if replace_existing:
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=max(old_rows, len(frame) + 1),
            min_col=1,
            max_col=max(old_columns, len(frame.columns)),
        ):
            for cell in row:
                cell.value = None

    header_style = None
    if old_rows >= 1 and old_columns >= 1:
        source = worksheet.cell(1, 1)
        header_style = {
            "font": copy(source.font),
            "fill": copy(source.fill),
            "border": copy(source.border),
            "alignment": copy(source.alignment),
            "number_format": source.number_format,
            "protection": copy(source.protection),
        }

    for column_index, column in enumerate(frame.columns, start=1):
        cell = worksheet.cell(1, column_index, str(column))
        if header_style:
            for attribute, value in header_style.items():
                setattr(cell, attribute, value)
        else:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B5D3B")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            worksheet.cell(row_index, column_index, value)

    if worksheet.max_row > len(frame) + 1:
        for row in worksheet.iter_rows(
            min_row=len(frame) + 2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=max(worksheet.max_column, len(frame.columns)),
        ):
            for cell in row:
                cell.value = None
    worksheet.freeze_panes = worksheet.freeze_panes or "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(max(1, len(frame) + 1), max(1, len(frame.columns))).coordinate}"


def preserve_workbook_export(
    source_workbook: str | Path,
    destination: str | Path,
    tables: Mapping[str, pd.DataFrame],
    *,
    replace_sheets: Mapping[str, str] | None = None,
) -> Path:
    """Clone one XLSX/XLSM and write table results without flattening the book.

    ``replace_sheets`` maps result-table names to existing worksheet names.  Any
    result without a mapping is appended as a new sheet, which is the safest
    default for customer workbooks.
    """

    source = Path(source_workbook).resolve()
    target = Path(destination).resolve()
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("保真模式只支持 .xlsx 或 .xlsm 源文件")
    if not source.is_file():
        raise FileNotFoundError(f"源工作簿不存在：{source}")
    if target.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("保真模式输出必须是 .xlsx 或 .xlsm")
    target.parent.mkdir(parents=True, exist_ok=True)
    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = load_workbook(source, keep_vba=keep_vba, keep_links=True, data_only=False)
    try:
        used = {name.casefold() for name in workbook.sheetnames}
        mappings = {str(key): str(value) for key, value in (replace_sheets or {}).items()}
        for table_name, frame in tables.items():
            requested_sheet = mappings.get(str(table_name))
            if requested_sheet and requested_sheet in workbook.sheetnames:
                worksheet = workbook[requested_sheet]
                _write_frame(worksheet, frame, replace_existing=True)
            else:
                sheet_name = _safe_sheet_name(str(table_name), used)
                worksheet = workbook.create_sheet(sheet_name)
                _write_frame(worksheet, frame, replace_existing=False)
        handle, temp_name = tempfile.mkstemp(prefix=f".{target.stem}_", suffix=target.suffix, dir=target.parent)
        os.close(handle)
        temp = Path(temp_name)
        try:
            workbook.save(temp)
            # Reopen before commit so a corrupt clone never replaces a delivery.
            checked = load_workbook(temp, read_only=True, keep_vba=keep_vba, data_only=False)
            checked.close()
            os.replace(temp, target)
        finally:
            temp.unlink(missing_ok=True)
    finally:
        workbook.close()
    return target


def workbook_feature_inventory(path: str | Path) -> dict[str, object]:
    """Return a non-destructive inventory used in preflight and acceptance."""

    source = Path(path)
    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = load_workbook(source, read_only=False, keep_vba=keep_vba, keep_links=True, data_only=False)
    try:
        return {
            "filename": source.name,
            "worksheets": list(workbook.sheetnames),
            "hidden_sheets": [sheet.title for sheet in workbook.worksheets if sheet.sheet_state != "visible"],
            "charts": sum(len(sheet._charts) for sheet in workbook.worksheets),
            "images": sum(len(sheet._images) for sheet in workbook.worksheets),
            "merged_ranges": sum(len(sheet.merged_cells.ranges) for sheet in workbook.worksheets),
            "defined_names": len(list(workbook.defined_names.values())),
            "has_vba": bool(getattr(workbook, "vba_archive", None)),
            "external_links": len(getattr(workbook, "_external_links", [])),
        }
    finally:
        workbook.close()


__all__ = ["preserve_workbook_export", "workbook_feature_inventory"]
