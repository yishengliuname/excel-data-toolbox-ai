from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
import pandas as pd

from excel_data_toolbox.core import export_tables


class GlobalExcelLayoutTests(unittest.TestCase):
    def test_all_generated_sheets_wrap_narrative_text_and_expand_rows(self) -> None:
        long_review = (
            "模型与结论需要进一步核验；参考文献格式需要统一；章节层级较多，"
            "建议补充关键结果解释、修正公式编号，并把风险、处理动作和人工复核边界写清楚。"
        ) * 4
        tables = {
            "普通问题表": pd.DataFrame({
                "编号": [1, 2],
                "风险提示": ["明显格式与可读性问题", "结论需要人工复核"],
                "评语摘要": [long_review, long_review[:180]],
                "金额": [123456.78, 98765.43],
            }),
            "处理说明": pd.DataFrame({
                "项目": ["版式", "交付"],
                "处理建议": [long_review, "保持表头、筛选、冻结窗格和一致的数字格式。"],
            }),
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "adaptive-layout.xlsx"
            export_tables(tables, path, include_log=False)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                for sheet_name, narrative_header in (
                    ("普通问题表", "评语摘要"),
                    ("处理说明", "处理建议"),
                ):
                    worksheet = workbook[sheet_name]
                    headers = {
                        str(cell.value or ""): cell.column
                        for cell in worksheet[1]
                    }
                    column = headers[narrative_header]
                    letter = worksheet.cell(1, column).column_letter
                    self.assertGreaterEqual(worksheet.column_dimensions[letter].width, 36)
                    self.assertLessEqual(worksheet.column_dimensions[letter].width, 58)
                    self.assertTrue(worksheet.cell(2, column).alignment.wrap_text)
                    self.assertEqual(worksheet.cell(2, column).alignment.vertical, "top")
                    self.assertGreater(worksheet.row_dimensions[2].height or 0, 80)
                    self.assertEqual(worksheet.sheet_format.defaultRowHeight, 21)
                    self.assertTrue(worksheet.cell(1, column).alignment.wrap_text)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
