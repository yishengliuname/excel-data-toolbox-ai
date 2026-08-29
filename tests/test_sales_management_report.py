from __future__ import annotations

from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch
import zipfile

import pandas as pd
from openpyxl import load_workbook

import excel_data_toolbox.server as server_module
from excel_data_toolbox.core import export_tables
from excel_data_toolbox.nl_agent import build_table_catalog, execute_plan, validate_plan
from excel_data_toolbox.sales_report import (
    build_sales_management_report,
    infer_sales_report_columns,
)
from excel_data_toolbox.server import AppSession, ToolboxHandler


def _customer_sales() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "日期": [
                "2026-01-03", "2026-01-05", "2026-01-08", "2026-01-12", "2026-01-15",
                "2026-02-02", "2026-02-07", "2026-02-13", "2026-02-20", "2026-02-26",
            ],
            "产品类别": ["智能设备", "软件服务", "智能设备", "数据服务", "软件服务", "智能设备", "数据服务", "软件服务", "智能设备", "数据服务"],
            "地区": ["华东", "华南", "华北", "华东", "西南", "华南", "华北", "华东", "西南", "华南"],
            "销售人员": ["张伟", "李娜", "王强", "赵敏", "陈浩", "李娜", "王强", "张伟", "陈浩", "赵敏"],
            "订单数量": [12, 8, 15, 20, 10, 18, 25, 13, 9, 30],
            "销售金额": [36000, 24000, 45000, 50000, 30000, 54000, 62500, 39000, 27000, 75000],
            "成本": [22000, 9000, 28000, 18000, 12000, 33000, 22000, 15000, 17000, 26000],
            "客户满意度": [5, 4, 5, 4, 3, 5, 4, 5, 3, 5],
        }
    )


def _params(frame: pd.DataFrame) -> dict[str, object]:
    return {**infer_sales_report_columns(frame), "satisfaction_threshold": 4}


def test_sales_report_calculates_requested_kpis_rankings_and_alerts() -> None:
    frame = _customer_sales()
    result = build_sales_management_report(frame, **_params(frame))

    assert list(result.outputs) == ["管理层数据总览", "产品分析", "销售人员分析", "异常数据提醒", "图表展示"]
    assert result.report["total_sales"] == 442500.0
    assert result.report["total_cost"] == 202000.0
    assert result.report["total_profit"] == 240500.0
    assert result.report["overall_profit_margin"] == 240500.0 / 442500.0
    assert result.report["top_product_by_sales"] == "数据服务"
    assert result.report["top_product_by_profit"] == "数据服务"
    assert result.report["top_salesperson"] == "赵敏"
    assert result.report["attention_rows"] == 2
    assert result.outputs["异常数据提醒"]["客户满意度"].tolist() == [3, 3]
    assert result.outputs["图表展示"]["地区"].dropna().tolist() == ["西南", "华北", "华东", "华南"]


def test_allowlisted_plan_generates_exactly_five_friendly_output_names() -> None:
    frame = _customer_sales()
    catalog = build_table_catalog({"sales": frame}, display_names={"sales": "销售数据原始表"})
    plan = validate_plan(
        {
            "schema_version": 1,
            "status": "ready",
            "summary": "生成销售经营管理报告",
            "message": "可执行",
            "clarification_questions": [],
            "assumptions": [],
            "warnings": [],
            "steps": [
                {
                    "id": "sales_report_1",
                    "operation": "sales_management_report",
                    "input_ids": ["sales"],
                    "output_name": "销售经营分析报告",
                    "params": _params(frame),
                }
            ],
        },
        catalog,
    )
    executed = execute_plan(plan, {"sales": frame}, dry_run=False)

    assert list(executed.tables) == ["管理层数据总览", "产品分析", "销售人员分析", "异常数据提醒", "图表展示"]


def test_sales_report_export_contains_five_sheets_and_three_native_charts(tmp_path: Path) -> None:
    frame = _customer_sales()
    result = build_sales_management_report(frame, **_params(frame))
    path = tmp_path / "销售经营分析报告.xlsx"

    export_tables(result.outputs, path, include_log=False)

    workbook = load_workbook(path, data_only=False)
    try:
        assert workbook.sheetnames == ["管理层数据总览", "产品分析", "销售人员分析", "异常数据提醒", "图表展示"]
        assert len(workbook["图表展示"]._charts) == 3
        assert workbook["管理层数据总览"]["B5"].value == 442500.0
        assert workbook["管理层数据总览"]["B8"].number_format == "0.00%"
        assert workbook["管理层数据总览"]["A1"].value == "销售经营管理驾驶舱"
        assert workbook["产品分析"].column_dimensions["B"].width >= 17
        assert workbook["销售人员分析"].column_dimensions["D"].width >= 17
        assert workbook["异常数据提醒"].column_dimensions["F"].width >= 17
        assert workbook["图表展示"]["A40"].value == "月份"
        assert workbook["图表展示"].sheet_view.showGridLines is False
        assert workbook["图表展示"]._charts[1].__class__.__name__ == "DoughnutChart"
    finally:
        workbook.close()

    with zipfile.ZipFile(path) as archive:
        doughnut_xml = archive.read("xl/charts/chart2.xml").decode("utf-8")
        sheet_xml = archive.read("xl/worksheets/sheet5.xml").decode("utf-8")
    assert 'dLblPos val="bestFit"' not in doughnut_xml
    assert 'pane="bottomLeft"' not in sheet_xml


def test_unified_sales_report_uses_local_engine_when_deepseek_is_unavailable() -> None:
    session = AppSession()
    handler = object.__new__(ToolboxHandler)
    try:
        table_id = session.add_table("销售数据原始表", _customer_sales(), source="单元测试", original=True)
        prompt = (
            "分析当前销售数据，自动计算总销售额、总成本、总利润和平均利润率；"
            "生成产品分析、销售人员排名、月度销售、地区销售、客户满意度异常；"
            "输出管理层数据总览、产品分析、销售人员分析、异常数据提醒、图表展示五张工作表。"
        )
        with (
            patch.object(server_module, "SESSION", session),
            patch.object(
                server_module,
                "_project_ai_config",
                return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"},
            ),
            patch.object(
                server_module.DeepSeekClient,
                "classify_unified_request",
                side_effect=AssertionError("完整销售报告不应访问 DeepSeek"),
            ),
        ):
            response = handler._ai_unified({"prompt": prompt, "table_ids": [table_id]})

        assert response["mode"] == "data"
        assert response["status"] == "ready"
        assert response["route"]["reason"].startswith("已识别为完整销售经营报告")
    finally:
        session.close()


class SalesManagementReportTests(unittest.TestCase):
    def test_kpis_rankings_and_alerts(self) -> None:
        test_sales_report_calculates_requested_kpis_rankings_and_alerts()

    def test_allowlisted_plan_output_names(self) -> None:
        test_allowlisted_plan_generates_exactly_five_friendly_output_names()

    def test_export_sheets_and_native_charts(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            test_sales_report_export_contains_five_sheets_and_three_native_charts(Path(temporary))

    def test_sales_report_bypasses_unavailable_deepseek(self) -> None:
        test_unified_sales_report_uses_local_engine_when_deepseek_is_unavailable()


if __name__ == "__main__":
    unittest.main()
