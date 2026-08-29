from __future__ import annotations

from pathlib import Path
import json
import zipfile

import pandas as pd
import pandas.testing as pdt
import pytest
from openpyxl import load_workbook

from excel_data_toolbox import (
    CleaningConfig,
    OperationLog,
    concat_tables,
    export_tables,
    group_summary,
    join_tables,
    load_tables,
    lookup_match,
    mask_columns,
    profile_dataframe,
    select_rename_sort,
    smart_clean,
    split_dataframe,
)


def test_profile_dataframe_reports_shape_missing_duplicates_and_samples() -> None:
    frame = pd.DataFrame(
        {
            "客户编号": ["000001", "000002", "000002"],
            "金额": [10.0, None, None],
        }
    )

    profile = profile_dataframe(frame, sample_values=2)

    assert profile.row_count == 3
    assert profile.column_count == 2
    assert profile.duplicate_row_count == 1
    assert profile.missing_cell_count == 2
    assert profile.columns[0].semantic_type == "identifier"
    assert profile.columns[0].sample_values == ("000001", "000002")
    assert profile.columns[1].missing_percent == pytest.approx(66.67)

    dated = profile_dataframe(pd.DataFrame({"日期": [pd.Timestamp("2026-01-01")]}))
    assert json.loads(json.dumps(dated.to_dict(), ensure_ascii=False))["columns"][0][
        "sample_values"
    ] == ["2026-01-01T00:00:00"]


def test_profile_dataframe_handles_nested_list_and_dict_values() -> None:
    frame = pd.DataFrame(
        {
            "指标": ["候选键", "候选键", "规则"],
            "结果": [["A", "B"], ["A", "B"], {"容差": 0.05}],
        }
    )

    profile = profile_dataframe(frame)

    assert profile.row_count == 3
    assert profile.columns[1].unique_count == 2
    assert profile.duplicate_row_count == 1


def test_smart_clean_is_non_mutating_and_cleans_common_problems() -> None:
    original = pd.DataFrame(
        {
            "姓名": [" 张三 ", " 张三 ", "", None],
            "金额": ["10", "10", "20", None],
            "日期": ["2026-01-02", "2026-01-02", "2026-02-03", None],
            "空列": [None, None, None, None],
        }
    )
    untouched = original.copy(deep=True)
    log = OperationLog()

    cleaned, report = smart_clean(original, log=log, table_name="订单")

    pdt.assert_frame_equal(original, untouched)
    assert list(cleaned.columns) == ["姓名", "金额", "日期"]
    assert cleaned.loc[0, "姓名"] == "张三"
    assert pd.isna(cleaned.loc[1, "姓名"])
    assert str(cleaned["金额"].dtype) == "Int64"
    assert pd.api.types.is_datetime64_any_dtype(cleaned["日期"])
    assert report.rows_before == 4
    assert report.rows_after == 2
    assert report.empty_rows_removed == 1
    assert report.empty_columns_removed == ("空列",)
    assert report.duplicate_rows_removed == 1
    assert report.inferred_types == {"金额": "integer", "日期": "datetime"}
    assert len(log) == 1
    assert log.entries[0].action == "智能清洗"


def test_smart_clean_fill_and_drop_missing_modes() -> None:
    frame = pd.DataFrame({"组": ["A", None], "数量": [None, 2]})

    filled, _ = smart_clean(
        frame,
        CleaningConfig(infer_types=False, drop_duplicates=False, missing_strategy="fill"),
    )
    assert filled.loc[0, "数量"] == 0
    assert filled.loc[1, "组"] == "未填写"

    dropped, _ = smart_clean(
        frame,
        CleaningConfig(
            infer_types=False,
            drop_duplicates=False,
            missing_strategy="drop_rows",
            missing_subset=("组",),
        ),
    )
    assert len(dropped) == 1
    assert dropped.loc[0, "组"] == "A"
    assert pd.isna(dropped.loc[0, "数量"])


def test_smart_clean_preserves_identifiers_and_large_integer_text() -> None:
    frame = pd.DataFrame(
        {
            "代码": ["00123", "00456"],
            "超长序列": ["9007199254740993", "9007199254740995"],
            "普通整数": ["10", "20"],
        }
    )

    cleaned, report = smart_clean(frame)

    assert cleaned["代码"].tolist() == ["00123", "00456"]
    assert cleaned["超长序列"].tolist() == ["9007199254740993", "9007199254740995"]
    assert cleaned["普通整数"].tolist() == [10, 20]
    assert report.inferred_types == {"普通整数": "integer"}

    duplicate_index = pd.DataFrame(
        {"普通整数": ["1", "2"], "日期": ["2026-01-01", "2026-01-02"]},
        index=[7, 7],
    )
    duplicate_index_cleaned, _ = smart_clean(duplicate_index)
    assert duplicate_index_cleaned["普通整数"].tolist() == [1, 2]
    assert duplicate_index_cleaned["日期"].dt.day.tolist() == [1, 2]


def test_select_rename_sort_preserves_original() -> None:
    frame = pd.DataFrame({"姓名": ["乙", "甲"], "金额": [2, 1], "备注": ["b", "a"]})
    untouched = frame.copy(deep=True)

    result = select_rename_sort(
        frame,
        columns=["金额", "姓名"],
        rename={"金额": "销售额"},
        sort_by="销售额",
    )

    pdt.assert_frame_equal(frame, untouched)
    assert result.to_dict("records") == [
        {"销售额": 1, "姓名": "甲"},
        {"销售额": 2, "姓名": "乙"},
    ]


def test_concat_join_and_group_summary() -> None:
    january = pd.DataFrame({"客户": ["A", "B"], "金额": [10, 20]})
    february = pd.DataFrame({"客户": ["A"], "金额": [5], "数量": [2]})

    combined = concat_tables({"一月": january, "二月": february})
    assert len(combined) == 3
    assert combined["来源表"].tolist() == ["一月", "一月", "二月"]
    assert pd.isna(combined.loc[0, "数量"])

    customers = pd.DataFrame({"客户": ["A", "B"], "区域": ["东", "西"]})
    joined = join_tables(combined, customers, on="客户", validate="many_to_one")
    assert joined["区域"].tolist() == ["东", "西", "东"]

    summary = group_summary(
        joined,
        by="区域",
        aggregations={"金额": ["sum", "mean"], "客户": "count"},
    )
    assert list(summary.columns) == ["区域", "金额_sum", "金额_mean", "客户_count"]
    east = summary.loc[summary["区域"] == "东"].iloc[0]
    assert east["金额_sum"] == 15
    assert east["客户_count"] == 2


def test_lookup_split_and_mask() -> None:
    orders = pd.DataFrame(
        {
            "客户号": [1, 2, 3],
            "手机号": ["13812345678", "13987654321", None],
            "邮箱": ["alice@example.com", "bob@example.com", None],
        }
    )
    lookup = pd.DataFrame(
        {"编号": [1, 1, 2], "等级": ["金", "旧值", "银"]}
    )

    matched = lookup_match(
        orders,
        lookup,
        source_key="客户号",
        lookup_key="编号",
        value_columns=["等级"],
    )
    assert len(matched) == len(orders)
    assert matched["等级"].tolist()[:2] == ["金", "银"]
    assert matched["匹配状态"].tolist() == ["已匹配", "已匹配", "未匹配"]

    parts = split_dataframe(matched, by="匹配状态")
    assert set(parts) == {"已匹配", "未匹配"}
    assert [len(parts[name]) for name in ("已匹配", "未匹配")] == [2, 1]

    masked = mask_columns(matched, {"手机号": "phone", "邮箱": "email"})
    assert masked.loc[0, "手机号"] == "138****5678"
    assert masked.loc[0, "邮箱"] == "a****@example.com"
    assert pd.isna(masked.loc[2, "手机号"])
    assert orders.loc[0, "手机号"] == "13812345678"


def test_lookup_does_not_drop_same_named_source_column() -> None:
    source = pd.DataFrame({"客户号": [1], "编号": ["源表原值"]})
    lookup = pd.DataFrame({"编号": [1], "等级": ["金"]})

    result = lookup_match(
        source,
        lookup,
        source_key="客户号",
        lookup_key="编号",
        value_columns=["等级"],
    )

    assert result.loc[0, "编号"] == "源表原值"
    assert result.loc[0, "等级"] == "金"


def test_load_multisheet_excel_and_csv(tmp_path: Path) -> None:
    workbook = tmp_path / "业务.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"a": [1]}).to_excel(writer, sheet_name="订单", index=False)
        pd.DataFrame({"b": [2]}).to_excel(writer, sheet_name="客户", index=False)
    csv_path = tmp_path / "补充.csv"
    pd.DataFrame({"c": [3]}).to_csv(csv_path, index=False, encoding="utf-8-sig")

    tables = load_tables([workbook, csv_path])

    assert set(tables) == {"业务.xlsx::订单", "业务.xlsx::客户", "补充.csv::CSV"}
    assert tables["业务.xlsx::订单"].iloc[0, 0] == 1
    assert tables["补充.csv::CSV"].iloc[0, 0] == "3"

    gbk_path = tmp_path / "国标编码.csv"
    pd.DataFrame({"城市": ["上海"]}).to_csv(
        gbk_path, index=False, encoding="gb18030"
    )
    assert load_tables(gbk_path)["国标编码.csv::CSV"].loc[0, "城市"] == "上海"

    identifiers_path = tmp_path / "标识符.csv"
    identifiers_path.write_text("代码,备注\n00123,NA\n00456,N/A\n", encoding="utf-8")
    identifiers = load_tables(identifiers_path)["标识符.csv::CSV"]
    assert identifiers.to_dict("records") == [
        {"代码": "00123", "备注": "NA"},
        {"代码": "00456", "备注": "N/A"},
    ]


def test_export_xlsx_zip_and_single_csv_with_operation_log(tmp_path: Path) -> None:
    tables = {
        "订单/明细": pd.DataFrame({"订单号": [1, 2], "金额": [10, 20]}),
        "客户": pd.DataFrame({"客户": ["甲"]}),
    }
    log = OperationLog()
    log.record("测试操作", input_tables=["原表"], output_tables=list(tables))

    xlsx_path = tmp_path / "result.xlsx"
    xlsx_result = export_tables(tables, xlsx_path, operation_log=log)
    assert xlsx_result.output_path == xlsx_path
    assert xlsx_result.operation_count == 1
    assert set(pd.ExcelFile(xlsx_path).sheet_names) == {"订单_明细", "客户", "操作日志"}
    workbook = load_workbook(xlsx_path)
    worksheet = workbook["订单_明细"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].font.color.rgb == "00FFFFFF"
    assert worksheet.sheet_view.showGridLines is False

    zip_path = tmp_path / "result.zip"
    export_tables(tables, zip_path, operation_log=log)
    with zipfile.ZipFile(zip_path) as archive:
        names = set(archive.namelist())
    assert names == {"订单_明细.csv", "客户.csv", "操作日志.csv"}

    csv_path = tmp_path / "one.csv"
    csv_result = export_tables({"订单": tables["订单/明细"]}, csv_path, operation_log=log)
    assert set(csv_result.files) == {csv_path, tmp_path / "one_操作日志.csv"}
    assert pd.read_csv(csv_path).shape == (2, 2)

    with pytest.raises(FileExistsError):
        export_tables({"订单": tables["订单/明细"]}, csv_path)


def test_export_escapes_formula_text_and_handles_long_duplicate_sheet_names(
    tmp_path: Path,
) -> None:
    dangerous = pd.DataFrame(
        {"=危险表头": ["=2+2", "+cmd", "正常"], "数值": [-1, 2, 3]}
    )
    long_prefix = "非常长的工作表名称" * 4
    path = tmp_path / "safe.xlsx"

    export_tables(
        {f"{long_prefix}?": dangerous, f"{long_prefix}*": dangerous},
        path,
    )

    workbook = pd.ExcelFile(path)
    data_sheets = [name for name in workbook.sheet_names if name != "操作日志"]
    assert len(data_sheets) == 2
    assert len(set(name.casefold() for name in data_sheets)) == 2
    assert all(len(name) <= 31 for name in data_sheets)
    exported = pd.read_excel(path, sheet_name=data_sheets[0])
    assert "'=危险表头" in exported.columns
    assert exported.loc[0, "'=危险表头"] == "'=2+2"
    assert exported.loc[0, "数值"] == -1


def test_export_wraps_long_text_without_losing_values(tmp_path: Path) -> None:
    long_text = "摘要偏少；目录格式错误；图中文字太小；结果分析需要进一步展开。" * 5
    frame = pd.DataFrame({"序号": [1, 2], "问题说明": [long_text, "正常"]})
    path = tmp_path / "readable.xlsx"

    export_tables({"成绩问题": frame}, path, include_log=False)

    exported = pd.read_excel(path, sheet_name="成绩问题")
    assert exported.shape == frame.shape
    assert exported.loc[0, "问题说明"] == long_text
    workbook = load_workbook(path)
    worksheet = workbook["成绩问题"]
    assert worksheet["B2"].alignment.wrap_text is True
    # Source sheets stay compact; the professional export adds a separate
    # long-text detail sheet for full-width reading.
    assert 36 <= worksheet.column_dimensions["B"].width <= 58
    assert worksheet.row_dimensions[2].height > 20


def test_invalid_operations_raise_clear_errors(tmp_path: Path) -> None:
    frame = pd.DataFrame({"a": [1]})
    with pytest.raises(KeyError, match="不存在的列"):
        select_rename_sort(frame, columns=["missing"])
    with pytest.raises(ValueError, match="只能提供"):
        split_dataframe(frame, by="a", rows_per_table=1)
    with pytest.raises(ValueError, match="只能有一个"):
        export_tables({"a": frame, "b": frame}, tmp_path / "bad.csv")
    with pytest.raises(ValueError, match="重复数据表名称"):
        export_tables({"A": frame, " A ": frame}, tmp_path / "bad.xlsx")
    duplicate_columns = pd.DataFrame([[1, 2]], columns=["a", "a"])
    with pytest.raises(ValueError, match="重复列名"):
        smart_clean(duplicate_columns)

    log = OperationLog()
    log.record("审计", details={"嵌套": {"数量": 1}})
    with pytest.raises(TypeError):
        log.entries[0].details["新增"] = "不可修改"  # type: ignore[index]
