from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
import sqlite3
import sys

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_data_toolbox.advanced_automation import (
    build_vba_bundle,
    document_capabilities,
    extract_image_text,
    query_sqlite_read_only,
    validate_vba_module,
)
from excel_data_toolbox.core import export_tables
from excel_data_toolbox.delivery_qa import verify_delivery
from excel_data_toolbox.large_data import query_files, validate_read_only_sql
from excel_data_toolbox.nl_agent import build_table_catalog, normalize_plan_envelope, validate_plan
from excel_data_toolbox.order_intake import quote_order
from excel_data_toolbox.scheduler import LocalScheduler
from excel_data_toolbox.secure_secrets import SecureSecretStore
from excel_data_toolbox.task_store import TaskRepository
from excel_data_toolbox.workbook_fidelity import preserve_workbook_export, workbook_feature_inventory


def test_ai_envelope_repairs_aliases_and_summary_parameters() -> None:
    frame = pd.DataFrame({"地区": ["华东"], "销售额": [100]})
    catalog = build_table_catalog({"sales": frame})
    raw = {
        "version": "v1",
        "status": "ready",
        "normalized_request": "按地区汇总销售额",
        "description": "可执行",
        "questions": [],
        "notes": [],
        "assumption_list": [],
        "operations": [
            {
                "step_id": "step_1",
                "op": "summary",
                "inputs": "sales",
                "output": "地区汇总",
                "parameters": {"group_by": "地区", "value_column": "销售额", "aggregation": "sum"},
            }
        ],
    }
    plan = validate_plan(normalize_plan_envelope(raw), catalog)
    assert plan.executable
    assert plan.steps[0].params["aggregations"] == {"销售额": "sum"}


def test_delivery_reopens_and_verifies_xlsx(tmp_path: Path) -> None:
    tables = {
        "订单": pd.DataFrame(
            {"订单": ["A", "B"], "金额": [100, 250.5], "说明": ["=SAFE", None]}
        )
    }
    path = tmp_path / "delivery.xlsx"
    export_tables(tables, path, include_log=False)
    report = verify_delivery(path, tables)
    assert report.status == "passed"
    assert report.checks_passed == report.checks_total
    assert len(report.artifact_sha256) == 64


def test_fidelity_clone_keeps_features_and_adds_result(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原表"
    sheet["A1"] = "标题"
    sheet["A2"] = "保留"
    sheet["B2"] = "=1+1"
    sheet.merge_cells("A4:B4")
    hidden = workbook.create_sheet("隐藏表")
    hidden.sheet_state = "hidden"
    workbook.save(source)
    target = tmp_path / "preserved.xlsx"
    preserve_workbook_export(source, target, {"处理结果": pd.DataFrame({"值": [1, 2]})})
    inventory = workbook_feature_inventory(target)
    assert "隐藏表" in inventory["hidden_sheets"]
    assert inventory["merged_ranges"] == 1
    checked = load_workbook(target, data_only=False)
    try:
        assert checked["原表"]["B2"].value == "=1+1"
        assert checked["处理结果"]["A3"].value == 2
    finally:
        checked.close()


def test_task_repository_roundtrip_and_retention(tmp_path: Path) -> None:
    repository = TaskRepository(tmp_path / "tasks", retention_days=3)
    repository.create("20260823-ABCD", "测试任务")
    frame = pd.DataFrame({"编号": ["001"], "金额": [12.5]})
    repository.save(
        "20260823-ABCD",
        task_name="测试任务",
        tables={"abc123": ("订单", frame, "测试", True)},
        active_table="abc123",
        operations=[],
        file_names=["source.xlsx"],
        import_warnings=[],
    )
    restored = repository.load("20260823-ABCD")
    pd.testing.assert_frame_equal(restored["loaded_tables"]["abc123"][1], frame)
    removed = repository.purge_expired(now=datetime.now() + timedelta(days=4))
    assert removed == ["20260823-ABCD"]


def test_read_only_sqlite_and_sql_guard(tmp_path: Path) -> None:
    database = tmp_path / "sales.db"
    connection = sqlite3.connect(database)
    connection.execute("CREATE TABLE sales(region TEXT, amount REAL)")
    connection.executemany("INSERT INTO sales VALUES (?,?)", [("华东", 10), ("华南", 20)])
    connection.commit()
    connection.close()
    result = query_sqlite_read_only(database, "SELECT region, amount FROM sales ORDER BY amount")
    assert result["amount"].tolist() == [10.0, 20.0]
    assert validate_read_only_sql("WITH x AS (SELECT 1 a) SELECT * FROM x")
    with pytest.raises(ValueError):
        validate_read_only_sql("DELETE FROM sales")
    with pytest.raises(ValueError):
        validate_read_only_sql("SELECT * FROM read_csv_auto('secret.csv')")


def test_vba_bundle_is_scanned_and_auditable(tmp_path: Path) -> None:
    code = "Option Explicit\nPublic Sub FormatReport()\n  Worksheets(1).Range(\"A1\").Font.Bold = True\nEnd Sub"
    assert validate_vba_module(code) == code
    destination = tmp_path / "vba.zip"
    result = build_vba_bundle(code, destination)
    assert result.status == "ready" and destination.exists()
    with pytest.raises(ValueError):
        validate_vba_module("Sub Bad(): Shell \"cmd.exe\": End Sub")


def test_order_quote_and_scheduler(tmp_path: Path) -> None:
    quote = quote_order(
        "合并三张订单表，复杂对账并制作经营看板，每周自动运行",
        table_count=3,
        total_rows=500_000,
        has_sample=True,
    )
    assert quote.capability in {"supported", "needs_review"}
    assert quote.suggested_price[0] > 0
    ran: list[dict[str, object]] = []
    scheduler = LocalScheduler(tmp_path / "scheduler.db")
    scheduler.register_job("test", lambda payload: ran.append(dict(payload)))
    schedule = scheduler.add("测试", "interval_minutes", "1", "test", {"value": 1})
    count = scheduler.run_due(datetime.fromisoformat(schedule.next_run) + timedelta(seconds=1))
    assert count == 1 and ran == [{"value": 1}]


def test_duckdb_queries_real_csv(tmp_path: Path) -> None:
    pytest.importorskip("duckdb")
    source = tmp_path / "sales.csv"
    pd.DataFrame(
        {"region": ["华东", "华东", "华南"], "sales": [10, 20, 7]}
    ).to_csv(source, index=False)
    result = query_files(
        [source],
        "SELECT region, SUM(sales) total FROM input_1 GROUP BY region ORDER BY total DESC",
    )
    assert result.to_dict("records") == [
        {"region": "华东", "total": 30.0},
        {"region": "华南", "total": 7.0},
    ]


@pytest.mark.skipif(not document_capabilities().get("image_ocr"), reason="OCR engine unavailable")
def test_real_chinese_ocr_asset() -> None:
    image = Path(__file__).resolve().parents[1] / "assets" / "xianyu" / "01_主封面_v2_数据分析.png"
    result = extract_image_text(image)
    assert not result.empty
    assert {"conf", "text"}.issubset(result.columns)


@pytest.mark.skipif(sys.platform != "win32", reason="Windows DPAPI only")
def test_dpapi_secret_roundtrip(tmp_path: Path) -> None:
    store = SecureSecretStore(tmp_path / "secrets.dpapi")
    store.set("test", "not-a-real-secret")
    assert store.get("test") == "not-a-real-secret"
    store.delete("test")
    assert store.get("test") == ""
