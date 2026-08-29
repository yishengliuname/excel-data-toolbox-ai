from __future__ import annotations

import tempfile
from pathlib import Path
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

import excel_data_toolbox.server as server_module
from excel_data_toolbox.adaptive_report import build_adaptive_analysis_report, infer_column_roles
from excel_data_toolbox.core import export_tables
from excel_data_toolbox.server import AppSession, ToolboxHandler


def _unfamiliar_tables() -> list[pd.DataFrame]:
    january = pd.DataFrame({
        "业务流水号": ["T001", "T002", "T003"],
        "发生日期": ["2026-01-03", "2026-01-11", "2026-01-25"],
        "服务类型": ["咨询", "实施", "咨询"],
        "运营区域": ["华东", "华南", "华东"],
        "触达渠道": ["线上", "转介绍", "线下"],
        "合同金额": [12000, 36000, 18000],
        "交付成本": [5000, 21000, 7000],
        "评价得分": [4.8, 4.2, 2.1],
    })
    february = pd.DataFrame({
        "业务流水号": ["T003", "T004", "T005"],
        "发生日期": ["2026-01-25", "2026-02-08", "2026-02-19"],
        "服务类型": ["咨询", "运维", "实施"],
        "运营区域": ["华东", "华北", "华南"],
        "触达渠道": ["线下", "线上", "转介绍"],
        "合同金额": [18000, 9000, 180000],
        "交付成本": [7000, 4500, 52000],
        "评价得分": [2.1, 4.9, 4.5],
    })
    region_owner = pd.DataFrame({
        "运营区域": ["华东", "华南", "华北"],
        "负责人": ["甲", "乙", "丙"],
        "预算额度": [100000, 160000, 80000],
    })
    notes = pd.DataFrame({"口径说明": ["合同金额为含税口径，异常值需要业务复核。"]})
    return [january, february, region_owner, notes]


class AdaptiveReportTests(unittest.TestCase):
    def test_roles_merge_metrics_trend_relations_and_anomalies(self) -> None:
        tables = _unfamiliar_tables()
        result = build_adaptive_analysis_report(
            tables,
            source_names=["一月业务", "二月业务", "区域负责人", "口径说明"],
            user_request="全面分析整体表现、趋势、排名、异常并生成老板能直接看的 Excel 报表",
        )
        self.assertEqual(len(result.outputs), 9)
        self.assertEqual(result.report["combined_table_count"], 2)
        self.assertEqual(result.report["primary_row_count"], 5)
        self.assertGreaterEqual(result.report["metric_count"], 3)
        self.assertGreaterEqual(result.report["dimension_count"], 2)
        self.assertFalse(result.outputs["分类排名"].empty)
        self.assertFalse(result.outputs["时间趋势"].empty)
        self.assertFalse(result.outputs["表关系建议"].empty)
        self.assertGreaterEqual(result.report["anomaly_count"], 1)
        roles = infer_column_roles(tables[0])
        self.assertEqual(roles["发生日期"]["role"], "日期")
        self.assertEqual(roles["业务流水号"]["role"], "标识符")

    def test_unknown_non_date_structure_still_produces_auditable_report(self) -> None:
        frame = pd.DataFrame({
            "学员编号": ["S1", "S2", "S3", "S4"],
            "班组": ["A", "A", "B", "B"],
            "测评成绩": [86, 93, 55, 78],
            "出勤率": [0.95, 0.98, 0.72, 0.88],
            "评语": ["稳定", "优秀", "需关注", "正常"],
        })
        result = build_adaptive_analysis_report(
            [frame], source_names=["训练营记录"], user_request="分析排名和需要关注的问题，输出管理报表"
        )
        self.assertEqual(result.report["primary_row_count"], 4)
        self.assertTrue(result.outputs["时间趋势"].empty)
        self.assertFalse(result.outputs["分类排名"].empty)
        self.assertIn("推断依据", result.outputs["数据字典"].columns)

    def test_secondary_fact_metrics_appear_in_management_overview(self) -> None:
        orders = pd.DataFrame({
            "订单日期": ["2026-01-01", "2026-01-02"], "订单号": ["O1", "O2"],
            "渠道": ["线上", "门店"], "销售额": [100, 200],
        })
        refunds = pd.DataFrame({
            "退款日期": ["2026-02-01", "2026-02-02"], "退款单号": ["R1", "R2"],
            "原因": ["质量", "时效"], "退款金额": [20, 30],
        })
        result = build_adaptive_analysis_report(
            [orders, refunds], source_names=["订单事实", "退款事实"],
            user_request="关联订单和退款分析经营情况",
        )
        overview = result.outputs["管理层通用总览"]
        assert result.report["fact_count"] == 2
        assert overview["指标"].astype(str).str.contains("事实域：退款事实.退款金额", regex=False).any()

    def test_native_xlsx_export_contains_nine_sheets_and_charts(self) -> None:
        result = build_adaptive_analysis_report(
            _unfamiliar_tables(),
            source_names=["一月业务", "二月业务", "区域负责人", "口径说明"],
            user_request="全面分析并输出经营看板",
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "adaptive.xlsx"
            export_tables(result.outputs, path, include_log=False)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, list(result.outputs))
                self.assertGreaterEqual(len(workbook["自适应图表看板"]._charts), 3)
                self.assertEqual(workbook["管理层通用总览"]["A1"].value, "AI 通用经营分析驾驶舱")
            finally:
                workbook.close()

    def test_unified_command_uses_local_fallback_without_deepseek(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            names = ["一月业务", "二月业务", "区域负责人", "口径说明"]
            ids = [session.add_table(name, frame, source="单元测试", original=True) for name, frame in zip(names, _unfamiliar_tables())]
            prompt = "请全面分析这些陌生数据，找出指标、趋势、排名、异常和数据质量问题，生成老板能直接看的 Excel 经营看板。"
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(server_module, "_project_ai_config", return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"}),
                patch.object(server_module.DeepSeekClient, "classify_unified_request", side_effect=AssertionError("通用本地流程不应访问 DeepSeek")),
            ):
                response = handler._ai_unified({"prompt": prompt, "table_ids": ids})
            self.assertEqual(response["mode"], "data")
            self.assertTrue(response["auto_execute"])
            self.assertEqual(response["plan"]["steps"][0]["operation"], "adaptive_analysis_report")
        finally:
            session.close()


if __name__ == "__main__":
    unittest.main()
