from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import excel_data_toolbox.server as server_module
import pandas as pd
from excel_data_toolbox.core import export_tables
from excel_data_toolbox.enterprise_report import build_enterprise_diagnosis_report
from excel_data_toolbox.server import AppSession, ToolboxHandler
from openpyxl import load_workbook


def _tables() -> list[pd.DataFrame]:
    transactions = pd.DataFrame(
        {
            "日期": ["2026-01-05", "2026-05-03"],
            "单据号": ["X1", "X2"],
            "业务类型": ["销售", "销售"],
            "部门": ["软件部", "数据部"],
            "客户/供应商": ["A科技", "G公司"],
            "收入": [50_000, 200_000],
            "成本": [18_000, 190_000],
            "付款状态": ["已回款", "未回款"],
            "负责人": ["张伟", "王强"],
        }
    )
    customers = pd.DataFrame(
        {
            "客户": ["A科技", "G公司"],
            "行业": ["制造业", "互联网"],
            "合作年限": [3, "半年"],
            "满意度": [5, 4],
            "信用等级": ["A级", "C级"],
            "回款风险": ["低", "高"],
        }
    )
    performance = pd.DataFrame(
        {
            "负责人": ["张伟", "王强"],
            "销售额": [50_000, 200_000],
            "目标完成率": [1.2, 1.5],
            "客户评分": [4.8, 4.1],
            "投诉次数": [0, 1],
        }
    )
    expenses = pd.DataFrame(
        {"月份": ["2026-01", "2026-05"], "费用类别": ["人员成本", "营销"], "金额": [30_000, 120_000]}
    )
    inventory = pd.DataFrame(
        {
            "产品": ["产品C", "产品D"],
            "库存数量": [300, 20],
            "库存金额": [150_000, 5_000],
            "月销量": [5, 60],
            "状态": ["积压", "缺货风险"],
        }
    )
    notes = pd.DataFrame({"老板需求": ["全面诊断增长与利润"]})
    return [transactions, customers, performance, expenses, inventory, notes]


class EnterpriseDiagnosisTests(unittest.TestCase):
    def test_calculates_cross_domain_diagnosis(self) -> None:
        result = build_enterprise_diagnosis_report(
            _tables(),
            source_names=["流水", "客户", "绩效", "费用", "库存", "老板原话"],
            user_request="全面分析客户、销售、成本和库存风险并给出下一步动作",
        )
        self.assertEqual(result.report["net_revenue"], 250_000)
        self.assertEqual(result.report["gross_profit"], 42_000)
        self.assertEqual(result.report["estimated_operating_profit"], -108_000)
        self.assertEqual(result.report["collection_risk_exposure"], 200_000)
        self.assertEqual(len(result.outputs), 10)
        self.assertEqual(result.outputs["客户与回款风险"].iloc[0]["客户"], "G公司")
        self.assertEqual(result.outputs["库存风险分析"].set_index("产品").loc["产品C", "诊断状态"], "偏高线索")

    def test_export_contains_ten_sheets_and_four_native_charts(self) -> None:
        result = build_enterprise_diagnosis_report(
            _tables(),
            source_names=["流水", "客户", "绩效", "费用", "库存", "老板原话"],
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "enterprise.xlsx"
            export_tables(result.outputs, path, include_log=False)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, list(result.outputs))
                self.assertEqual(len(workbook["经营诊断看板"]._charts), 4)
                self.assertEqual(workbook["管理层诊断总览"]["A1"].value, "企业集团经营诊断驾驶舱")
            finally:
                workbook.close()

    def test_manufacturing_multifact_model_handles_unmatched_master_data_and_refunds(self) -> None:
        transactions = pd.DataFrame(
            {
                "日期": ["2026-01-05", "2026-01-20", "2026-02-03"],
                "订单号": ["M1", "M2", "M3"],
                "客户": ["海外客户A", "海外客户A", "临时客户"],
                "产品": ["工业软件", "工业软件", "传感器"],
                "销售人员": ["张伟", "张伟", "李明"],
                "销售金额": [1_000_000, 120_000, 100_000],
                "订单状态": ["未回款", "退款", "已回款"],
            }
        )
        customers = pd.DataFrame(
            {
                "客户名称": ["海外客户A"],
                "信用等级": ["C"],
                "满意度": [3],
                "回款风险": ["高"],
            }
        )
        performance = pd.DataFrame(
            {
                "销售人员": ["张伟", "李明"],
                "销售额": [1_120_000, 100_000],
                "毛利": [280_000, 20_000],
                "回款金额": [0, 100_000],
                "客户评分": [3.0, 3.5],
                "投诉次数": [2, 3],
            }
        )
        expenses = pd.DataFrame(
            {"月份": ["2026-01", "2026-02"], "费用类别": ["市场", "管理"], "费用金额": [250_000, 150_000]}
        )
        inventory = pd.DataFrame(
            {
                "产品": ["工业软件授权", "传感器"],
                "库存数量": [100, 10],
                "库存金额": [500_000, 20_000],
                "月销量": [5, 20],
            }
        )
        production = pd.DataFrame(
            {
                "月份": ["2026-01", "2026-02"],
                "产品": ["工业软件授权", "传感器"],
                "材料成本": [300_000, 50_000],
                "人工成本": [100_000, 20_000],
                "制造费用": [50_000, 10_000],
                "总成本": [450_000, 80_000],
                "产量": [10, 100],
            }
        )
        result = build_enterprise_diagnosis_report(
            [transactions, production, customers, inventory, performance, expenses],
            source_names=["订单销售数据", "生产成本数据", "客户经营画像", "库存数据", "人员绩效", "费用支出"],
        )
        self.assertEqual(result.report["gross_profit"], 300_000)
        self.assertEqual(result.report["operating_expense"], 400_000)
        self.assertEqual(result.report["estimated_operating_profit"], -100_000)
        self.assertEqual(result.report["management_revenue_excluding_refunds"], 1_100_000)
        self.assertEqual(result.report["production_cost"], 530_000)
        customer_risk = result.outputs["客户与回款风险"].set_index("客户")
        self.assertIn("客户主数据未匹配", customer_risk.loc["临时客户", "主要风险"])
        self.assertEqual(customer_risk.loc["临时客户", "综合风险"], "未知/待核验")
        self.assertTrue(result.outputs["销售团队诊断"]["订单级成本"].isna().all())
        self.assertTrue(result.outputs["风险行动计划"]["风险事项"].eq("销售质量与客户体验").any())
        audits = result.outputs["数据口径与验收"]
        self.assertTrue(audits["审计项"].eq("退款处理").any())
        self.assertTrue(audits["审计项"].astype(str).str.contains("工业软件↔工业软件授权", regex=False).any())

    def test_alternate_business_vocabulary_is_portable(self) -> None:
        transactions = pd.DataFrame(
            {
                "交易日期": ["2026-01-01", "2026-02-01"],
                "订单编号": ["A1", "A2"],
                "客户单位": ["甲方", "乙方"],
                "商品名称": ["咨询服务", "培训服务"],
                "成交金额": [80_000, 120_000],
                "销售员": ["顾问甲", "顾问乙"],
                "结算状态": ["已回款", "未回款"],
            }
        )
        customers = pd.DataFrame(
            {
                "客户单位": ["甲方", "乙方"],
                "客户满意度": [5, 4],
                "客户信用": ["A", "B"],
                "风险等级": ["低", "中"],
            }
        )
        performance = pd.DataFrame(
            {
                "员工姓名": ["顾问甲", "顾问乙"],
                "业绩金额": [80_000, 120_000],
                "销售毛利": [40_000, 45_000],
                "实收金额": [80_000, 0],
                "绩效评分": [4.8, 4.0],
                "投诉": [0, 1],
            }
        )
        expenses = pd.DataFrame(
            {
                "会计期间": ["2026-01", "2026-02"],
                "费用科目": ["人员", "市场"],
                "发生额": [30_000, 20_000],
            }
        )
        inventory = pd.DataFrame(
            {
                "物料名称": ["培训教材", "咨询工时包"],
                "当前库存": [50, 20],
                "存货金额": [5_000, 10_000],
                "近30天出库": [25, 5],
            }
        )
        result = build_enterprise_diagnosis_report(
            [transactions, customers, performance, expenses, inventory],
            source_names=["交易", "客户", "顾问绩效", "费用", "资源库存"],
        )
        self.assertEqual(result.report["net_revenue"], 200_000)
        self.assertEqual(result.report["gross_profit"], 85_000)
        self.assertEqual(result.outputs["客户与回款风险"].set_index("客户").loc["乙方", "综合风险"], "高")
        self.assertEqual(len(result.outputs), 10)

    def test_unified_command_bypasses_deepseek_and_engineering_route(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            ids = [
                session.add_table(name, frame, source="单元测试", original=True)
                for name, frame in zip(
                    ["流水", "客户", "绩效", "费用", "库存", "老板原话"],
                    _tables(),
                    strict=True,
                )
            ]
            prompt = "最近公司增长很快但是利润没有提升，全面分析客户、销售、成本和库存风险，告诉我下一步怎么做。"
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(
                    server_module,
                    "_project_ai_config",
                    return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"},
                ),
                patch.object(
                    server_module.DeepSeekClient,
                    "classify_unified_request",
                    side_effect=AssertionError("不应访问DeepSeek"),
                ),
            ):
                response = handler._ai_unified({"prompt": prompt, "table_ids": ids})
            self.assertEqual(response["mode"], "data")
            self.assertTrue(response["auto_execute"])
            self.assertEqual(response["plan"]["steps"][0]["operation"], "enterprise_diagnosis_report")
        finally:
            session.close()


if __name__ == "__main__":
    unittest.main()
