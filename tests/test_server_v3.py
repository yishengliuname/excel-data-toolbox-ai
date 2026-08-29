from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

import excel_data_toolbox.server as server_module
from excel_data_toolbox.server import (
    AppSession,
    TableEntry,
    ToolboxHandler,
    _audit_xlsx_structure,
    _long_text_detail_frame,
)


def test_xlsx_structure_audit_warns_without_exposing_cell_values(tmp_path: Path) -> None:
    path = tmp_path / "客户样例.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet["A1"] = "高度敏感的原始值"
    sheet["B2"] = "=1+1"
    sheet.merge_cells("C1:D1")
    hidden = workbook.create_sheet("内部")
    hidden.sheet_state = "hidden"
    workbook.save(path)

    warnings = _audit_xlsx_structure(path)

    combined = "\n".join(warnings)
    assert "公式" in combined
    assert "合并单元格" in combined
    assert "隐藏工作表" in combined
    assert "高度敏感的原始值" not in combined


def test_review_queue_tracks_decisions_without_operation_log_values() -> None:
    session = AppSession()
    try:
        identifiers = session.add_review_items(
            "模糊匹配",
            "单元测试",
            [
                {
                    "title": "客户名称待确认",
                    "record_key": "第3行",
                    "evidence": {"原值": "甲有限公司", "候选": "甲公司"},
                }
            ],
        )
        assert len(identifiers) == 1
        assert session.review_payload()["counts"]["pending"] == 1

        changed = session.decide_reviews(identifiers, "accepted", "人工核验通过")

        assert changed == 1
        payload = session.review_payload()
        assert payload["counts"]["accepted"] == 1
        assert payload["items"][0]["decision_note"] == "人工核验通过"
        assert len(session.operation_log()) == 0
    finally:
        session.close()


def test_session_can_redo_a_generated_table() -> None:
    session = AppSession()
    try:
        original_id = session.add_table(
            "原表", pd.DataFrame({"订单": ["001"]}), source="测试", original=True
        )
        result_id = session.add_table(
            "结果", pd.DataFrame({"订单": ["001"]}), source="测试"
        )
        session.record(
            "测试处理",
            "不含原始值",
            inputs=[session.tables[original_id].name],
            produced=[result_id],
            before_rows=1,
            after_rows=1,
        )

        session.undo()
        assert result_id not in session.tables
        assert session.state_payload()["can_redo"] is True

        session.redo()
        assert result_id in session.tables
        assert session.active_table == result_id
        assert session.state_payload()["can_redo"] is False
    finally:
        session.close()


def test_each_reset_creates_one_named_isolated_task_folder() -> None:
    session = AppSession()
    try:
        first_task_id = session.task_id
        first_task_dir = session.task_dir
        assert first_task_dir.name == first_task_id
        assert session.upload_dir.parent == first_task_dir
        assert session.output_dir.parent == first_task_dir

        session.add_table("旧任务表", pd.DataFrame({"值": [1]}), source="测试")
        session.reset()

        assert session.task_id != first_task_id
        assert session.task_dir.name == session.task_id
        # V9 keeps one durable folder per task so a restart can recover work.
        assert first_task_dir.exists()
        assert (first_task_dir / "manifest.json").exists()
        assert session.tables == {}
    finally:
        session.close()


def test_long_text_detail_expands_issue_items_without_changing_source() -> None:
    frame = pd.DataFrame(
        {"序号": [1], "问题说明": ["摘要偏少；目录错误；图中文字太小。"], "得分": [70]}
    )
    original = frame.copy(deep=True)

    detail = _long_text_detail_frame(
        [TableEntry("table-1", "成绩表", frame, "测试", True)]
    )

    assert detail["内容"].tolist() == ["摘要偏少；", "目录错误；", "图中文字太小。"]
    assert detail["记录标识"].tolist() == [1, 1, 1]
    pd.testing.assert_frame_equal(frame, original)


def test_recipe_http_handlers_persist_rules_only_and_support_dry_run(
    tmp_path: Path, monkeypatch,
) -> None:
    session = AppSession()
    monkeypatch.setattr(server_module, "SESSION", session)
    monkeypatch.setattr(server_module, "RECIPE_STORE_DIR", tmp_path / "recipes")
    handler = object.__new__(ToolboxHandler)
    try:
        source_id = session.add_table(
            "客户表",
            pd.DataFrame(
                {
                    "订单号": ["001", "001", "002"],
                    "备注": ["高度敏感内容", "高度敏感内容", "正常"],
                }
            ),
            source="测试",
            original=True,
        )
        saved = handler._recipe_save(
            {
                "name": "标准去重",
                "description": "测试方案",
                "steps": [
                    {
                        "operation": "drop_duplicates",
                        "params": {"subset": ["订单号"], "keep": "first"},
                    }
                ],
            }
        )
        recipe_id = saved["saved"]["id"]
        stored_text = (tmp_path / "recipes" / f"{recipe_id}.json").read_text(
            encoding="utf-8"
        )
        assert "高度敏感内容" not in stored_text
        assert saved["saved"]["steps"][0]["operation"] == "drop_duplicates"

        before_tables = len(session.tables)
        preview = handler._recipe_run(
            {
                "recipe_id": recipe_id,
                "table_id": source_id,
                "output_name": "去重结果",
                "dry_run": True,
            }
        )
        assert preview["before_rows"] == 3
        assert preview["after_rows"] == 2
        assert len(session.tables) == before_tables

        committed = handler._recipe_run(
            {
                "recipe_id": recipe_id,
                "table_id": source_id,
                "output_name": "去重结果",
                "dry_run": False,
            }
        )
        assert len(session.get(committed["table_id"]).frame) == 2
    finally:
        session.close()


def test_validation_and_reconciliation_handlers_create_evidence_tables(
    monkeypatch,
) -> None:
    session = AppSession()
    monkeypatch.setattr(server_module, "SESSION", session)
    handler = object.__new__(ToolboxHandler)
    try:
        left_id = session.add_table(
            "账面表",
            pd.DataFrame(
                {
                    "编号": ["A", "B", "C"],
                    "金额": [100, 200, 300],
                    "日期": ["2026-08-01", "2026-08-02", "2026-08-03"],
                }
            ),
            source="测试",
            original=True,
        )
        right_id = session.add_table(
            "流水表",
            pd.DataFrame(
                {
                    "编号": ["A", "B", "D"],
                    "金额": [100, 200.05, 400],
                    "日期": ["2026-08-01", "2026-08-02", "2026-08-04"],
                }
            ),
            source="测试",
            original=True,
        )

        validation = handler._validate(
            {
                "table_id": left_id,
                "output_name": "质量验收",
                "rules": [
                    {"type": "not_null", "column": "编号"},
                    {"type": "range", "column": "金额", "min": 0, "max": 250},
                ],
            }
        )
        assert validation["total_rules"] == 2
        assert validation["failed_count"] == 1
        assert set(validation["table_ids"]) == {"规则汇总", "失败明细"}

        reconciliation = handler._reconcile_advanced(
            {
                "left_id": left_id,
                "right_id": right_id,
                "output_name": "高级对账",
                "config": {
                    "left_keys": ["编号"],
                    "right_keys": ["编号"],
                    "amount": {
                        "left_column": "金额",
                        "right_column": "金额",
                        "tolerance": 0.1,
                    },
                    "date": {
                        "left_column": "日期",
                        "right_column": "日期",
                        "tolerance_days": 0,
                    },
                },
            }
        )
        assert reconciliation["matched_count"] == 1
        assert reconciliation["difference_count"] == 1
        assert reconciliation["left_only_count"] == 1
        assert reconciliation["right_only_count"] == 1
        assert len(reconciliation["table_ids"]) == 8
        assert session.review_payload()["counts"]["pending"] >= 2
    finally:
        session.close()


def test_columns_handler_sorts_by_renamed_column(monkeypatch) -> None:
    session = AppSession()
    monkeypatch.setattr(server_module, "SESSION", session)
    handler = object.__new__(ToolboxHandler)
    try:
        source_id = session.add_table(
            "订单表",
            pd.DataFrame({"订单号": ["A", "B"], "销售额": [100, 300]}),
            source="测试",
            original=True,
        )

        handler._columns(
            {
                "table": source_id,
                "columns": ["订单号", "销售额"],
                "rename_column": "销售额",
                "rename_value": "净销售额",
                "sort_column": "销售额",
                "ascending": False,
                "output_name": "字段整理结果",
            }
        )

        result = session.get(session.active_table or "").frame
        assert result.columns.tolist() == ["订单号", "净销售额"]
        assert result["净销售额"].tolist() == [300, 100]
    finally:
        session.close()
