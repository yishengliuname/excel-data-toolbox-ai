from __future__ import annotations

from pathlib import Path
import tempfile

from openpyxl import load_workbook
import pandas as pd

from excel_data_toolbox.core import export_tables
from excel_data_toolbox.source_guard import assess_prompt_data_alignment, detect_generated_workbook


def test_generated_workbook_metadata_survives_sheet_renaming() -> None:
    frame = pd.DataFrame({"指标": ["收入"], "结果": [100]})
    frame.attrs["toolbox_report_kind"] = "test_report"
    with tempfile.TemporaryDirectory() as directory:
        path = Path(directory) / "output.xlsx"
        export_tables({"任意结果": frame}, path, include_log=False)
        workbook = load_workbook(path)
        workbook[workbook.sheetnames[0]].title = "用户已经重命名"
        workbook.save(path)
        workbook.close()

        assessment = detect_generated_workbook(path)
        assert assessment.generated
        assert assessment.report_kind == "test_report"
        assert assessment.matched_sheets == ("工作簿生成元数据",)


def test_prompt_alignment_can_use_small_local_value_samples() -> None:
    frame = pd.DataFrame(
        {
            "主体": ["中心门店", "大学城门店"],
            "品项": ["招牌菜品", "套餐菜品"],
            "渠道": ["堂食", "美团外卖"],
            "金额": [100, 200],
        }
    )
    assessment = assess_prompt_data_alignment(
        "分析餐饮门店的菜品、食材损耗和外卖经营情况",
        [frame],
        ["业务数据"],
    )
    assert assessment.prompt_domain == "餐饮门店"
    assert assessment.data_domain == "餐饮门店"
    assert assessment.aligned
