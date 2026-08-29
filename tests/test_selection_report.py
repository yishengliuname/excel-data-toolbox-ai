from __future__ import annotations

import tempfile
from pathlib import Path
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

import excel_data_toolbox.server as server_module
from excel_data_toolbox.core import export_tables
from excel_data_toolbox.selection_report import (
    build_selection_recommendation_report,
    infer_selection_columns,
    parse_selection_count,
)
from excel_data_toolbox.server import AppSession, ToolboxHandler


def _candidate_frame() -> pd.DataFrame:
    return pd.DataFrame({
        "序号": [1, 2, 3, 4, 5],
        "第1轮得分": [92, 86, 90, 78, "?"],
        "第1轮问题": ["结构完整", "主要结果合理", "完成度较高", "内容偏少", "无有效提交"],
        "第2轮得分": [90, 88, 91, 80, None],
        "第2轮评语": ["质量较好", "与3组内容重复", "结果有误", "近期明显提升", "不合格"],
        "第3轮得分": [93, 89, 92, 90, None],
        "第3轮备注": ["表现稳定", "需复核", "已修正", "主要结果合理", ""],
    })


class SelectionReportTests(unittest.TestCase):
    def test_infers_columns_count_and_builds_structured_outputs(self) -> None:
        frame = _candidate_frame()
        inferred = infer_selection_columns(frame)
        self.assertEqual(inferred["identifier_column"], "序号")
        self.assertEqual(inferred["score_columns"], ["第1轮得分", "第2轮得分", "第3轮得分"])
        self.assertEqual(parse_selection_count("从这些序号里挑出前三名参加比赛"), 3)
        self.assertEqual(
            parse_selection_count("按照需求和每一个组的数据，从这些组中选取最优秀的八个组参加比赛"),
            8,
        )
        self.assertEqual(parse_selection_count("从一百二十组中筛选综合最好的十二组"), 12)

        result = build_selection_recommendation_report(
            [frame], source_names=["评审记录"], user_request="选出3个参加比赛", top_n=3,
        )
        self.assertEqual(list(result.outputs), [
            "评选管理总览", "建议入选名单", "全部候选排序",
            "风险复核清单", "评选规则与字段", "评选图表看板",
        ])
        self.assertEqual(result.report["selected_count"], 3)
        self.assertEqual(result.report["selected_ids"][0], "1")
        self.assertNotIn("5", result.report["selected_ids"])
        self.assertIn("风险扣分", result.outputs["全部候选排序"].columns)
        self.assertIn("推荐理由", result.outputs["建议入选名单"].columns)

    def test_export_contains_six_sheets_and_native_charts(self) -> None:
        result = build_selection_recommendation_report(
            [_candidate_frame()], source_names=["评审记录"], user_request="选出3个参加比赛", top_n=3,
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "selection.xlsx"
            export_tables(result.outputs, path, include_log=False)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, list(result.outputs))
                self.assertEqual(len(workbook["评选图表看板"]._charts), 2)
                self.assertEqual(workbook["评选管理总览"]["A1"].value, "候选对象结构化评选驾驶舱")
            finally:
                workbook.close()

    def test_ai_can_omit_visualization_when_it_has_no_decision_value(self) -> None:
        result = build_selection_recommendation_report(
            [_candidate_frame()],
            source_names=["评审记录"],
            user_request="只给出入选名单和复核依据",
            top_n=3,
            include_charts=False,
        )
        self.assertNotIn("评选图表看板", result.outputs)
        self.assertEqual(result.report["chart_count"], 0)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "selection-no-charts.xlsx"
            export_tables(result.outputs, path, include_log=False)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertNotIn("评选图表看板", workbook.sheetnames)
                self.assertEqual(sum(len(sheet._charts) for sheet in workbook.worksheets), 0)
            finally:
                workbook.close()

    def test_unified_command_bypasses_deepseek_and_engineering_route(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            table_id = session.add_table("评审记录", _candidate_frame(), source="单元测试", original=True)
            prompt = "需要从这些序号中选3个去参加比赛，你觉得选哪些合适？"
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(server_module, "_project_ai_config", return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"}),
                patch.object(server_module.DeepSeekClient, "classify_unified_request", side_effect=AssertionError("评选本地流程不应访问DeepSeek")),
            ):
                response = handler._ai_unified({"prompt": prompt, "table_ids": [table_id]})
            self.assertEqual(response["mode"], "data")
            self.assertTrue(response["auto_execute"])
            self.assertEqual(response["plan"]["steps"][0]["operation"], "selection_recommendation_report")
            self.assertEqual(response["plan"]["steps"][0]["params"]["top_n"], 3)
        finally:
            session.close()

    def test_natural_selection_wording_uses_structured_local_route(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            table_id = session.add_table("评审记录", _candidate_frame(), source="单元测试", original=True)
            prompt = "按照需求和每一个组的数据，从这些组中选取最优秀的八个组参加比赛"
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(server_module, "_project_ai_config", return_value={"configured": False, "api_key": "", "model": "deepseek-v4-flash"}),
                patch.object(server_module.DeepSeekClient, "classify_unified_request", side_effect=AssertionError("评选本地流程不应访问DeepSeek")),
            ):
                response = handler._ai_unified({"prompt": prompt, "table_ids": [table_id]})
            self.assertEqual(response["mode"], "data")
            self.assertEqual(response["plan"]["steps"][0]["operation"], "selection_recommendation_report")
            self.assertEqual(response["plan"]["steps"][0]["params"]["top_n"], 8)
        finally:
            session.close()

    def test_deepseek_stage_one_normalisation_reenters_local_specialist(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            table_id = session.add_table("评审记录", _candidate_frame(), source="单元测试", original=True)
            colloquial = "请根据各组材料，替公司安排八支代表队出战"
            normalized = "根据各组材料，为公司安排八支代表队出战"
            route = {
                "intent": "data",
                "normalized_request": normalized,
                "data_request": normalized,
                "chart_request": "",
                "engineering_category": None,
                "business_action": "select_candidates",
                "target_count": 8,
                "business_subject": "候选组",
                "interpretation_confidence": "high",
                "visualization_need": "recommended",
                "visualization_reason": "多组综合得分和风险对比使用图表更直观",
                "reason": "已理解为候选组评选",
            }
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(server_module, "_project_ai_config", return_value={"configured": True, "api_key": "test-key", "model": "deepseek-v4-flash"}),
                patch.object(server_module.DeepSeekClient, "classify_unified_request", return_value=route),
                patch.object(server_module.DeepSeekClient, "create_plan", side_effect=AssertionError("专业模块已匹配，不应进入通用计划生成")),
            ):
                response = handler._ai_unified({"prompt": colloquial, "table_ids": [table_id]})
            self.assertEqual(response["mode"], "data")
            self.assertEqual(response["plan"]["steps"][0]["operation"], "selection_recommendation_report")
            self.assertEqual(response["plan"]["steps"][0]["params"]["top_n"], 8)
            self.assertIn("stage_1", response["ai_pipeline"])
            self.assertIn("DeepSeek 第一阶段", response["route"]["reason"])
            self.assertEqual(response["ai_pipeline"]["visualization_decision"], "recommended")
        finally:
            session.close()

    def test_ai_visualization_judgement_controls_local_selection_output(self) -> None:
        session = AppSession()
        handler = object.__new__(ToolboxHandler)
        try:
            table_id = session.add_table("评审记录", _candidate_frame(), source="单元测试", original=True)
            visual_route = {
                "visualization_need": "not_needed",
                "visualization_reason": "本次只需名单与复核依据，表格表达更直接",
            }
            with (
                patch.object(server_module, "SESSION", session),
                patch.object(server_module, "_project_ai_config", return_value={"configured": True, "api_key": "test-key", "model": "deepseek-v4-flash"}),
                patch.object(server_module.DeepSeekClient, "classify_unified_request", return_value=visual_route),
            ):
                response = handler._ai_unified({
                    "prompt": "按照每一个组的数据选取最优秀的三个组参加比赛",
                    "table_ids": [table_id],
                })
            self.assertFalse(response["plan"]["steps"][0]["params"]["include_charts"])
            self.assertEqual(response["ai_pipeline"]["visualization_decision"], "not_needed")
            self.assertNotIn("评选图表看板", response["dry_run"]["tables"])
            with patch.object(server_module, "SESSION", session):
                executed = handler._ai_execute({
                    "plan_token": response["plan_token"],
                    "confirmed": True,
                })
            self.assertIn(executed["status"], {"completed", "needs_review"})
            destination = session.output_dir / "候选对象结构化评选报告.xlsx"
            self.assertTrue(destination.exists())
            workbook = load_workbook(destination, read_only=False, data_only=False)
            try:
                self.assertNotIn("评选图表看板", workbook.sheetnames)
                self.assertEqual(sum(len(sheet._charts) for sheet in workbook.worksheets), 0)
            finally:
                workbook.close()
        finally:
            session.close()


if __name__ == "__main__":
    unittest.main()
